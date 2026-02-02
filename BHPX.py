from __future__ import annotations

import csv
import calendar
import hashlib
import json
import os
import random
import sqlite3
import threading
import traceback
import webbrowser
import sys
import time
import tkinter.font as tkfont
import re
from http.cookiejar import CookieJar
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from html.parser import HTMLParser
from pathlib import Path
from queue import Empty, Queue
from urllib.request import Request, urlopen
from urllib.error import URLError, HTTPError
from urllib.parse import quote, urlencode
from tkinter import BOTH, END, LEFT, RIGHT, BooleanVar, Button, Frame, IntVar, Menu, Spinbox, StringVar, Tk, Toplevel
from tkinter import filedialog, messagebox
from tkinter import simpledialog
from tkinter.scrolledtext import ScrolledText
from tkinter import ttk

APP_NAME = "BHPX"
CARRIER_BHP = "BHP"
DEFAULT_RESET_PASSWORD = "bhpx-admin"
UI_SCALE = 1.3


def _is_writable_dir(path: Path) -> bool:
    try:
        path.mkdir(parents=True, exist_ok=True)
        probe = path / ".write_test"
        probe.write_text("ok", encoding="utf-8")
        probe.unlink(missing_ok=True)
        return True
    except Exception:
        return False


def _resolve_app_base_dir() -> Path:
    # When packaged with PyInstaller --onefile, __file__ points to a TEMP extraction folder (_MEI...).
    # Persist data in a stable location instead of TEMP.
    if getattr(sys, "frozen", False):
        exe_dir = Path(sys.executable).resolve().parent
        # Prefer portable mode: keep `data/` next to the EXE.
        if _is_writable_dir(exe_dir):
            return exe_dir

        appdata = os.environ.get("APPDATA") or os.environ.get("LOCALAPPDATA")
        if appdata:
            return Path(appdata) / APP_NAME
        return Path.home() / f".{APP_NAME.lower()}"

    return Path(__file__).resolve().parent


ROOT_DIR = _resolve_app_base_dir()
DATA_DIR = ROOT_DIR / "data"
LOGS_DIR = ROOT_DIR / "logs"
DATA_DIR.mkdir(parents=True, exist_ok=True)
LOGS_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH = DATA_DIR / "bhpx.sqlite3"
SETTINGS_PATH = DATA_DIR / "settings.json"
LOCK_PATH = DATA_DIR / "bhpx.lock"
DUPLICATE_PAYMENTS_LOG_PATH = LOGS_DIR / "duple_uplate.log"

_APP_LOCK_FH = None


def resource_path(relative_path: str) -> str:
    """
    Resolve a resource path for both dev runs and PyInstaller --onefile mode.
    When frozen, resources are extracted under sys._MEIPASS.
    """
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    return str(base / relative_path)


def apply_window_icon(win) -> None:
    try:
        win.iconbitmap(resource_path("bhpico.ico"))
    except Exception:
        pass


def now_utc_iso() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat(timespec="seconds") + "Z"


def normalize_tracking(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip().upper()
    if not text:
        return None
    return "".join(text.split())


def parse_date_to_iso(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    if not text:
        return None

    for transform in (lambda s: s, lambda s: s.split("T")[0], lambda s: s.split(" ")[0]):
        candidate = transform(text)
        try:
            if len(candidate) == 10 and "-" in candidate:
                return date.fromisoformat(candidate).isoformat()
        except ValueError:
            pass

    formats = [
        "%d.%m.%Y",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y.%m.%d",
        "%Y/%m/%d",
        "%Y-%m-%d",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue

    return text


_STATUS_DATE_RE = re.compile(r"(\\d{1,2})[./-](\\d{1,2})[./-](\\d{2,4})")
_TRACKING_RE = re.compile(r"^[A-Z0-9]+$")


def infer_delivery_date_from_status(status_text: str | None) -> str | None:
    if not status_text:
        return None
    m = _STATUS_DATE_RE.search(status_text)
    if not m:
        return None
    dd, mm, yy = m.group(1), m.group(2), m.group(3)
    try:
        d = int(dd)
        mth = int(mm)
        y = int(yy)
        if y < 100:
            y = 2000 + y
        return date(y, mth, d).isoformat()
    except Exception:
        return None


_BCS_FOLD_MAP = str.maketrans(
    {
        "č": "c",
        "ć": "c",
        "š": "s",
        "đ": "d",
        "ž": "z",
        "Č": "c",
        "Ć": "c",
        "Š": "s",
        "Đ": "d",
        "Ž": "z",
    }
)


def _fold_bcs(text: str) -> str:
    return " ".join(text.translate(_BCS_FOLD_MAP).lower().split())


def should_auto_remove_from_claims_by_status(status_text: str | None) -> bool:
    """
    Auto move to UNDELIVERED_REMOVED for statuses like:
      "Pošiljka uručena/vraćena pošiljaocu"
    """
    if not status_text:
        return False
    t = _fold_bcs(status_text)
    return ("urucena" in t) and ("vracena" in t) and ("posiljaoc" in t)


def should_auto_remove_from_claims_by_tracking_events(events: list[dict[str, str]] | None) -> bool:
    if not events:
        return False
    # Use the latest event (most recent). WebForms returns rows in chronological order.
    last = events[-1] if events else {}
    text = _fold_bcs(f"{last.get('event','')} {last.get('note','')}")
    return ("urucena" in text) and ("vracena" in text) and ("posiljaoc" in text)


def should_auto_remove_from_claims_by_recipient(recipient_name: str | None) -> bool:
    if not recipient_name:
        return False
    return "luxehair" in " ".join(str(recipient_name).lower().split())


def auto_remove_reason_note(status_text: str | None, recipient_name: str | None) -> tuple[str, str]:
    # Prefer explicit delivery/return status when present.
    if should_auto_remove_from_claims_by_status(status_text):
        return "Vraćeno pošiljaocu", "auto-status"
    if should_auto_remove_from_claims_by_recipient(recipient_name):
        return "Luxehair", "auto-recipient"
    return "Ostalo", "auto"


def parse_bhp_uplate_row(
    cols: list[str], *, override_paid_date_iso: str | None = None
) -> tuple[str | None, str | None, int | None, str | None, str]:
    """
    Returns (tracking, payer_name, amount_minor, paid_date_iso, error_key).
    error_key is "" when parsing succeeded.

    Deterministic parsing to support both fixed-layout CSVs and rows where extra ';' splits the payer name/description.
    """
    if len(cols) < 4:
        return None, None, None, None, "parse_error"

    # Tracking: choose the rightmost token that looks like a tracking id.
    tracking = None
    tracking_idx = None
    for i in range(len(cols) - 1, -1, -1):
        cand = normalize_tracking(cols[i])
        if not cand:
            continue
        if _TRACKING_RE.match(cand) and len(cand) >= 5:
            tracking = cand
            tracking_idx = i
            break
    if not tracking:
        return None, None, None, None, "missing_tracking"

    # Amount: choose the rightmost parseable amount before the tracking column (if possible).
    amount_minor = None
    amount_idx = None
    search_end = tracking_idx if tracking_idx is not None else len(cols)
    for i in range(search_end - 1, -1, -1):
        cand = cols[i]
        parsed = parse_payment_amount_to_minor(cand)
        if parsed is not None:
            amount_minor = parsed
            amount_idx = i
            break
    if amount_minor is None:
        return tracking, None, None, None, "invalid_amount"

    # Date: either overridden, or choose the rightmost parseable date before tracking.
    if override_paid_date_iso:
        paid_date_iso = override_paid_date_iso
    else:
        paid_date_iso = None
        for i in range(search_end - 1, -1, -1):
            # avoid accidentally taking amount field as date
            if amount_idx is not None and i == amount_idx:
                continue
            cand = parse_date_to_iso(cols[i])
            if parse_iso_date(cand) is not None:
                paid_date_iso = cand
                break
        if paid_date_iso is None:
            return tracking, None, amount_minor, None, "invalid_date"

    # Payer name: join columns between col3 (index 2) and amount column (exclusive) if present.
    payer_name = None
    if amount_idx is not None and amount_idx > 2:
        parts = [p.strip() for p in cols[2:amount_idx] if p is not None and str(p).strip()]
        if parts:
            payer_name = " ".join(parts).strip()
    return tracking, payer_name, amount_minor, paid_date_iso, ""

def parse_cod_major_to_minor(value: object) -> int | None:
    if value is None:
        return None

    if isinstance(value, int):
        return value * 100

    text = str(value).strip()
    if not text:
        return None

    try:
        dec = Decimal(text)
    except InvalidOperation:
        return None

    minor = int((dec.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) * 100).to_integral_value())
    return minor


def parse_minor_int(value: object) -> int | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(text)
    except ValueError:
        return None


def parse_payment_amount_to_minor(value: object) -> int | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None

    text = text.replace("KM", "").replace("km", "").replace(" ", "")
    if not text:
        return None

    if "," in text or "." in text:
        text = text.replace(",", ".")
        try:
            dec = Decimal(text)
        except InvalidOperation:
            return None
        return int((dec.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) * 100).to_integral_value())

    try:
        return int(text)
    except ValueError:
        return None


def parse_km_major_to_minor(value: object) -> int | None:
    """
    Parse a user-entered KM amount into minor units.
    This is intended for UI inputs labeled as KM (major units), NOT for raw CSV fields.

    Rules:
      - "154"   -> 15400
      - "154,0" -> 15400
      - "154,50" -> 15450
      - "154.50" -> 15450
    """
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("KM", "").replace("km", "").replace(" ", "")
    if not text:
        return None

    if "," in text or "." in text:
        text = text.replace(",", ".")
        try:
            dec = Decimal(text)
        except InvalidOperation:
            return None
        return int((dec.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) * 100).to_integral_value())

    # No decimal separator: treat as whole KM (major) for UI inputs.
    try:
        return int(text) * 100
    except ValueError:
        return None


def parse_bank_amount_to_minor(value: object) -> int | None:
    """
    Parse bank statement amount (typically with thousand separator '.' and decimal comma ','):
      - '2.205,17' -> 220517
      - '71,95' -> 7195
      - 2205.17 (float) -> 220517 (best-effort)
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            dec = Decimal(str(value))
        except InvalidOperation:
            return None
        return int((dec.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) * 100).to_integral_value())

    text = str(value).strip()
    if not text:
        return None
    text = text.replace("KM", "").replace("km", "").replace(" ", "")
    if not text:
        return None
    # Remove thousands separators and normalize decimal comma.
    text = text.replace(".", "").replace(",", ".")
    try:
        dec = Decimal(text)
    except InvalidOperation:
        return None
    return int((dec.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) * 100).to_integral_value())


def parse_bank_ts_to_iso(value: object) -> tuple[str | None, str | None]:
    """
    Parse bank statement timestamp from Excel:
      - '29.01.2026 11:47:10' -> ('2026-01-29T11:47:10', '2026-01-29')
      - datetime -> ISO equivalents
    """
    if value is None:
        return None, None
    if isinstance(value, datetime):
        ts = value.replace(microsecond=0).isoformat(sep="T")
        return ts, value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat() + "T00:00:00", value.isoformat()
    text = str(value).strip()
    if not text:
        return None, None
    # Expected 'DD.MM.YYYY HH:MM:SS'
    try:
        dt = datetime.strptime(text, "%d.%m.%Y %H:%M:%S")
        return dt.isoformat(sep="T"), dt.date().isoformat()
    except Exception:
        pass
    # Fallback: try without seconds
    try:
        dt = datetime.strptime(text, "%d.%m.%Y %H:%M")
        return dt.isoformat(sep="T"), dt.date().isoformat()
    except Exception:
        return None, None


def detect_csv_encoding(path: Path, log: callable) -> str:
    encodings = ["utf-8-sig", "cp1250", "iso-8859-2", "cp1252"]

    def looks_ok(sample: str) -> bool:
        if "�" in sample:
            return False
        # Common mojibake markers for our locale (e.g. MATIÄ†, BEKRIÆ)
        suspicious = ("Ã", "Ä", "Â", "Æ", "†")
        return not any(s in sample for s in suspicious)

    for enc in encodings:
        try:
            with open(path, "r", encoding=enc, errors="replace", newline="") as fh:
                lines: list[str] = []
                for _ in range(501):  # header + first 500 lines (avoid false positives)
                    line = fh.readline()
                    if not line:
                        break
                    lines.append(line)
                sample = "".join(lines)
            if looks_ok(sample):
                log(f"[CSV] Encoding odabran: {enc}")
                return enc
        except Exception:
            continue

    # Fallback: last resort
    log("[CSV] Encoding fallback: cp1252")
    return "cp1252"


def infer_paid_date_from_csv(csv_path: Path) -> str | None:
    # Infer the most common valid date in column 5 (index 4).
    encoding = detect_csv_encoding(csv_path, lambda _msg: None)
    counts: dict[str, int] = {}
    try:
        with open(csv_path, "r", encoding=encoding, errors="replace", newline="") as fh:
            reader = csv.reader(fh, delimiter=";")
            for _line_no, cols in enumerate(reader, start=1):
                if not cols:
                    continue
                _tracking, _payer, _amount_minor, paid_date_iso, err = parse_bhp_uplate_row(cols)
                if err or not paid_date_iso:
                    continue
                counts[paid_date_iso] = counts.get(paid_date_iso, 0) + 1
                if sum(counts.values()) >= 2000:
                    break
    except Exception:
        return None

    if not counts:
        return None
    return max(counts.items(), key=lambda kv: (kv[1], kv[0]))[0]


def preview_bhp_uplate_csv(csv_path: Path) -> dict[str, object]:
    # Preview totals/counts from a CSV before importing.
    # Valid row (for preview): has tracking + valid amount. Date is optional for counting,
    # but used for "found date" inference.
    encoding = detect_csv_encoding(csv_path, lambda _msg: None)
    total_minor = 0
    valid_count = 0
    invalid_count = 0
    missing_tracking = 0
    invalid_amount = 0
    invalid_date = 0
    parse_error = 0
    date_counts: dict[str, int] = {}

    try:
        with open(csv_path, "r", encoding=encoding, errors="replace", newline="") as fh:
            reader = csv.reader(fh, delimiter=";")
            for _line_no, cols in enumerate(reader, start=1):
                if not cols:
                    continue
                tracking, _payer, amount_minor, paid_date_iso, err = parse_bhp_uplate_row(cols)
                if err:
                    invalid_count += 1
                    if err == "parse_error":
                        parse_error += 1
                    elif err == "missing_tracking":
                        missing_tracking += 1
                    elif err == "invalid_amount":
                        invalid_amount += 1
                    elif err == "invalid_date":
                        invalid_date += 1
                    continue

                total_minor += int(amount_minor or 0)
                valid_count += 1
                if paid_date_iso:
                    date_counts[paid_date_iso] = date_counts.get(paid_date_iso, 0) + 1

                if valid_count >= 200000:
                    break
    except Exception:
        pass

    found_iso = None
    if date_counts:
        found_iso = max(date_counts.items(), key=lambda kv: (kv[1], kv[0]))[0]

    return {
        "encoding": encoding,
        "found_date_iso": found_iso,
        "total_minor": total_minor,
        "valid_count": valid_count,
        "invalid_count": invalid_count,
        "invalid_breakdown": {
            "missing_tracking": missing_tracking,
            "invalid_amount": invalid_amount,
            "invalid_date": invalid_date,
            "parse_error": parse_error,
        },
    }


def parse_km_text_to_minor(value: str) -> int:
    text = (value or "").strip()
    if not text:
        raise ValueError("Empty amount")
    text = text.replace(",", ".")
    dec = Decimal(text)
    return int((dec.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) * 100).to_integral_value())


def compute_payment_hash(
    *,
    carrier: str,
    tracking: str | None,
    payer_name: str | None,
    paid_amount_minor: int | None,
    paid_date: str | None,
) -> str:
    payload = "\x1f".join(
        [
            carrier.strip().upper(),
            (tracking or "").strip().upper(),
            (payer_name or "").strip(),
            "" if paid_amount_minor is None else str(int(paid_amount_minor)),
            (paid_date or "").strip(),
        ]
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def compute_payment_row_hash(*, carrier: str, cols: list[str]) -> str:
    # Dedupe by the source CSV row content (normalized), independent of any UI date override.
    # This prevents importing the same file/row twice even if the user selects a different paid_date.
    norm_cols = [("" if c is None else str(c).strip()) for c in cols]
    payload = "\x1f".join([carrier.strip().upper(), *norm_cols])
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def format_minor_km(value: int | None) -> str:
    if value is None:
        return ""
    sign = "-" if value < 0 else ""
    value = abs(int(value))
    return f"{sign}{value // 100},{value % 100:02d}"


def _same_person_name(a: str | None, b: str | None) -> bool:
    aa = " ".join((a or "").strip().lower().split())
    bb = " ".join((b or "").strip().lower().split())
    if not aa or not bb:
        return False
    return aa == bb


def _is_older_than_12_months(prev_iso: str | None, cur_iso: str | None) -> bool:
    # Strict > 365 days rule.
    prev_d = parse_iso_date(prev_iso)
    cur_d = parse_iso_date(cur_iso)
    if prev_d is None or cur_d is None:
        return False
    return (cur_d - prev_d).days > 365


def _append_duplicate_payment_log(line: str) -> None:
    try:
        DUPLICATE_PAYMENTS_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with DUPLICATE_PAYMENTS_LOG_PATH.open("a", encoding="utf-8") as f:
            f.write(f"{ts} {line}\n")
    except Exception:
        # Never crash imports due to file logging.
        pass


def _payment_amount_display(paid_amount_minor: int) -> str:
    # Display as KM for user editing.
    return format_minor_km(int(paid_amount_minor))


def ui_status_label(state: str | None) -> str:
    mapping = {
        "OPEN": "Otvoreno",
        "PARTIAL": "Djelimično uplaćeno",
        "PAID": "Uplaćeno",
        "MISMATCH": "Neusklađeno",
        "UNDELIVERED_REMOVED": "Nepreuzeto",
    }
    if not state:
        return ""
    return mapping.get(state, state)


def compute_lifecycle_and_settlement(
    cod_amount_minor: int | None, paid_total_minor: int, payments_count: int
) -> tuple[str, str, int | None]:
    # If there is no COD amount (0 or missing) and no payments, this shipment should not be treated as a claim.
    if (cod_amount_minor is None or int(cod_amount_minor) <= 0) and (payments_count <= 0 or paid_total_minor == 0):
        return "PAID", "FULL", 0

    if payments_count <= 0 or paid_total_minor == 0:
        return "OPEN", "UNMATCHED", (cod_amount_minor if cod_amount_minor is not None else None)

    if cod_amount_minor is None or cod_amount_minor <= 0 or paid_total_minor < 0:
        return "MISMATCH", "MISMATCH", None

    difference = int(cod_amount_minor) - int(paid_total_minor)
    if paid_total_minor == cod_amount_minor:
        return "PAID", "FULL", 0
    if 0 < paid_total_minor < cod_amount_minor:
        return "PARTIAL", "PARTIAL", difference
    return "MISMATCH", "MISMATCH", difference


def month_bounds(year: int, month: int) -> tuple[str, str]:
    start = date(year, month, 1)
    if month == 12:
        next_start = date(year + 1, 1, 1)
    else:
        next_start = date(year, month + 1, 1)
    return start.isoformat(), next_start.isoformat()


def bhp_tracking_url(tracking: str) -> str:
    normalized = "".join(str(tracking).split())
    encoded = quote(normalized, safe="")
    return f"https://www.posta.ba/pracenje-posiljaka/?prijemnibroj={encoded}"


def bhp_tracking_events_url(tracking: str) -> str:
    normalized = "".join(str(tracking).split())
    encoded = quote(normalized, safe="")
    return f"https://bhpwebout.posta.ba/trackwebapp/pracenje-posiljaka/?prijemniBroj={encoded}"


def bhp_trackwebapp_root_url() -> str:
    return "https://bhpwebout.posta.ba/trackwebapp/"


def bhp_trackwebapp_post_url(tracking: str) -> str:
    normalized = "".join(str(tracking).split())
    encoded = quote(normalized, safe="")
    # WebForms expects lower-case query param name here (as provided by user).
    return f"https://bhpwebout.posta.ba/trackwebapp/?prijemnibroj={encoded}"


def _decode_html_bytes(data: bytes) -> str:
    for enc in ("utf-8", "utf-8-sig", "cp1250", "iso-8859-2", "cp1252"):
        try:
            return data.decode(enc)
        except Exception:
            continue
    return data.decode("utf-8", errors="replace")


class _HtmlTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self._in_table = False
        self._in_tr = False
        self._in_td = False
        self._cell_parts: list[str] = []
        self._row: list[str] = []
        self._tables: list[list[list[str]]] = []
        self._table: list[list[str]] | None = None

    def handle_starttag(self, tag: str, attrs) -> None:
        t = tag.lower()
        if t == "table":
            self._in_table = True
            self._table = []
        elif t == "tr" and self._in_table:
            self._in_tr = True
            self._row = []
        elif t in ("td", "th") and self._in_table and self._in_tr:
            self._in_td = True
            self._cell_parts = []

    def handle_endtag(self, tag: str) -> None:
        t = tag.lower()
        if t in ("td", "th") and self._in_td:
            text = " ".join("".join(self._cell_parts).split())
            self._row.append(text)
            self._cell_parts = []
            self._in_td = False
        elif t == "tr" and self._in_tr:
            if self._table is not None and self._row:
                self._table.append(self._row)
            self._row = []
            self._in_tr = False
        elif t == "table" and self._in_table:
            if self._table is not None:
                self._tables.append(self._table)
            self._table = None
            self._in_table = False

    def handle_data(self, data: str) -> None:
        if self._in_td and data:
            self._cell_parts.append(data)

    def tables(self) -> list[list[list[str]]]:
        return self._tables


class TrackingNoDataError(RuntimeError):
    pass


class TrackingCaptchaError(RuntimeError):
    pass


def _extract_webforms_hidden_fields(html: str) -> dict[str, str]:
    fields = {"__VIEWSTATE": "", "__VIEWSTATEGENERATOR": "", "__EVENTVALIDATION": ""}

    # Prefer BeautifulSoup when available.
    try:
        from bs4 import BeautifulSoup  # type: ignore

        soup = BeautifulSoup(html, "html.parser")
        for key in list(fields.keys()):
            inp = soup.find("input", attrs={"name": key})
            if inp and inp.has_attr("value"):
                fields[key] = str(inp.get("value") or "")
        return fields
    except Exception:
        pass

    # Fallback regex extraction.
    for key in list(fields.keys()):
        m = re.search(
            rf"""name\s*=\s*['"]{re.escape(key)}['"][^>]*\bvalue\s*=\s*['"]([^'"]*)['"]""",
            html,
            flags=re.IGNORECASE,
        )
        if m:
            fields[key] = m.group(1)
    return fields


def fetch_bhp_events_webforms(
    tracking: str, *, timeout_s: int = 12, debug_log: callable | None = None
) -> list[dict[str, str]]:
    """
    Fetch tracking events from BHP WebForms:
      - GET /trackwebapp/ to obtain session cookies + WebForms hidden fields
      - POST /trackwebapp/?prijemnibroj=<TRACKING> with hidden fields + tbBrojPosiljke + btnIzlistaj
    """
    tracking_norm = "".join(str(tracking).split())
    ua = (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"
    )

    jar = CookieJar()
    try:
        import urllib.request as _urllib_request

        opener = _urllib_request.build_opener(_urllib_request.HTTPCookieProcessor(jar))
    except Exception:
        opener = None

    def _open(req: Request, timeout: int) -> object:
        if opener is not None:
            return opener.open(req, timeout=timeout)
        return urlopen(req, timeout=timeout)

    # 1) GET root
    get_url = bhp_trackwebapp_root_url()
    get_req = Request(
        get_url,
        headers={
            "User-Agent": ua,
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "bs-BA,hr-HR,bs;q=0.9,en-US;q=0.7,en;q=0.5",
            "Referer": "https://bhpwebout.posta.ba/",
        },
        method="GET",
    )
    try:
        with _open(get_req, timeout_s) as resp:
            status = int(getattr(resp, "status", None) or getattr(resp, "getcode", lambda: 0)() or 0)
            data = resp.read()
            if debug_log:
                try:
                    url_final = ""
                    try:
                        url_final = str(resp.geturl() or "")
                    except Exception:
                        url_final = ""
                    debug_log(f"webforms_get status={status} url={url_final or get_url} bytes={len(data)}")
                except Exception:
                    pass
            if status != 200:
                raise RuntimeError("HTTP error")
            get_html = _decode_html_bytes(data)
    except Exception:
        raise RuntimeError("Nije moguće dohvatiti podatke sa BHP sistema.")

    lower_get = get_html.lower()
    if "recaptcha" in lower_get or "g-recaptcha" in lower_get:
        raise TrackingCaptchaError("Potrebna je verifikacija (reCAPTCHA).")

    hidden = _extract_webforms_hidden_fields(get_html)
    if debug_log:
        try:
            debug_log(
                "webforms_hidden "
                f"viewstate={len(hidden.get('__VIEWSTATE','') or '')} "
                f"generator={len(hidden.get('__VIEWSTATEGENERATOR','') or '')} "
                f"eventvalidation={len(hidden.get('__EVENTVALIDATION','') or '')}"
            )
        except Exception:
            pass
    if not hidden.get("__VIEWSTATE") or not hidden.get("__EVENTVALIDATION"):
        # WebForms page structure changed / blocked.
        if debug_log:
            try:
                debug_log(f"webforms_hidden_missing head={repr(get_html[:500])}")
            except Exception:
                pass

    # 2) POST with same session
    post_url = bhp_trackwebapp_post_url(tracking_norm)
    form = {
        "__EVENTTARGET": "",
        "__EVENTARGUMENT": "",
        "__LASTFOCUS": "",
        "__VIEWSTATE": hidden.get("__VIEWSTATE", ""),
        "__VIEWSTATEGENERATOR": hidden.get("__VIEWSTATEGENERATOR", ""),
        "__EVENTVALIDATION": hidden.get("__EVENTVALIDATION", ""),
        "tbBrojPosiljke": tracking_norm,
        "btnIzlistaj": "Prikaži",
    }
    body = urlencode(form).encode("utf-8")
    post_req = Request(
        post_url,
        data=body,
        headers={
            "User-Agent": ua,
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "bs-BA,hr-HR,bs;q=0.9,en-US;q=0.7,en;q=0.5",
            "Referer": get_url,
            "Content-Type": "application/x-www-form-urlencoded",
        },
        method="POST",
    )
    try:
        with _open(post_req, timeout_s) as resp:
            status = int(getattr(resp, "status", None) or getattr(resp, "getcode", lambda: 0)() or 0)
            data = resp.read()
            if debug_log:
                try:
                    url_final = ""
                    try:
                        url_final = str(resp.geturl() or "")
                    except Exception:
                        url_final = ""
                    debug_log(f"webforms_post status={status} url={url_final or post_url} bytes={len(data)}")
                except Exception:
                    pass
            if status != 200:
                raise RuntimeError("HTTP error")
            post_html = _decode_html_bytes(data)
    except TrackingCaptchaError:
        raise
    except Exception:
        raise RuntimeError("Nije moguće dohvatiti podatke sa BHP sistema.")

    lower_post = post_html.lower()
    if "recaptcha" in lower_post or "g-recaptcha" in lower_post:
        raise TrackingCaptchaError("Potrebna je verifikacija (reCAPTCHA).")

    date_re = re.compile(r"^\d{2}\.\d{2}\.\d{4}")

    # Parse events table (<table class='tabela'> ... "Pregled događaja" ...)
    try:
        from bs4 import BeautifulSoup  # type: ignore

        soup = BeautifulSoup(post_html, "html.parser")
        tables = soup.find_all("table", class_=lambda c: c and "tabela" in str(c))
        if debug_log:
            try:
                debug_log(f"webforms_tables count={len(tables)}")
            except Exception:
                pass
        best_rows: list[list[str]] = []
        best_score = 0
        for t in tables:
            text = t.get_text(" ", strip=True).lower()
            # Prefer the one containing the "Pregled događaja" label.
            bonus = 2 if "pregled događaja" in text or "pregled dogadaja" in text else 0
            rows: list[list[str]] = []
            score = 0
            for tr in t.find_all("tr"):
                cells = [c.get_text(" ", strip=True) for c in tr.find_all(["td", "th"])]
                if not cells:
                    continue
                rows.append(cells)
                if date_re.match(cells[0] or "") and len(cells) >= 5:
                    score += 1
            score += bonus
            if score > best_score:
                best_score = score
                best_rows = rows

        events: list[dict[str, str]] = []
        for cells in best_rows:
            if not cells or not date_re.match(cells[0] or ""):
                continue
            c5 = (cells + ["", "", "", "", ""])[:5]
            events.append(
                {
                    "datetime": c5[0],
                    "country": c5[1],
                    "location": c5[2],
                    "event": c5[3],
                    "note": c5[4],
                }
            )
        if debug_log:
            try:
                debug_log(f"webforms_parsed_events count={len(events)}")
            except Exception:
                pass
        if not events and debug_log:
            try:
                debug_log(f"webforms_no_events head={repr(post_html[:500])}")
            except Exception:
                pass
        if not events:
            raise TrackingNoDataError("Nema događaja ili BHP blokirao zahtjev.")
        return events
    except TrackingNoDataError:
        raise
    except Exception:
        pass

    # Fallback parser (no bs4): best-effort based on any table containing date rows with >= 5 cells.
    parser = _HtmlTableParser()
    parser.feed(post_html)
    best: list[list[str]] = []
    best_score = 0
    for tbl in parser.tables():
        score = 0
        for row in tbl:
            if row and date_re.match((row[0] or "").strip()) and len(row) >= 5:
                score += 1
        if score > best_score:
            best_score = score
            best = tbl

    events: list[dict[str, str]] = []
    for row in best:
        if not row or not date_re.match((row[0] or "").strip()):
            continue
        c5 = (row + ["", "", "", "", ""])[:5]
        events.append(
            {
                "datetime": c5[0],
                "country": c5[1],
                "location": c5[2],
                "event": c5[3],
                "note": c5[4],
            }
        )
    if debug_log:
        try:
            debug_log(f"webforms_fallback_parsed_events count={len(events)}")
        except Exception:
            pass
    if not events:
        raise TrackingNoDataError("Nema događaja ili BHP blokirao zahtjev.")
    return events


def _extract_ajax_candidates_from_html(html: str, *, base: str = "https://www.posta.ba") -> list[tuple[str, dict[str, str]]]:
    """
    Best-effort extraction of potential WP AJAX endpoints from HTML.
    Returns list of (url, params_hint) tuples (params may be empty).
    """
    candidates: list[tuple[str, dict[str, str]]] = []
    seen: set[str] = set()

    # Endpoint URLs.
    url_patterns = [
        r"(https?://[^\s'\"<>]*wp-admin/admin-ajax\.php[^\s'\"<>]*)",
        r"(https?://[^\s'\"<>]*/wp-json/[^\s'\"<>]*)",
        r"([/][^\s'\"<>]*wp-admin/admin-ajax\.php[^\s'\"<>]*)",
        r"([/][^\s'\"<>]*/wp-json/[^\s'\"<>]*)",
    ]
    for pat in url_patterns:
        for m in re.findall(pat, html, flags=re.IGNORECASE):
            u = m
            if u.startswith("/"):
                u = base.rstrip("/") + u
            if u not in seen:
                seen.add(u)
                candidates.append((u, {}))

    # Action/nonces hints inside inline scripts.
    # Keep it conservative: only hints, not guaranteed correct.
    actions: list[str] = []
    for m in re.finditer(r"\baction\s*[:=]\s*['\"]([a-zA-Z0-9_:-]{2,80})['\"]", html):
        actions.append(m.group(1))
    nonces: list[str] = []
    for m in re.finditer(r"\bnonce\s*[:=]\s*['\"]([a-zA-Z0-9_-]{6,120})['\"]", html):
        nonces.append(m.group(1))

    # Attach hints to admin-ajax candidates.
    if actions or nonces:
        for i, (u, _) in enumerate(list(candidates)):
            if "admin-ajax.php" in u.lower():
                params: dict[str, str] = {}
                if actions:
                    params["action"] = actions[0]
                if nonces:
                    params["nonce"] = nonces[0]
                candidates[i] = (u, params)

    return candidates


def fetch_bhp_tracking_events(
    tracking: str, *, timeout_s: int = 12, debug_log: callable | None = None
) -> list[dict[str, str]]:
    # IMPORTANT: do NOT attempt to parse WordPress (www.posta.ba) HTML.
    # Tracking events are fetched exclusively via BHP WebForms (GET+POST with session cookies).
    if debug_log:
        try:
            debug_log("endpoint=bhpwebout trackwebapp (WebForms GET+POST)")
        except Exception:
            pass

    return fetch_bhp_events_webforms(tracking, timeout_s=timeout_s, debug_log=debug_log)




def parse_iso_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except Exception:
        return None


def days_between(start_iso: str | None, end: date) -> int | None:
    d = parse_iso_date(start_iso)
    if d is None:
        return None
    return (end - d).days


def iso_to_ddmmyyyy(value: str | None) -> str:
    d = parse_iso_date(value)
    if d is None:
        return "N/A"
    return d.strftime("%d.%m.%Y")


def parse_ddmmyy_from_filename(name: str) -> date | None:
    # Expected pattern contains DDMMYY, e.g. FEMMA_MO_270126.csv -> 27.01.2026
    stem = Path(name).stem
    parts = stem.split("_")
    if not parts:
        return None
    candidate = parts[-1]
    if len(candidate) != 6 or not candidate.isdigit():
        return None
    dd = int(candidate[0:2])
    mm = int(candidate[2:4])
    yy = int(candidate[4:6])
    year = 2000 + yy
    try:
        return date(year, mm, dd)
    except Exception:
        return None


def export_xlsx(path: Path, sheet_name: str, headers: list[str], rows: list[list[object]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception:
        raise RuntimeError("Nedostaje zavisnost: openpyxl. Instaliraj: pip install -r requirements.txt")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] if sheet_name else "Sheet1"

    ws.append(headers)
    for row in rows:
        ws.append(row)

    widths = [len(str(h)) for h in headers]
    for row in rows:
        for i, val in enumerate(row):
            widths[i] = max(widths[i], len("" if val is None else str(val)))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), 60)

    wb.save(str(path))


def export_xlsx_workbook(path: Path, sheets: list[tuple[str, list[str], list[list[object]]]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception:
        raise RuntimeError("Nedostaje zavisnost: openpyxl. Instaliraj: pip install -r requirements.txt")

    wb = Workbook()
    first = True
    for sheet_name, headers, rows in sheets:
        if first:
            ws = wb.active
            first = False
        else:
            ws = wb.create_sheet()
        ws.title = (sheet_name or "Sheet")[:31]

        ws.append(headers)
        for row in rows:
            ws.append(row)

        widths = [len(str(h)) for h in headers]
        for row in rows:
            for i, val in enumerate(row):
                widths[i] = max(widths[i], len("" if val is None else str(val)))
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), 60)

    wb.save(str(path))


def _km_number(minor: int | None) -> float | None:
    if minor is None:
        return None
    return round(int(minor) / 100.0, 2)


def export_cashbook_dnevnik_xlsx(
    *,
    db: "Database",
    year: int,
    month: int,
    out_path: Path,
    log: callable,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        raise RuntimeError("Nedostaje zavisnost: openpyxl. Instaliraj: pip install -r requirements.txt")

    month_row = db.get_or_create_cashbook_month(int(year), int(month))
    month_id = int(month_row["id"])
    status = str(month_row["status"])
    entries = db.fetch_cashbook_entries(month_id)
    if not entries and status == "OPEN":
        month_row = db.rebuild_cashbook_entries(int(year), int(month))
        month_id = int(month_row["id"])
        entries = db.fetch_cashbook_entries(month_id)

    opening_minor = int(month_row["opening_balance_minor"])
    total_in_minor = int(month_row["total_in_minor"])
    total_out_minor = int(month_row["total_out_minor"])
    closing_minor = int(month_row["closing_balance_minor"])

    start_iso, next_start_iso = month_bounds(int(year), int(month))
    start_d = parse_iso_date(start_iso) or date(int(year), int(month), 1)
    last_d = (parse_iso_date(next_start_iso) or (start_d.replace(day=1) + timedelta(days=32))).replace(day=1) - timedelta(
        days=1
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Blagajna"

    green = "2e7d32"
    thin = Side(style="thin", color=green)
    thick = Side(style="medium", color=green)
    dotted = Side(style="dotted", color=green)

    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)
    border_dotted = Border(left=dotted, right=dotted, top=dotted, bottom=dotted)

    header_font = Font(bold=True, size=14)
    sub_font = Font(bold=False, italic=True, size=11)
    table_head_font = Font(bold=True, size=11)
    normal_font = Font(size=11)
    green_font = Font(color=green, bold=True, size=11)

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    fill_light = PatternFill("solid", fgColor="E8F5E9")

    # Column widths
    widths = {"A": 6, "B": 12, "C": 46, "D": 14, "E": 14, "F": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Header block
    broj = f"{int(month):02d}/{int(year) % 100:02d}"
    ws.merge_cells("A1:F1")
    ws["A1"].value = f"DNEVNIK BLAGAJNE broj: {broj}"
    ws["A1"].font = header_font
    ws["A1"].alignment = center
    ws["A1"].border = border_thick

    ws.merge_cells("A2:F2")
    ws["A2"].value = "(Blagajnički izvještaj)"
    ws["A2"].font = sub_font
    ws["A2"].alignment = center
    ws["A2"].border = border_thick

    for r in (1, 2):
        ws.row_dimensions[r].height = 24 if r == 1 else 18
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.border = border_thick

    # Table header
    table_header_row = 4
    headers = ["R.br.", "Temeljnica", "OPIS", "ULAZ", "IZLAZ", "Br.računa"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=table_header_row, column=i, value=h)
        cell.font = table_head_font
        cell.alignment = center
        cell.border = border_thin
        cell.fill = fill_light
    ws.row_dimensions[table_header_row].height = 18

    # Entries
    row = table_header_row + 1
    for e in entries:
        rb = int(e["rb"])
        entry_date = str(e["entry_date"])
        desc = str(e["description"])
        source_type = str(e["source_type"])
        in_minor = int(e["in_minor"])
        out_minor = int(e["out_minor"])

        # Match requested wording for payments
        if source_type == "PAYMENTS_OUT":
            desc = f"Uplata BHP {iso_to_ddmmyyyy(entry_date)}"

        ws.cell(row=row, column=1, value=rb).alignment = center
        ws.cell(row=row, column=2, value=rb).alignment = center
        ws.cell(row=row, column=3, value=desc).alignment = left

        in_cell = ws.cell(row=row, column=4, value=_km_number(in_minor) if in_minor else None)
        out_cell = ws.cell(row=row, column=5, value=_km_number(out_minor) if out_minor else None)
        in_cell.alignment = right
        out_cell.alignment = right
        in_cell.number_format = "#,##0.00"
        out_cell.number_format = "#,##0.00"

        ws.cell(row=row, column=6, value="").alignment = center

        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.font = normal_font
            cell.border = border_dotted
        ws.row_dimensions[row].height = 18
        row += 1

    # Outline border for the whole table area
    last_table_row = row - 1 if row > table_header_row + 1 else table_header_row
    for c in range(1, 7):
        ws.cell(row=table_header_row, column=c).border = border_thin
        ws.cell(row=last_table_row, column=c).border = border_thin

    # Apply thick outer border to table block
    for r in range(table_header_row, last_table_row + 1):
        ws.cell(row=r, column=1).border = Border(left=thick, right=thin, top=thin, bottom=thin)
        ws.cell(row=r, column=6).border = Border(left=thin, right=thick, top=thin, bottom=thin)
    for c in range(1, 7):
        top_cell = ws.cell(row=table_header_row, column=c)
        bottom_cell = ws.cell(row=last_table_row, column=c)
        top_cell.border = Border(left=top_cell.border.left, right=top_cell.border.right, top=thick, bottom=top_cell.border.bottom)
        bottom_cell.border = Border(left=bottom_cell.border.left, right=bottom_cell.border.right, top=bottom_cell.border.top, bottom=thick)

    # Summary block (approximate layout)
    summary_row = last_table_row + 2
    label_col = 3  # C
    val_in_col = 4  # D
    val_out_col = 5  # E

    def sum_row(label: str, in_minor: int | None = None, out_minor: int | None = None, single_minor: int | None = None):
        nonlocal summary_row
        ws.cell(row=summary_row, column=label_col, value=label).alignment = left
        ws.cell(row=summary_row, column=label_col).font = green_font
        if single_minor is not None:
            ws.merge_cells(
                start_row=summary_row, start_column=val_in_col, end_row=summary_row, end_column=val_out_col
            )
            v = ws.cell(row=summary_row, column=val_in_col, value=_km_number(single_minor))
            v.alignment = right
            v.number_format = "#,##0.00"
            v.font = Font(bold=True, size=11)
        else:
            v_in = ws.cell(row=summary_row, column=val_in_col, value=_km_number(in_minor) if in_minor is not None else None)
            v_out = ws.cell(row=summary_row, column=val_out_col, value=_km_number(out_minor) if out_minor is not None else None)
            v_in.alignment = right
            v_out.alignment = right
            v_in.number_format = "#,##0.00"
            v_out.number_format = "#,##0.00"
            v_in.font = Font(bold=True, size=11)
            v_out.font = Font(bold=True, size=11)

        for c in range(label_col, val_out_col + 1):
            ws.cell(row=summary_row, column=c).border = border_thin
        summary_row += 1

    sum_row("Promet blagajne", in_minor=total_in_minor, out_minor=total_out_minor)
    sum_row(f"Saldo od {start_d.strftime('%d.%m.%Y')}", single_minor=opening_minor)
    sum_row("Ukupni promet", single_minor=opening_minor + total_in_minor)
    sum_row("Odbija se izdatak", single_minor=total_out_minor)
    sum_row(f"Saldo od {last_d.strftime('%d.%m.%Y')}", single_minor=closing_minor)

    # Thick border around summary block
    summary_top = (last_table_row + 2)
    summary_bottom = summary_row - 1
    for r in range(summary_top, summary_bottom + 1):
        ws.cell(row=r, column=label_col).border = Border(left=thick, right=thin, top=thin, bottom=thin)
        ws.cell(row=r, column=val_out_col).border = Border(left=thin, right=thick, top=thin, bottom=thin)
    for c in range(label_col, val_out_col + 1):
        ws.cell(row=summary_top, column=c).border = Border(
            left=ws.cell(row=summary_top, column=c).border.left,
            right=ws.cell(row=summary_top, column=c).border.right,
            top=thick,
            bottom=ws.cell(row=summary_top, column=c).border.bottom,
        )
        ws.cell(row=summary_bottom, column=c).border = Border(
            left=ws.cell(row=summary_bottom, column=c).border.left,
            right=ws.cell(row=summary_bottom, column=c).border.right,
            top=ws.cell(row=summary_bottom, column=c).border.top,
            bottom=thick,
        )

    # Simple signature placeholders (optional, like the form)
    sig_row = summary_bottom + 3
    ws.merge_cells(start_row=sig_row, start_column=2, end_row=sig_row, end_column=3)
    ws.merge_cells(start_row=sig_row, start_column=5, end_row=sig_row, end_column=6)
    ws.cell(row=sig_row, column=2, value="Kontrolisao").alignment = center
    ws.cell(row=sig_row, column=5, value="Blagajnik").alignment = center
    for c in range(2, 4):
        ws.cell(row=sig_row, column=c).border = border_thin
    for c in range(5, 7):
        ws.cell(row=sig_row, column=c).border = border_thin

    wb.save(str(out_path))
    log(f"[IZVOZ] Dnevnik blagajne: {out_path}")


def apply_ui_scale(root: Tk, scale: float = UI_SCALE) -> None:
    # Increase overall readability without changing business logic.
    # We scale named Tk fonts and key ttk widget styles.
    for font_name in ("TkDefaultFont", "TkTextFont", "TkFixedFont", "TkMenuFont", "TkHeadingFont"):
        try:
            f = tkfont.nametofont(font_name)
            size = f.cget("size")
            if isinstance(size, int) and size != 0:
                f.configure(size=int(round(size * float(scale))))
        except Exception:
            pass

    try:
        style = ttk.Style(root)
        # Slightly tighter layout to avoid excessive whitespace while keeping fonts larger.
        style.configure("TNotebook.Tab", padding=(int(8 * scale), int(5 * scale)))
        # Make active tab easier to spot.
        try:
            style.map(
                "TNotebook.Tab",
                background=[("selected", "#dbeafe")],  # slight blue
                foreground=[("selected", "#0b3d91")],
            )
        except Exception:
            pass
        style.configure("Treeview", rowheight=int(round(18 * scale)))
        style.configure("Treeview.Heading", font=tkfont.nametofont("TkHeadingFont"))
    except Exception:
        pass


def center_window(win: Toplevel, parent: Tk) -> None:
    win.update_idletasks()
    w = win.winfo_width() or win.winfo_reqwidth()
    h = win.winfo_height() or win.winfo_reqheight()
    px = parent.winfo_rootx()
    py = parent.winfo_rooty()
    pw = parent.winfo_width()
    ph = parent.winfo_height()
    x = px + max(int((pw - w) / 2), 0)
    y = py + max(int((ph - h) / 2), 0)
    win.geometry(f"{w}x{h}+{x}+{y}")


class Tooltip:
    def __init__(self, widget, text: str, *, wrap: int = 420) -> None:
        self.widget = widget
        self.text = text
        self.wrap = wrap
        self._tip: Toplevel | None = None
        self._after_id: str | None = None

        widget.bind("<Enter>", self._on_enter)
        widget.bind("<Leave>", self._on_leave)
        widget.bind("<ButtonPress>", self._on_leave)

    def _on_enter(self, _e=None) -> None:
        try:
            self._after_id = self.widget.after(400, self.show)
        except Exception:
            self.show()

    def _on_leave(self, _e=None) -> None:
        try:
            if self._after_id:
                self.widget.after_cancel(self._after_id)
        except Exception:
            pass
        self._after_id = None
        self.hide()

    def show(self) -> None:
        if self._tip is not None:
            return
        try:
            x = self.widget.winfo_rootx() + 12
            y = self.widget.winfo_rooty() + self.widget.winfo_height() + 10
        except Exception:
            x, y = 100, 100

        tip = Toplevel(self.widget)
        apply_window_icon(tip)
        tip.wm_overrideredirect(True)
        tip.wm_geometry(f"+{x}+{y}")
        try:
            tip.attributes("-topmost", True)
        except Exception:
            pass

        frm = Frame(tip, bd=1, relief="solid", background="#ffffe0")
        frm.pack(fill="both", expand=True)
        lbl = ttk.Label(frm, text=self.text, justify="left", wraplength=self.wrap, background="#ffffe0")
        lbl.pack(padx=10, pady=8)
        self._tip = tip

    def hide(self) -> None:
        if self._tip is None:
            return
        try:
            self._tip.destroy()
        except Exception:
            pass
        self._tip = None


_MONTH_NAMES_BS = [
    "januar",
    "februar",
    "mart",
    "april",
    "maj",
    "juni",
    "juli",
    "august",
    "septembar",
    "oktobar",
    "novembar",
    "decembar",
]


class CalendarPicker(Frame):
    def __init__(self, parent, *, initial: date | None = None) -> None:
        super().__init__(parent)
        self._today = date.today()
        self._selected = initial or self._today
        self._display_year = self._selected.year
        self._display_month = self._selected.month

        header = Frame(self)
        header.pack(fill="x", pady=(0, 6))

        ttk.Label(header, text="Mjesec:").pack(side=LEFT)
        self.month_var = StringVar(value=_MONTH_NAMES_BS[self._display_month - 1])
        self.month_cb = ttk.Combobox(header, width=12, values=_MONTH_NAMES_BS, state="readonly", textvariable=self.month_var)
        self.month_cb.pack(side=LEFT, padx=(6, 10))

        ttk.Label(header, text="Godina:").pack(side=LEFT)
        self.year_var = StringVar(value=str(self._display_year))
        self.year_sb = Spinbox(header, from_=2000, to=2100, width=6, textvariable=self.year_var)
        self.year_sb.pack(side=LEFT, padx=(6, 0))

        self.month_cb.bind("<<ComboboxSelected>>", lambda _e: self._on_month_year_changed())
        self.year_sb.bind("<FocusOut>", lambda _e: self._on_month_year_changed())
        self.year_sb.bind("<Return>", lambda _e: self._on_month_year_changed())

        # Weekday header
        grid = Frame(self)
        grid.pack()
        self._grid = grid
        weekdays = ["pon", "uto", "sri", "čet", "pet", "sub", "ned"]
        for col, name in enumerate(weekdays):
            lbl = ttk.Label(grid, text=name)
            lbl.grid(row=0, column=col, padx=2, pady=2)

        self._day_buttons: list[Button] = []
        for r in range(1, 7):
            for c in range(7):
                b = Button(grid, text="", width=3, command=lambda: None)
                b.grid(row=r, column=c, padx=1, pady=1)
                self._day_buttons.append(b)

        self._render()

    def _on_month_year_changed(self) -> None:
        try:
            month_name = self.month_var.get()
            month = _MONTH_NAMES_BS.index(month_name) + 1
        except Exception:
            month = self._display_month
        try:
            year = int(self.year_var.get())
        except Exception:
            year = self._display_year
        self._display_year = year
        self._display_month = month
        self._render()

    def _render(self) -> None:
        cal = calendar.Calendar(firstweekday=0)  # Monday
        weeks = cal.monthdatescalendar(self._display_year, self._display_month)
        # Ensure 6 weeks
        while len(weeks) < 6:
            last = weeks[-1][-1]
            weeks.append([last + timedelta(days=i) for i in range(1, 8)])

        def make_cmd(d: date):
            return lambda: self.selection_set(d)

        idx = 0
        for week in weeks[:6]:
            for d in week:
                b = self._day_buttons[idx]
                idx += 1
                b.configure(text=str(d.day), command=make_cmd(d))
                if d.month != self._display_month:
                    b.configure(state="disabled", fg="#999999")
                else:
                    b.configure(state="normal", fg="#000000")

                if d == self._selected:
                    b.configure(relief="sunken", bg="#d9ead3")
                else:
                    b.configure(relief="raised", bg="SystemButtonFace")

    def selection_get(self) -> date:
        return self._selected

    def selection_set(self, d: date) -> None:
        self._selected = d
        self._display_year = d.year
        self._display_month = d.month
        self.month_var.set(_MONTH_NAMES_BS[self._display_month - 1])
        self.year_var.set(str(self._display_year))
        self._render()


class DateField(Frame):
    def __init__(self, parent, *, var: StringVar, width: int = 12, on_change: callable | None = None) -> None:
        super().__init__(parent)
        self.var = var
        self.on_change = on_change
        self.entry = ttk.Entry(self, width=width, textvariable=self.var)
        self.entry.pack(side=LEFT)
        Button(self, text="▼", width=2, command=self._open).pack(side=LEFT, padx=(4, 0))
        self.entry.bind("<Return>", lambda _e: self._changed())
        self.entry.bind("<FocusOut>", lambda _e: self._changed())

    def _open(self) -> None:
        initial_iso = parse_date_to_iso(self.var.get())
        initial = parse_iso_date(initial_iso) if initial_iso else None
        d = date_picker_dialog(self.winfo_toplevel(), initial=initial)
        if d is None:
            return
        self.var.set(d.isoformat())
        self._changed()

    def _changed(self) -> None:
        if self.on_change:
            try:
                self.on_change()
            except Exception:
                pass


def date_picker_dialog(parent: Tk, *, initial: date | None = None, title: str = "Odaberi datum") -> date | None:
    win = Toplevel(parent)
    apply_window_icon(win)
    win.title(title)
    win.transient(parent)
    win.grab_set()
    win.resizable(False, False)

    result: dict[str, object] = {"ok": False}
    picker = CalendarPicker(win, initial=initial)
    picker.pack(padx=12, pady=12)

    btns = Frame(win)
    btns.pack(fill="x", padx=12, pady=(0, 12))

    def ok():
        result["ok"] = True
        win.destroy()

    def cancel():
        win.destroy()

    Button(btns, text="Odustani", command=cancel).pack(side=RIGHT, padx=(6, 0))
    Button(btns, text="Potvrdi", command=ok).pack(side=RIGHT)

    win.bind("<Escape>", lambda _e: cancel())
    win.bind("<Return>", lambda _e: ok())
    center_window(win, parent)
    win.focus_set()
    parent.wait_window(win)

    if not result.get("ok"):
        return None
    return picker.selection_get()


class SettingsStore:
    def __init__(self, path: Path) -> None:
        self.path = path
        self.data: dict[str, object] = {}
        self.load()

    def load(self) -> None:
        try:
            if self.path.exists():
                self.data = json.loads(self.path.read_text(encoding="utf-8"))
            else:
                self.data = {}
        except Exception:
            self.data = {}

    def save(self) -> None:
        tmp = self.path.with_suffix(".tmp")
        tmp.write_text(json.dumps(self.data, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp.replace(self.path)

    def get(self, key: str, default: object = None) -> object:
        return self.data.get(key, default)

    def set(self, key: str, value: object) -> None:
        self.data[key] = value


class Database:
    def __init__(self, path: Path) -> None:
        self.path = path
        self.conn = sqlite3.connect(self.path, check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self._lock = threading.RLock()
        with self._lock:
            self.conn.execute("PRAGMA foreign_keys = ON;")
            self.conn.execute("PRAGMA journal_mode = WAL;")
            self.conn.execute("PRAGMA synchronous = NORMAL;")
            self.migrate()

    def close(self) -> None:
        with self._lock:
            self.conn.close()

    def migrate(self) -> None:
        version = int(self.conn.execute("PRAGMA user_version;").fetchone()[0])
        if version < 1:
            self.conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS shipments (
                    id INTEGER PRIMARY KEY,
                    carrier TEXT NOT NULL,
                    tracking TEXT NOT NULL,
                    status_text TEXT,
                    recipient_name TEXT,
                    recipient_phone TEXT,
                    pickup_date TEXT,
                    cod_amount_minor INTEGER,
                    lifecycle_state TEXT NOT NULL DEFAULT 'active',
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    UNIQUE (carrier, tracking)
                );

                CREATE TABLE IF NOT EXISTS payments (
                    id INTEGER PRIMARY KEY,
                    carrier TEXT NOT NULL,
                    tracking TEXT NOT NULL,
                    payer_name TEXT,
                    paid_amount_minor INTEGER NOT NULL,
                    paid_date TEXT,
                    raw_hash TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    UNIQUE (raw_hash)
                );

                CREATE INDEX IF NOT EXISTS idx_shipments_carrier_tracking ON shipments(carrier, tracking);
                CREATE INDEX IF NOT EXISTS idx_payments_carrier_tracking ON payments(carrier, tracking);
                """
            )
            self.conn.execute("PRAGMA user_version = 1;")
            self.conn.commit()
            version = 1

        if version < 2:
            self.conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS settlements (
                    id INTEGER PRIMARY KEY,
                    shipment_id INTEGER NOT NULL,
                    paid_total_minor INTEGER NOT NULL,
                    difference_minor INTEGER,
                    match_type TEXT NOT NULL,
                    last_matched_at TEXT NOT NULL,
                    UNIQUE (shipment_id),
                    FOREIGN KEY (shipment_id) REFERENCES shipments(id)
                );

                CREATE INDEX IF NOT EXISTS idx_settlements_shipment_id ON settlements(shipment_id);
                """
            )
            self.conn.execute("PRAGMA user_version = 2;")
            self.conn.commit()
            version = 2

        if version < 3:
            self.conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS cashbook_months (
                    id INTEGER PRIMARY KEY,
                    year INTEGER NOT NULL,
                    month INTEGER NOT NULL,
                    opening_balance_minor INTEGER NOT NULL DEFAULT 0,
                    total_in_minor INTEGER NOT NULL DEFAULT 0,
                    total_out_minor INTEGER NOT NULL DEFAULT 0,
                    closing_balance_minor INTEGER NOT NULL DEFAULT 0,
                    status TEXT NOT NULL DEFAULT 'OPEN',
                    closed_at TEXT,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    UNIQUE (year, month)
                );

                CREATE TABLE IF NOT EXISTS cashbook_entries (
                    id INTEGER PRIMARY KEY,
                    cashbook_month_id INTEGER NOT NULL,
                    rb INTEGER NOT NULL,
                    entry_date TEXT NOT NULL,
                    description TEXT NOT NULL,
                    in_minor INTEGER NOT NULL DEFAULT 0,
                    out_minor INTEGER NOT NULL DEFAULT 0,
                    source_type TEXT NOT NULL,
                    source_ref TEXT,
                    created_at TEXT NOT NULL,
                    FOREIGN KEY (cashbook_month_id) REFERENCES cashbook_months(id)
                );

                CREATE INDEX IF NOT EXISTS idx_cashbook_entries_month_id ON cashbook_entries(cashbook_month_id);
                CREATE INDEX IF NOT EXISTS idx_cashbook_entries_month_id_rb ON cashbook_entries(cashbook_month_id, rb);

                CREATE TABLE IF NOT EXISTS audit_log (
                    id INTEGER PRIMARY KEY,
                    ts TEXT NOT NULL,
                    actor TEXT NOT NULL,
                    action TEXT NOT NULL,
                    target TEXT NOT NULL,
                    details TEXT
                );
                """
            )
            self.conn.execute("PRAGMA user_version = 3;")
            self.conn.commit()
            version = 3

        if version < 4:
            cols = {r["name"] for r in self.conn.execute("PRAGMA table_info(shipments);").fetchall()}
            alters: list[str] = []
            if "removed_at" not in cols:
                alters.append("ALTER TABLE shipments ADD COLUMN removed_at TEXT;")
            if "removed_reason" not in cols:
                alters.append("ALTER TABLE shipments ADD COLUMN removed_reason TEXT;")
            if "removed_note" not in cols:
                alters.append("ALTER TABLE shipments ADD COLUMN removed_note TEXT;")
            if alters:
                for stmt in alters:
                    self.conn.execute(stmt)
                self.conn.commit()

            # Enforce uniqueness even if the table was created without a UNIQUE constraint.
            # Note: this will fail if duplicates already exist in the DB.
            self.conn.execute(
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_shipments_carrier_tracking ON shipments(carrier, tracking);"
            )
            self.conn.commit()
            self.conn.execute("PRAGMA user_version = 4;")
            self.conn.commit()
            version = 4

        if version < 5:
            self.conn.executescript(
                """
                CREATE INDEX IF NOT EXISTS idx_shipments_pickup_date ON shipments(pickup_date);
                CREATE INDEX IF NOT EXISTS idx_shipments_lifecycle_state ON shipments(lifecycle_state);
                CREATE INDEX IF NOT EXISTS idx_shipments_recipient_name ON shipments(recipient_name);

                CREATE INDEX IF NOT EXISTS idx_payments_paid_date ON payments(paid_date);
                """
            )
            self.conn.execute("PRAGMA user_version = 5;")
            self.conn.commit()
            version = 5

        if version < 6:
            cols = {r["name"] for r in self.conn.execute("PRAGMA table_info(shipments);").fetchall()}
            alters: list[str] = []
            if "status_updated_at" not in cols:
                alters.append("ALTER TABLE shipments ADD COLUMN status_updated_at TEXT;")
            if "delivery_date" not in cols:
                alters.append("ALTER TABLE shipments ADD COLUMN delivery_date TEXT;")
            if alters:
                for stmt in alters:
                    self.conn.execute(stmt)
                self.conn.commit()
            self.conn.execute("PRAGMA user_version = 6;")
            self.conn.commit()
            version = 6

        if version < 7:
            # Bank statement support + linking payments to bank transactions.
            self.conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS bank_transactions (
                    id INTEGER PRIMARY KEY,
                    bank_ts TEXT NOT NULL,
                    bank_date TEXT NOT NULL,
                    description TEXT,
                    amount_minor INTEGER NOT NULL,
                    raw_hash TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    UNIQUE(raw_hash)
                );

                CREATE INDEX IF NOT EXISTS idx_bank_tx_date_amount ON bank_transactions(bank_date, amount_minor);
                CREATE INDEX IF NOT EXISTS idx_bank_tx_amount ON bank_transactions(amount_minor);
                """
            )
            # payments.bank_tx_id is optional link to a bank transaction. Use a partial unique index.
            pcols = {r["name"] for r in self.conn.execute("PRAGMA table_info(payments);").fetchall()}
            if "bank_tx_id" not in pcols:
                self.conn.execute("ALTER TABLE payments ADD COLUMN bank_tx_id INTEGER;")
            # NOTE: payments can belong to a batch that matches a single bank transaction,
            # so bank_tx_id is NOT unique on payments. (Claiming is handled via bank_tx_claims.)
            self.conn.commit()
            self.conn.execute("PRAGMA user_version = 7;")
            self.conn.commit()
            version = 7

        if version < 8:
            # Bank transaction claiming (one bank transaction can cover a whole CSV batch).
            self.conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS bank_tx_claims (
                    id INTEGER PRIMARY KEY,
                    bank_tx_id INTEGER NOT NULL,
                    source_file TEXT,
                    note TEXT,
                    claimed_at TEXT NOT NULL,
                    UNIQUE(bank_tx_id),
                    FOREIGN KEY (bank_tx_id) REFERENCES bank_transactions(id)
                );
                CREATE INDEX IF NOT EXISTS idx_payments_bank_tx_id ON payments(bank_tx_id);
                """
            )
            # Drop old unique index if it exists from a previous build.
            try:
                self.conn.execute("DROP INDEX IF EXISTS uq_payments_bank_tx_id;")
            except Exception:
                pass
            self.conn.commit()
            self.conn.execute("PRAGMA user_version = 8;")
            self.conn.commit()

    def upsert_shipment(self, shipment: "Shipment") -> str:
        with self._lock:
            existing = self.conn.execute(
                "SELECT id FROM shipments WHERE carrier=? AND tracking=?;",
                (shipment.carrier, shipment.tracking),
            ).fetchone()

            if existing is None:
                self.conn.execute(
                    """
                    INSERT INTO shipments(
                        carrier, tracking, status_text, recipient_name, recipient_phone,
                        pickup_date, cod_amount_minor, lifecycle_state, created_at, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
                    """,
                    (
                        shipment.carrier,
                        shipment.tracking,
                        shipment.status_text,
                        shipment.recipient_name,
                        shipment.recipient_phone,
                        shipment.pickup_date,
                        shipment.cod_amount_minor,
                        shipment.lifecycle_state,
                        shipment.created_at,
                        shipment.updated_at,
                    ),
                )
                return "inserted"

            self.conn.execute(
                """
                UPDATE shipments
                SET status_text=?,
                    recipient_name=?,
                    recipient_phone=?,
                    pickup_date=?,
                    cod_amount_minor=?,
                    lifecycle_state=?,
                    updated_at=?
                WHERE carrier=? AND tracking=?;
                """,
                (
                    shipment.status_text,
                    shipment.recipient_name,
                    shipment.recipient_phone,
                    shipment.pickup_date,
                    shipment.cod_amount_minor,
                    shipment.lifecycle_state,
                    shipment.updated_at,
                    shipment.carrier,
                    shipment.tracking,
                ),
            )
            return "updated"

    def insert_shipment_new_only(self, shipment: "Shipment") -> str:
        with self._lock:
            try:
                self.conn.execute(
                    """
                    INSERT INTO shipments(
                        carrier, tracking, status_text, recipient_name, recipient_phone,
                        pickup_date, cod_amount_minor, lifecycle_state, created_at, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
                    """,
                    (
                        shipment.carrier,
                        shipment.tracking,
                        shipment.status_text,
                        shipment.recipient_name,
                        shipment.recipient_phone,
                        shipment.pickup_date,
                        shipment.cod_amount_minor,
                        shipment.lifecycle_state,
                        shipment.created_at,
                        shipment.updated_at,
                    ),
                )
                return "inserted"
            except sqlite3.IntegrityError:
                return "duplicate"

    def insert_payment(self, payment: "Payment") -> str:
        with self._lock:
            try:
                self.conn.execute(
                    """
                    INSERT INTO payments(
                        carrier, tracking, payer_name, paid_amount_minor, paid_date, raw_hash, created_at, bank_tx_id
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?);
                    """,
                    (
                        payment.carrier,
                        payment.tracking,
                        payment.payer_name,
                        payment.paid_amount_minor,
                        payment.paid_date,
                        payment.raw_hash,
                        payment.created_at,
                        getattr(payment, "bank_tx_id", None),
                    ),
                )
                return "inserted"
            except sqlite3.IntegrityError:
                return "duplicate"

    def upsert_shipment_controlled(
        self,
        shipment: "Shipment",
        *,
        file_pickup_date: str | None,
        file_delivery_date: str | None,
    ) -> tuple[str, tuple[str, str, str] | None]:
        """
        Returns (action, pickup_conflict) where:
          - action is "inserted" or "updated"
          - pickup_conflict is (tracking, old_pickup, new_pickup) when file_pickup differs from existing non-empty pickup_date
        """
        pickup_conflict: tuple[str, str, str] | None = None
        auto_remove = should_auto_remove_from_claims_by_status(
            shipment.status_text
        ) or should_auto_remove_from_claims_by_recipient(shipment.recipient_name)
        remove_reason, remove_note = auto_remove_reason_note(shipment.status_text, shipment.recipient_name)
        with self._lock:
            existing = self.conn.execute(
                """
                SELECT id, pickup_date, delivery_date, lifecycle_state, removed_at, removed_reason, removed_note
                FROM shipments
                WHERE carrier=? AND tracking=?;
                """,
                (shipment.carrier, shipment.tracking),
            ).fetchone()

            if existing is None:
                lifecycle_state = "UNDELIVERED_REMOVED" if auto_remove else shipment.lifecycle_state
                removed_at = shipment.updated_at if auto_remove else None
                removed_reason = remove_reason if auto_remove else None
                removed_note = remove_note if auto_remove else None
                self.conn.execute(
                    """
                    INSERT INTO shipments(
                        carrier, tracking, status_text, recipient_name, recipient_phone,
                        pickup_date, cod_amount_minor, lifecycle_state, created_at, updated_at,
                        status_updated_at, delivery_date,
                        removed_at, removed_reason, removed_note
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
                    """,
                    (
                        shipment.carrier,
                        shipment.tracking,
                        shipment.status_text,
                        shipment.recipient_name,
                        shipment.recipient_phone,
                        shipment.pickup_date,
                        shipment.cod_amount_minor,
                        lifecycle_state,
                        shipment.created_at,
                        shipment.updated_at,
                        shipment.status_updated_at,
                        shipment.delivery_date,
                        removed_at,
                        removed_reason,
                        removed_note,
                    ),
                )
                return "inserted", None

            shipment_id = int(existing["id"])
            existing_pickup = (existing["pickup_date"] or "").strip()
            set_pickup = None
            if not existing_pickup:
                set_pickup = file_pickup_date
            else:
                if file_pickup_date and file_pickup_date != existing_pickup:
                    pickup_conflict = (shipment.tracking, existing_pickup, file_pickup_date)

            delivery_to_set = None
            if file_delivery_date:
                delivery_to_set = file_delivery_date

            if delivery_to_set is not None and set_pickup is not None:
                self.conn.execute(
                    """
                    UPDATE shipments
                    SET status_text=?,
                        status_updated_at=?,
                        delivery_date=?,
                        pickup_date=?,
                        updated_at=?
                    WHERE id=?;
                    """,
                    (
                        shipment.status_text,
                        shipment.status_updated_at,
                        delivery_to_set,
                        set_pickup,
                        shipment.updated_at,
                        shipment_id,
                    ),
                )
            elif delivery_to_set is not None:
                self.conn.execute(
                    """
                    UPDATE shipments
                    SET status_text=?,
                        status_updated_at=?,
                        delivery_date=?,
                        updated_at=?
                    WHERE id=?;
                    """,
                    (
                        shipment.status_text,
                        shipment.status_updated_at,
                        delivery_to_set,
                        shipment.updated_at,
                        shipment_id,
                    ),
                )
            elif set_pickup is not None:
                self.conn.execute(
                    """
                    UPDATE shipments
                    SET status_text=?,
                        status_updated_at=?,
                        pickup_date=?,
                        updated_at=?
                    WHERE id=?;
                    """,
                    (
                        shipment.status_text,
                        shipment.status_updated_at,
                        set_pickup,
                        shipment.updated_at,
                        shipment_id,
                    ),
                )
            else:
                self.conn.execute(
                    """
                    UPDATE shipments
                    SET status_text=?,
                        status_updated_at=?,
                        updated_at=?
                    WHERE id=?;
                    """,
                    (
                        shipment.status_text,
                        shipment.status_updated_at,
                        shipment.updated_at,
                        shipment_id,
                    ),
                )

            if auto_remove and str(existing["lifecycle_state"] or "") != "UNDELIVERED_REMOVED":
                self.conn.execute(
                    """
                    UPDATE shipments
                    SET lifecycle_state='UNDELIVERED_REMOVED',
                        removed_at=COALESCE(NULLIF(removed_at, ''), ?),
                        removed_reason=COALESCE(NULLIF(removed_reason, ''), ?),
                        removed_note=COALESCE(NULLIF(removed_note, ''), ?),
                        updated_at=?
                    WHERE id=?;
                    """,
                    (
                        shipment.updated_at,
                        remove_reason,
                        remove_note,
                        shipment.updated_at,
                        shipment_id,
                    ),
                )

            return "updated", pickup_conflict

    def auto_remove_shipments_by_status(self, carrier: str) -> int:
        """
        Moves shipments to UNDELIVERED_REMOVED based on XLSX-imported fields:
          - status_text indicates "Pošiljka uručena/vraćena pošiljaocu"
          - recipient_name contains "Luxehair"

        This is safe to run repeatedly (idempotent) and does not overwrite existing removal info.
        """
        ts = now_utc_iso()
        changed = 0
        with self._lock:
            rows = self.conn.execute(
                """
                SELECT id, tracking, status_text, recipient_name
                FROM shipments
                WHERE carrier=?
                  AND lifecycle_state <> 'UNDELIVERED_REMOVED'
                  AND (status_text IS NOT NULL OR recipient_name IS NOT NULL);
                """,
                (carrier,),
            ).fetchall()
            for r in rows:
                status_text = r["status_text"]
                recipient_name = r["recipient_name"]
                if not (
                    should_auto_remove_from_claims_by_status(status_text)
                    or should_auto_remove_from_claims_by_recipient(recipient_name)
                ):
                    continue
                reason, note = auto_remove_reason_note(status_text, recipient_name)
                self.conn.execute(
                    """
                    UPDATE shipments
                    SET lifecycle_state='UNDELIVERED_REMOVED',
                        removed_at=COALESCE(NULLIF(removed_at, ''), ?),
                        removed_reason=COALESCE(NULLIF(removed_reason, ''), ?),
                        removed_note=COALESCE(NULLIF(removed_note, ''), ?),
                        updated_at=?
                    WHERE id=?;
                    """,
                    (ts, reason, note, ts, int(r["id"])),
                )
                changed += 1
            self.conn.commit()
        return changed

    def commit(self) -> None:
        with self._lock:
            self.conn.commit()

    def run_matching(self, carrier: str) -> dict[str, int]:
        matched_at = now_utc_iso()
        stats = {"shipments": 0, "OPEN": 0, "PARTIAL": 0, "PAID": 0, "MISMATCH": 0, "unmatched_payments": 0}

        with self._lock:
            shipment_rows = self.conn.execute(
                """
                SELECT
                    s.id AS shipment_id,
                    s.tracking AS tracking,
                    s.cod_amount_minor AS cod_amount_minor,
                    COALESCE(p.paid_total_minor, 0) AS paid_total_minor,
                    COALESCE(p.payments_count, 0) AS payments_count
                FROM shipments s
                LEFT JOIN (
                    SELECT carrier, tracking,
                           SUM(paid_amount_minor) AS paid_total_minor,
                           COUNT(*) AS payments_count
                    FROM payments
                    WHERE carrier=?
                    GROUP BY carrier, tracking
                ) p
                ON p.carrier=s.carrier AND p.tracking=s.tracking
                WHERE s.carrier=?
                  AND s.lifecycle_state <> 'UNDELIVERED_REMOVED';
                """,
                (carrier, carrier),
            ).fetchall()

            for row in shipment_rows:
                shipment_id = int(row["shipment_id"])
                cod_amount_minor = row["cod_amount_minor"]
                paid_total_minor = int(row["paid_total_minor"])
                payments_count = int(row["payments_count"])

                lifecycle_state, match_type, difference_minor = compute_lifecycle_and_settlement(
                    cod_amount_minor, paid_total_minor, payments_count
                )

                self.conn.execute(
                    """
                    INSERT INTO settlements(
                        shipment_id, paid_total_minor, difference_minor, match_type, last_matched_at
                    )
                    VALUES (?, ?, ?, ?, ?)
                    ON CONFLICT(shipment_id) DO UPDATE SET
                        paid_total_minor=excluded.paid_total_minor,
                        difference_minor=excluded.difference_minor,
                        match_type=excluded.match_type,
                        last_matched_at=excluded.last_matched_at;
                    """,
                    (shipment_id, paid_total_minor, difference_minor, match_type, matched_at),
                )

                self.conn.execute(
                    "UPDATE shipments SET lifecycle_state=?, updated_at=? WHERE id=?;",
                    (lifecycle_state, matched_at, shipment_id),
                )

                stats["shipments"] += 1
                stats[lifecycle_state] += 1

            stats["unmatched_payments"] = int(
                self.conn.execute(
                    """
                    SELECT COUNT(*)
                    FROM payments p
                    LEFT JOIN shipments s
                      ON s.carrier=p.carrier AND s.tracking=p.tracking
                    WHERE p.carrier=?
                      AND s.id IS NULL;
                    """,
                    (carrier,),
                ).fetchone()[0]
            )

            self.conn.commit()

        return stats

    def fetch_matching_overview(self, carrier: str, limit: int = 2000) -> list[sqlite3.Row]:
        with self._lock:
            return self.conn.execute(
                """
                SELECT
                    s.tracking AS tracking,
                    s.cod_amount_minor AS cod_amount_minor,
                    COALESCE(se.paid_total_minor, 0) AS paid_total_minor,
                    se.difference_minor AS difference_minor,
                    s.lifecycle_state AS lifecycle_state,
                    COALESCE(pc.payments_count, 0) AS payments_count
                FROM shipments s
                LEFT JOIN settlements se ON se.shipment_id=s.id
                LEFT JOIN (
                    SELECT carrier, tracking, COUNT(*) AS payments_count
                    FROM payments
                    WHERE carrier=?
                    GROUP BY carrier, tracking
                ) pc ON pc.carrier=s.carrier AND pc.tracking=s.tracking
                WHERE s.carrier=?
                  AND s.lifecycle_state <> 'UNDELIVERED_REMOVED'
                ORDER BY s.tracking
                LIMIT ?;
                """,
                (carrier, carrier, int(limit)),
            ).fetchall()

    def fetch_unmatched_payments(self, carrier: str, limit: int = 2000) -> list[sqlite3.Row]:
        with self._lock:
            return self.conn.execute(
                """
                SELECT
                    p.tracking AS tracking,
                    p.payer_name AS payer_name,
                    p.paid_amount_minor AS paid_amount_minor,
                    p.paid_date AS paid_date,
                    p.created_at AS created_at
                FROM payments p
                LEFT JOIN shipments s
                  ON s.carrier=p.carrier AND s.tracking=p.tracking
                WHERE p.carrier=?
                  AND s.id IS NULL
                ORDER BY p.id DESC
                LIMIT ?;
                """,
                (carrier, int(limit)),
            ).fetchall()

    def fetch_payments_for_tracking(self, carrier: str, tracking: str) -> list[sqlite3.Row]:
        with self._lock:
            return self.conn.execute(
                """
                SELECT id, carrier, tracking, payer_name, paid_amount_minor, paid_date, created_at
                FROM payments
                WHERE carrier=? AND tracking=?
                ORDER BY
                    CASE WHEN paid_date IS NULL OR TRIM(paid_date)='' THEN 1 ELSE 0 END,
                    paid_date DESC,
                    id DESC;
                """,
                (carrier, tracking),
            ).fetchall()

    def update_payment_amount_minor(self, payment_id: int, new_minor: int) -> None:
        with self._lock:
            self.conn.execute(
                "UPDATE payments SET paid_amount_minor=? WHERE id=?;",
                (int(new_minor), int(payment_id)),
            )
            self.conn.commit()

    def insert_bank_transaction(
        self,
        *,
        bank_ts: str,
        bank_date: str,
        description: str | None,
        amount_minor: int,
        raw_hash: str,
        created_at: str,
    ) -> str:
        with self._lock:
            try:
                self.conn.execute(
                    """
                    INSERT INTO bank_transactions(bank_ts, bank_date, description, amount_minor, raw_hash, created_at)
                    VALUES (?, ?, ?, ?, ?, ?);
                    """,
                    (bank_ts, bank_date, description, int(amount_minor), raw_hash, created_at),
                )
                return "inserted"
            except sqlite3.IntegrityError:
                return "duplicate"

    def claim_bank_tx(self, *, bank_tx_id: int, source_file: str | None, note: str | None) -> str:
        ts = now_utc_iso()
        with self._lock:
            try:
                self.conn.execute(
                    """
                    INSERT INTO bank_tx_claims(bank_tx_id, source_file, note, claimed_at)
                    VALUES (?, ?, ?, ?);
                    """,
                    (int(bank_tx_id), source_file, note, ts),
                )
                self.conn.commit()
                return "inserted"
            except sqlite3.IntegrityError:
                return "duplicate"

    def fetch_unmatched_bank_transactions(self, limit: int = 2000) -> list[sqlite3.Row]:
        with self._lock:
            return self.conn.execute(
                """
                SELECT
                    b.bank_date AS bank_date,
                    b.description AS description,
                    b.amount_minor AS amount_minor
                FROM bank_transactions b
                LEFT JOIN bank_tx_claims c ON c.bank_tx_id=b.id
                WHERE c.id IS NULL
                ORDER BY b.bank_date DESC, b.id DESC
                LIMIT ?;
                """,
                (int(limit),),
            ).fetchall()

    def fetch_all_bank_transactions_with_claims(self, limit: int = 5000) -> list[sqlite3.Row]:
        with self._lock:
            return self.conn.execute(
                """
                SELECT
                    b.bank_date AS bank_date,
                    b.bank_ts AS bank_ts,
                    b.description AS description,
                    b.amount_minor AS amount_minor,
                    c.source_file AS source_file,
                    c.claimed_at AS claimed_at
                FROM bank_transactions b
                LEFT JOIN bank_tx_claims c ON c.bank_tx_id=b.id
                ORDER BY b.bank_date DESC, b.id DESC
                LIMIT ?;
                """,
                (int(limit),),
            ).fetchall()

    def has_bank_transactions(self) -> bool:
        with self._lock:
            try:
                c = int(self.conn.execute("SELECT COUNT(*) FROM bank_transactions;").fetchone()[0])
                return c > 0
            except Exception:
                return False

    def find_unlinked_bank_tx_in_window(
        self, *, amount_minor: int, start_date_iso: str, end_date_iso: str, limit: int = 10
    ) -> list[sqlite3.Row]:
        with self._lock:
            return self.conn.execute(
                """
                SELECT b.id, b.bank_date, b.bank_ts, b.description
                FROM bank_transactions b
                LEFT JOIN bank_tx_claims c ON c.bank_tx_id=b.id
                WHERE c.id IS NULL
                  AND b.amount_minor=?
                  AND b.bank_date >= ? AND b.bank_date <= ?
                ORDER BY b.bank_date ASC, b.id ASC
                LIMIT ?;
                """,
                (int(amount_minor), start_date_iso, end_date_iso, int(limit)),
            ).fetchall()

    def find_unlinked_bank_tx_by_amount(self, *, amount_minor: int, limit: int = 10) -> list[sqlite3.Row]:
        with self._lock:
            return self.conn.execute(
                """
                SELECT b.id, b.bank_date, b.bank_ts, b.description
                FROM bank_transactions b
                LEFT JOIN bank_tx_claims c ON c.bank_tx_id=b.id
                WHERE c.id IS NULL
                  AND b.amount_minor=?
                ORDER BY b.bank_date ASC, b.id ASC
                LIMIT ?;
                """,
                (int(amount_minor), int(limit)),
            ).fetchall()

    def audit(self, action: str, target: str, details: str | None = None, actor: str = "local") -> None:
        with self._lock:
            self.conn.execute(
                "INSERT INTO audit_log(ts, actor, action, target, details) VALUES (?, ?, ?, ?, ?);",
                (now_utc_iso(), actor, action, target, details),
            )
            self.conn.commit()

    def get_or_create_cashbook_month(self, year: int, month: int) -> sqlite3.Row:
        created_at = now_utc_iso()
        with self._lock:
            existing = self.conn.execute(
                "SELECT * FROM cashbook_months WHERE year=? AND month=?;",
                (int(year), int(month)),
            ).fetchone()
            if existing is not None:
                return existing

            prev_year, prev_month = (year - 1, 12) if month == 1 else (year, month - 1)
            prev = self.conn.execute(
                "SELECT * FROM cashbook_months WHERE year=? AND month=?;",
                (int(prev_year), int(prev_month)),
            ).fetchone()
            opening = 0
            if prev is not None and prev["status"] == "CLOSED":
                opening = int(prev["closing_balance_minor"])

            self.conn.execute(
                """
                INSERT INTO cashbook_months(
                    year, month, opening_balance_minor,
                    total_in_minor, total_out_minor, closing_balance_minor,
                    status, closed_at, created_at, updated_at
                ) VALUES (?, ?, ?, 0, 0, ?, 'OPEN', NULL, ?, ?);
                """,
                (int(year), int(month), int(opening), int(opening), created_at, created_at),
            )
            self.conn.commit()
            return self.conn.execute(
                "SELECT * FROM cashbook_months WHERE year=? AND month=?;",
                (int(year), int(month)),
            ).fetchone()

    def fetch_cashbook_entries(self, cashbook_month_id: int) -> list[sqlite3.Row]:
        with self._lock:
            return self.conn.execute(
                """
                SELECT rb, entry_date, description, in_minor, out_minor, source_type, source_ref
                FROM cashbook_entries
                WHERE cashbook_month_id=?
                ORDER BY rb ASC;
                """,
                (int(cashbook_month_id),),
            ).fetchall()

    def _update_cashbook_totals(self, cashbook_month_id: int) -> sqlite3.Row:
        row = self.conn.execute(
            """
            SELECT
              opening_balance_minor,
              COALESCE(SUM(in_minor), 0) AS total_in_minor,
              COALESCE(SUM(out_minor), 0) AS total_out_minor
            FROM cashbook_months m
            LEFT JOIN cashbook_entries e ON e.cashbook_month_id=m.id
            WHERE m.id=?
            GROUP BY opening_balance_minor;
            """,
            (int(cashbook_month_id),),
        ).fetchone()
        opening = int(row["opening_balance_minor"])
        total_in = int(row["total_in_minor"])
        total_out = int(row["total_out_minor"])
        closing = opening + total_in - total_out
        ts = now_utc_iso()
        self.conn.execute(
            """
            UPDATE cashbook_months
            SET total_in_minor=?,
                total_out_minor=?,
                closing_balance_minor=?,
                updated_at=?
            WHERE id=?;
            """,
            (total_in, total_out, closing, ts, int(cashbook_month_id)),
        )
        return self.conn.execute("SELECT * FROM cashbook_months WHERE id=?;", (int(cashbook_month_id),)).fetchone()

    def set_cashbook_opening(self, cashbook_month_id: int, opening_balance_minor: int) -> sqlite3.Row:
        ts = now_utc_iso()
        with self._lock:
            self.conn.execute(
                "UPDATE cashbook_months SET opening_balance_minor=?, updated_at=? WHERE id=?;",
                (int(opening_balance_minor), ts, int(cashbook_month_id)),
            )
            updated = self._update_cashbook_totals(int(cashbook_month_id))
            self.conn.commit()
            return updated

    def rebuild_cashbook_entries(self, year: int, month: int, *, allow_closed: bool = False) -> sqlite3.Row:
        with self._lock:
            month_row = self.get_or_create_cashbook_month(int(year), int(month))
            if month_row["status"] == "CLOSED" and not allow_closed:
                raise RuntimeError("Mjesec je zatvoren. Potrebno je otključavanje.")

            month_id = int(month_row["id"])
            start, next_start = month_bounds(int(year), int(month))
            ts = now_utc_iso()

            # Preserve manual and storno entries, rebuild only generated rows.
            preserved = self.conn.execute(
                """
                SELECT entry_date, description, in_minor, out_minor, source_type, source_ref, created_at
                FROM cashbook_entries
                WHERE cashbook_month_id=?
                  AND source_type IN ('MANUAL','STORNO_COD');
                """,
                (month_id,),
            ).fetchall()

            self.conn.execute("DELETE FROM cashbook_entries WHERE cashbook_month_id=?;", (month_id,))

            shipments_by_day = self.conn.execute(
                """
                SELECT pickup_date AS day, SUM(COALESCE(cod_amount_minor, 0)) AS total_minor
                FROM shipments
                WHERE pickup_date >= ? AND pickup_date < ?
                GROUP BY pickup_date
                ORDER BY pickup_date ASC;
                """,
                (start, next_start),
            ).fetchall()

            payments_by_day = self.conn.execute(
                """
                SELECT paid_date AS day, SUM(COALESCE(paid_amount_minor, 0)) AS total_minor
                FROM payments
                WHERE paid_date >= ? AND paid_date < ?
                GROUP BY paid_date
                ORDER BY paid_date ASC;
                """,
                (start, next_start),
            ).fetchall()

            entries: list[tuple[str, str, int, int, str, str | None, str]] = []
            for r in shipments_by_day:
                day = r["day"]
                total_minor = int(r["total_minor"] or 0)
                if total_minor == 0:
                    continue
                entries.append(
                    (
                        day,
                        f"Ulaz BHP pošiljke {day}",
                        total_minor,
                        0,
                        "SHIPMENTS_IN",
                        day,
                        ts,
                    )
                )
            for r in payments_by_day:
                day = r["day"]
                total_minor = int(r["total_minor"] or 0)
                if total_minor == 0:
                    continue
                entries.append(
                    (
                        day,
                        f"Izlaz BHP uplate {day}",
                        0,
                        total_minor,
                        "PAYMENTS_OUT",
                        day,
                        ts,
                    )
                )

            for r in preserved:
                entries.append(
                    (
                        str(r["entry_date"]),
                        str(r["description"] or ""),
                        int(r["in_minor"] or 0),
                        int(r["out_minor"] or 0),
                        str(r["source_type"]),
                        r["source_ref"],
                        str(r["created_at"] or ts),
                    )
                )

            order = {"SHIPMENTS_IN": 0, "STORNO_COD": 1, "PAYMENTS_OUT": 2, "MANUAL": 3}
            entries.sort(key=lambda x: (x[0], order.get(x[4], 9), x[1]))
            for rb, (entry_date, desc, in_minor, out_minor, source_type, source_ref, created_at) in enumerate(
                entries, start=1
            ):
                self.conn.execute(
                    """
                    INSERT INTO cashbook_entries(
                        cashbook_month_id, rb, entry_date, description, in_minor, out_minor,
                        source_type, source_ref, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?);
                    """,
                    (
                        month_id,
                        rb,
                        entry_date,
                        desc,
                        int(in_minor),
                        int(out_minor),
                        source_type,
                        source_ref,
                        created_at,
                    ),
                )

            updated_month = self._update_cashbook_totals(month_id)
            self.conn.commit()
            return updated_month

    def _storno_ref(self, carrier: str, tracking: str) -> str:
        return f"{carrier.strip().upper()}|{normalize_tracking(tracking) or tracking}"

    def storno_exists(self, carrier: str, tracking: str) -> bool:
        ref = self._storno_ref(carrier, tracking)
        with self._lock:
            row = self.conn.execute(
                """
                SELECT 1
                FROM cashbook_entries
                WHERE source_type='STORNO_COD' AND source_ref=?
                LIMIT 1;
                """,
                (ref,),
            ).fetchone()
            return row is not None

    def create_storno_cod(
        self,
        *,
        carrier: str,
        tracking: str,
        storno_date_iso: str,
        amount_minor: int,
        note: str,
        created_at: str | None = None,
        rebuild_month: bool = True,
    ) -> bool:
        """
        Create a single STORNO_COD cashbook entry linked to a tracking.
        Returns True if created, False if already exists.
        """
        tracking_n = normalize_tracking(tracking) or ""
        if not tracking_n:
            return False
        if parse_iso_date(storno_date_iso) is None:
            raise ValueError("Neispravan datum storna.")
        if int(amount_minor) <= 0:
            return False

        ref = self._storno_ref(carrier, tracking_n)
        with self._lock:
            if self.storno_exists(carrier, tracking_n):
                return False

            d = date.fromisoformat(storno_date_iso)
            month_row = self.get_or_create_cashbook_month(int(d.year), int(d.month))
            month_id = int(month_row["id"])
            ts = created_at or now_utc_iso()
            carrier_u = carrier.strip().upper() if carrier else CARRIER_BHP
            desc = f"Storno {carrier_u} pošiljke ({tracking_n})"
            if note:
                desc = f"{desc} – {note}"

            # Insert with provisional RB; rebuild will renumber.
            rb = int(
                self.conn.execute(
                    "SELECT COALESCE(MAX(rb), 0) + 1 FROM cashbook_entries WHERE cashbook_month_id=?;",
                    (month_id,),
                ).fetchone()[0]
            )
            self.conn.execute(
                """
                INSERT INTO cashbook_entries(
                    cashbook_month_id, rb, entry_date, description, in_minor, out_minor,
                    source_type, source_ref, created_at
                ) VALUES (?, ?, ?, ?, 0, ?, 'STORNO_COD', ?, ?);
                """,
                (month_id, rb, storno_date_iso, desc, int(amount_minor), ref, ts),
            )
            self.conn.execute(
                "INSERT INTO audit_log(ts, actor, action, target, details) VALUES (?, ?, ?, ?, ?);",
                (
                    ts,
                    "local",
                    "STORNO_COD",
                    tracking_n,
                    f"carrier={carrier_u}; date={storno_date_iso}; amount_minor={int(amount_minor)}; note={note}",
                ),
            )

            # Rebuild month to keep ordering/totals consistent. Allow rebuild even if month was CLOSED.
            self.conn.commit()

        if rebuild_month:
            # Rebuild outside nested lock.
            self.rebuild_cashbook_entries(int(d.year), int(d.month), allow_closed=True)
        return True

    def rebuild_cashbook_entries_safe(self, year: int, month: int) -> sqlite3.Row | None:
        """
        Rebuild cashbook entries only when it's safe to do so automatically:
          - month status is OPEN
          - no MANUAL entries exist (to avoid wiping user-entered rows)
        """
        with self._lock:
            month_row = self.get_or_create_cashbook_month(int(year), int(month))
            if str(month_row["status"]) != "OPEN":
                return None
            month_id = int(month_row["id"])
            manual_count = int(
                self.conn.execute(
                    "SELECT COUNT(*) FROM cashbook_entries WHERE cashbook_month_id=? AND source_type='MANUAL';",
                    (month_id,),
                ).fetchone()[0]
            )
            if manual_count > 0:
                return None
        # call outside lock nesting (rebuild does its own lock)
        return self.rebuild_cashbook_entries(int(year), int(month), allow_closed=False)

    def close_cashbook_month(self, year: int, month: int) -> tuple[sqlite3.Row, sqlite3.Row]:
        with self._lock:
            month_row = self.get_or_create_cashbook_month(int(year), int(month))
            if month_row["status"] == "CLOSED":
                raise RuntimeError("Mjesec je već zatvoren.")

            month_id = int(month_row["id"])
            updated = self._update_cashbook_totals(month_id)

            closed_at = now_utc_iso()
            self.conn.execute(
                "UPDATE cashbook_months SET status='CLOSED', closed_at=?, updated_at=? WHERE id=?;",
                (closed_at, closed_at, month_id),
            )
            closed_month = self.conn.execute("SELECT * FROM cashbook_months WHERE id=?;", (month_id,)).fetchone()

            next_year, next_month = (year + 1, 1) if month == 12 else (year, month + 1)
            next_existing = self.conn.execute(
                "SELECT * FROM cashbook_months WHERE year=? AND month=?;",
                (int(next_year), int(next_month)),
            ).fetchone()
            if next_existing is None:
                opening = int(closed_month["closing_balance_minor"])
                ts = now_utc_iso()
                self.conn.execute(
                    """
                    INSERT INTO cashbook_months(
                        year, month, opening_balance_minor,
                        total_in_minor, total_out_minor, closing_balance_minor,
                        status, closed_at, created_at, updated_at
                    ) VALUES (?, ?, ?, 0, 0, ?, 'OPEN', NULL, ?, ?);
                    """,
                    (int(next_year), int(next_month), opening, opening, ts, ts),
                )
            else:
                # Business rule: next month opening balance should follow closing of previous month.
                # Update only if next month is still OPEN (do not mutate CLOSED months automatically).
                if str(next_existing["status"]) == "OPEN":
                    opening = int(closed_month["closing_balance_minor"])
                    next_id = int(next_existing["id"])
                    if int(next_existing["opening_balance_minor"] or 0) != opening:
                        ts = now_utc_iso()
                        self.conn.execute(
                            "UPDATE cashbook_months SET opening_balance_minor=?, updated_at=? WHERE id=?;",
                            (opening, ts, next_id),
                        )
                        # Keep totals consistent with the new opening.
                        self._update_cashbook_totals(next_id)

            self.conn.commit()
            next_month_row = self.conn.execute(
                "SELECT * FROM cashbook_months WHERE year=? AND month=?;",
                (int(next_year), int(next_month)),
            ).fetchone()
            return closed_month, next_month_row

    def remove_shipment_from_claims(
        self, carrier: str, tracking: str, reason: str, note: str | None
    ) -> None:
        ts = now_utc_iso()
        with self._lock:
            row = self.conn.execute(
                """
                SELECT id, lifecycle_state
                FROM shipments
                WHERE carrier=? AND tracking=?;
                """,
                (carrier, tracking),
            ).fetchone()
            if row is None:
                raise RuntimeError("Pošiljka nije pronađena.")
            if str(row["lifecycle_state"]) == "PAID":
                raise RuntimeError("Uplaćena pošiljka se ne može skinuti iz potraživanja.")

            self.conn.execute(
                """
                UPDATE shipments
                SET lifecycle_state='UNDELIVERED_REMOVED',
                    removed_at=?,
                    removed_reason=?,
                    removed_note=?,
                    updated_at=?
                WHERE carrier=? AND tracking=?;
                """,
                (ts, reason, note, ts, carrier, tracking),
            )
            self.conn.execute(
                "INSERT INTO audit_log(ts, actor, action, target, details) VALUES (?, ?, ?, ?, ?);",
                (
                    ts,
                    "local",
                    "REMOVE_FROM_CLAIMS",
                    tracking,
                    f"reason={reason}; note={(note or '')}",
                ),
            )
            self.conn.commit()

    def fetch_shipments_operational(
        self,
        *,
        carrier: str,
        search: str | None,
        pickup_from: str | None,
        pickup_to: str | None,
        states: list[str] | None,
        limit: int = 2000,
    ) -> list[sqlite3.Row]:
        where = ["s.carrier=?"]
        params: list[object] = [carrier]

        if search:
            # Note: Some SQLite builds can behave case-sensitive for LIKE.
            # We store tracking and names mostly uppercased; support both raw and upper search.
            search_raw = str(search).strip()
            search_upper = search_raw.upper()
            tracking_norm = normalize_tracking(search_raw) or search_raw

            like_raw = f"%{search_raw}%"
            like_upper = f"%{search_upper}%"
            like_tracking_norm = f"%{tracking_norm}%"

            where.append(
                "("
                "s.tracking LIKE ? OR s.tracking LIKE ? OR s.tracking LIKE ? "
                "OR COALESCE(s.recipient_name,'') LIKE ? OR COALESCE(s.recipient_name,'') LIKE ?"
                ")"
            )
            params.extend([like_raw, like_upper, like_tracking_norm, like_raw, like_upper])

        if pickup_from:
            where.append("s.pickup_date >= ?")
            params.append(pickup_from)
        if pickup_to:
            where.append("s.pickup_date <= ?")
            params.append(pickup_to)

        if states:
            placeholders = ",".join(["?"] * len(states))
            where.append(f"s.lifecycle_state IN ({placeholders})")
            params.extend(states)

        sql = f"""
            SELECT
                s.tracking AS tracking,
                s.recipient_name AS recipient_name,
                s.recipient_phone AS recipient_phone,
                s.pickup_date AS pickup_date,
                s.cod_amount_minor AS cod_amount_minor,
                s.lifecycle_state AS lifecycle_state,
                s.removed_reason AS removed_reason,
                s.removed_at AS removed_at,
                COALESCE(se.paid_total_minor, 0) AS paid_total_minor,
                se.difference_minor AS difference_minor,
                COALESCE(pc.payments_count, 0) AS payments_count
            FROM shipments s
            LEFT JOIN settlements se ON se.shipment_id=s.id
            LEFT JOIN (
                SELECT carrier, tracking, COUNT(*) AS payments_count
                FROM payments
                WHERE carrier=?
                GROUP BY carrier, tracking
            ) pc ON pc.carrier=s.carrier AND pc.tracking=s.tracking
            WHERE {" AND ".join(where)}
            ORDER BY s.pickup_date DESC, s.tracking ASC
            LIMIT ?;
        """
        params = [carrier] + params + [int(limit)]

        with self._lock:
            return self.conn.execute(sql, params).fetchall()

    def fetch_latest_shipment_basic(self, carrier: str) -> sqlite3.Row | None:
        with self._lock:
            return self.conn.execute(
                """
                SELECT tracking, pickup_date, updated_at
                FROM shipments
                WHERE carrier=?
                ORDER BY pickup_date DESC, updated_at DESC, id DESC
                LIMIT 1;
                """,
                (carrier,),
            ).fetchone()

    def fetch_shipment_settlement_by_tracking(self, carrier: str, tracking: str) -> sqlite3.Row | None:
        tracking_n = normalize_tracking(tracking) or tracking
        with self._lock:
            return self.conn.execute(
                """
                SELECT
                    s.lifecycle_state AS lifecycle_state,
                    s.cod_amount_minor AS cod_amount_minor,
                    COALESCE(se.paid_total_minor, 0) AS paid_total_minor,
                    se.difference_minor AS difference_minor
                FROM shipments s
                LEFT JOIN settlements se ON se.shipment_id=s.id
                WHERE s.carrier=? AND s.tracking=?
                LIMIT 1;
                """,
                (carrier, tracking_n),
            ).fetchone()

    def fetch_last_cashbook_payments_out_date(self) -> str | None:
        with self._lock:
            row = self.conn.execute(
                "SELECT MAX(entry_date) AS d FROM cashbook_entries WHERE source_type='PAYMENTS_OUT';"
            ).fetchone()
            if row is None:
                return None
            d = row["d"]
            return str(d) if d else None

    def fetch_latest_bank_transaction(self) -> sqlite3.Row | None:
        with self._lock:
            return self.conn.execute(
                """
                SELECT bank_ts, bank_date, amount_minor, description
                FROM bank_transactions
                ORDER BY bank_ts DESC, id DESC
                LIMIT 1;
                """
            ).fetchone()

    def fetch_unpaid_shipments(
        self,
        *,
        carrier: str,
        pickup_from: str | None,
        pickup_to: str | None,
        limit: int = 2000,
    ) -> list[sqlite3.Row]:
        return self.fetch_shipments_operational(
            carrier=carrier,
            search=None,
            pickup_from=pickup_from,
            pickup_to=pickup_to,
            states=["OPEN", "PARTIAL"],
            limit=limit,
        )

    def fetch_removed_shipments(
        self,
        *,
        carrier: str,
        pickup_from: str | None = None,
        pickup_to: str | None = None,
        limit: int = 2000,
    ) -> list[sqlite3.Row]:
        return self.fetch_shipments_operational(
            carrier=carrier,
            search=None,
            pickup_from=pickup_from,
            pickup_to=pickup_to,
            states=["UNDELIVERED_REMOVED"],
            limit=limit,
        )

    def fetch_mismatch_shipments(
        self,
        *,
        carrier: str,
        pickup_from: str | None = None,
        pickup_to: str | None = None,
        limit: int = 2000,
    ) -> list[sqlite3.Row]:
        return self.fetch_shipments_operational(
            carrier=carrier,
            search=None,
            pickup_from=pickup_from,
            pickup_to=pickup_to,
            states=["MISMATCH"],
            limit=limit,
        )

    def fetch_unmatched_payments_by_paid_date(
        self, *, carrier: str, paid_from: str | None, paid_to: str | None, limit: int = 5000
    ) -> list[sqlite3.Row]:
        where = ["p.carrier=?"]
        params: list[object] = [carrier]
        if paid_from:
            where.append("p.paid_date >= ?")
            params.append(paid_from)
        if paid_to:
            where.append("p.paid_date <= ?")
            params.append(paid_to)

        sql = f"""
            SELECT
                p.tracking AS tracking,
                p.payer_name AS payer_name,
                p.paid_amount_minor AS paid_amount_minor,
                p.paid_date AS paid_date,
                p.created_at AS created_at
            FROM payments p
            LEFT JOIN shipments s
              ON s.carrier=p.carrier AND s.tracking=p.tracking
            WHERE {" AND ".join(where)}
              AND s.id IS NULL
            ORDER BY p.id DESC
            LIMIT ?;
        """
        params.append(int(limit))
        with self._lock:
            return self.conn.execute(sql, params).fetchall()

    def fetch_claims_month_summary(self, *, carrier: str, year: int, month: int) -> sqlite3.Row:
        start, next_start = month_bounds(int(year), int(month))
        with self._lock:
            return self.conn.execute(
                """
                WITH base AS (
                    SELECT
                        s.lifecycle_state AS lifecycle_state,
                        s.cod_amount_minor AS cod_amount_minor,
                        COALESCE(se.paid_total_minor, 0) AS paid_total_minor,
                        CASE
                            WHEN se.difference_minor IS NOT NULL THEN se.difference_minor
                            WHEN s.cod_amount_minor IS NULL THEN NULL
                            ELSE (COALESCE(s.cod_amount_minor, 0) - COALESCE(se.paid_total_minor, 0))
                        END AS difference_minor
                    FROM shipments s
                    LEFT JOIN settlements se ON se.shipment_id=s.id
                    WHERE s.carrier=?
                      AND s.pickup_date >= ? AND s.pickup_date < ?
                )
                SELECT
                    SUM(CASE WHEN lifecycle_state='OPEN' THEN 1 ELSE 0 END) AS open_count,
                    SUM(CASE WHEN lifecycle_state='PARTIAL' THEN 1 ELSE 0 END) AS partial_count,
                    COALESCE(SUM(CASE WHEN lifecycle_state IN ('OPEN','PARTIAL') THEN COALESCE(difference_minor,0) ELSE 0 END), 0) AS unpaid_minor
                FROM base;
                """,
                (carrier, start, next_start),
            ).fetchone()

    def fetch_dashboard_metrics(
        self, *, carrier: str, pickup_from: str | None, pickup_to: str | None
    ) -> sqlite3.Row:
        where = ["s.carrier=?"]
        params: list[object] = [carrier]
        if pickup_from:
            where.append("s.pickup_date >= ?")
            params.append(pickup_from)
        if pickup_to:
            where.append("s.pickup_date <= ?")
            params.append(pickup_to)

        sql = f"""
            WITH base AS (
                SELECT
                    s.id AS id,
                    s.lifecycle_state AS lifecycle_state,
                    s.cod_amount_minor AS cod_amount_minor,
                    COALESCE(se.paid_total_minor, 0) AS paid_total_minor,
                    CASE
                        WHEN se.difference_minor IS NOT NULL THEN se.difference_minor
                        WHEN s.cod_amount_minor IS NULL THEN NULL
                        ELSE (COALESCE(s.cod_amount_minor, 0) - COALESCE(se.paid_total_minor, 0))
                    END AS difference_minor
                FROM shipments s
                LEFT JOIN settlements se ON se.shipment_id=s.id
                WHERE {" AND ".join(where)}
            )
            SELECT
                COUNT(*) AS total_shipments,
                SUM(CASE WHEN lifecycle_state <> 'UNDELIVERED_REMOVED' THEN 1 ELSE 0 END) AS total_active,
                SUM(CASE WHEN lifecycle_state = 'PAID' THEN 1 ELSE 0 END) AS paid_count,
                SUM(CASE WHEN lifecycle_state = 'PARTIAL' THEN 1 ELSE 0 END) AS partial_count,
                SUM(CASE WHEN lifecycle_state = 'OPEN' THEN 1 ELSE 0 END) AS open_count,
                SUM(CASE WHEN lifecycle_state = 'MISMATCH' THEN 1 ELSE 0 END) AS mismatch_count,
                SUM(CASE WHEN lifecycle_state = 'UNDELIVERED_REMOVED' THEN 1 ELSE 0 END) AS removed_count,
                COALESCE(SUM(COALESCE(cod_amount_minor, 0)), 0) AS sum_cod_minor,
                COALESCE(SUM(COALESCE(paid_total_minor, 0)), 0) AS sum_paid_minor,
                COALESCE(SUM(CASE WHEN lifecycle_state IN ('OPEN','PARTIAL') THEN COALESCE(difference_minor, 0) ELSE 0 END), 0) AS sum_unpaid_minor
            FROM base;
        """
        with self._lock:
            return self.conn.execute(sql, params).fetchone()


@dataclass(frozen=True)
class Shipment:
    carrier: str
    tracking: str
    status_text: str | None
    status_updated_at: str
    recipient_name: str | None
    recipient_phone: str | None
    pickup_date: str | None
    delivery_date: str | None
    cod_amount_minor: int | None
    lifecycle_state: str
    created_at: str
    updated_at: str


@dataclass(frozen=True)
class Payment:
    carrier: str
    tracking: str
    payer_name: str | None
    paid_amount_minor: int
    paid_date: str | None
    raw_hash: str
    created_at: str
    bank_tx_id: int | None = None


def import_bank_statement_xlsx(db: Database, xlsx_path: Path, log: callable) -> tuple[int, int, int, str | None]:
    """
    Import a bank statement XLSX (folder Izvodi/). Expected columns:
      A: datetime (e.g. '29.01.2026 11:47:10')
      B: description
      E: amount (e.g. '2.205,17')
    Returns (inserted, duplicates, total_minor, max_bank_date_iso).
    """
    try:
        from openpyxl import load_workbook
    except Exception:
        raise RuntimeError("Nedostaje zavisnost: openpyxl. Instaliraj: pip install -r requirements.txt")

    created_at = now_utc_iso()
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    ws = wb.active

    inserted = 0
    duplicates = 0
    total_minor = 0
    max_bank_date: str | None = None

    log(f"[BANK] Početak uvoza: {xlsx_path}")
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row:
            continue
        bank_ts, bank_date = parse_bank_ts_to_iso(row[0] if len(row) > 0 else None)  # A
        desc = (str(row[1]).strip() if len(row) > 1 and row[1] is not None else None)  # B
        amount_minor = parse_bank_amount_to_minor(row[4] if len(row) > 4 else None)  # E
        if bank_ts is None or bank_date is None or amount_minor is None:
            continue

        raw = f"{bank_ts}|{bank_date}|{amount_minor}|{desc or ''}"
        raw_hash = hashlib.sha256(raw.encode("utf-8")).hexdigest()
        res = db.insert_bank_transaction(
            bank_ts=bank_ts,
            bank_date=bank_date,
            description=desc,
            amount_minor=int(amount_minor),
            raw_hash=raw_hash,
            created_at=created_at,
        )
        if res == "inserted":
            inserted += 1
            total_minor += int(amount_minor)
            if max_bank_date is None or bank_date > max_bank_date:
                max_bank_date = bank_date
        else:
            duplicates += 1

        if (inserted + duplicates) % 1000 == 0:
            log(f"[BANK] Obrađen red {idx}...")

    db.commit()
    log(f"[BANK] Završeno. novi={inserted} duplikati={duplicates} ukupno={format_minor_km(total_minor)} KM")
    return inserted, duplicates, total_minor, max_bank_date


def import_bhp_export_xlsx(db: Database, xlsx_path: Path, log: callable) -> tuple[int, int, int]:
    try:
        from openpyxl import load_workbook
    except Exception:
        raise RuntimeError("Nedostaje zavisnost: openpyxl. Instaliraj: pip install -r requirements.txt")

    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    ws = wb.active

    inserted = 0
    updated = 0
    missing_tracking = 0
    pickup_conflicts = 0
    pickup_conflict_examples: list[tuple[str, str, str]] = []
    storno_candidates: list[tuple[str, str, int]] = []
    storno_created = 0
    storno_exists = 0
    storno_invalid_date = 0
    storno_sample_dates: list[str] = []

    started = now_utc_iso()
    log(f"[XLSX] Početak uvoza: {xlsx_path}")

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        tracking = normalize_tracking(row[1] if len(row) > 1 else None)  # B
        if not tracking:
            missing_tracking += 1
            continue

        # Status can come from different export variants:
        # - legacy: column C
        # - newer: column W (more detailed status)
        status_c = (str(row[2]).strip() if len(row) > 2 and row[2] is not None else None)  # C
        status_w = (str(row[22]).strip() if len(row) > 22 and row[22] is not None else None)  # W
        status_text = status_w or status_c
        recipient_name = (str(row[4]).strip() if len(row) > 4 and row[4] is not None else None)  # E
        recipient_phone = (str(row[6]).strip() if len(row) > 6 and row[6] is not None else None)  # G
        file_pickup_date = parse_date_to_iso(row[9] if len(row) > 9 else None)  # J
        cod_amount_minor = parse_cod_major_to_minor(row[15] if len(row) > 15 else None)  # P
        file_delivery_date = infer_delivery_date_from_status(status_text)
        file_return_date = parse_date_to_iso(row[23] if len(row) > 23 else None)  # X (datum povrata)

        status_updated_at = now_utc_iso()
        shipment = Shipment(
            carrier=CARRIER_BHP,
            tracking=tracking,
            status_text=status_text,
            status_updated_at=status_updated_at,
            recipient_name=recipient_name,
            recipient_phone=recipient_phone,
            pickup_date=file_pickup_date,
            delivery_date=file_delivery_date,
            cod_amount_minor=cod_amount_minor,
            lifecycle_state="active",
            created_at=started,
            updated_at=now_utc_iso(),
        )

        result, conflict = db.upsert_shipment_controlled(
            shipment, file_pickup_date=file_pickup_date, file_delivery_date=file_delivery_date
        )
        if result == "inserted":
            inserted += 1
        else:
            updated += 1
        if conflict is not None:
            pickup_conflicts += 1
            if len(pickup_conflict_examples) < 10:
                pickup_conflict_examples.append(conflict)

        # Auto storno COD for returned-to-sender shipments (balances earlier ULAZ in cashbook).
        if (
            should_auto_remove_from_claims_by_status(status_text)
            and int(cod_amount_minor or 0) > 0
        ):
            if not file_return_date or parse_iso_date(file_return_date) is None:
                storno_invalid_date += 1
            else:
                # Avoid duplicates: one storno per tracking.
                if db.storno_exists(CARRIER_BHP, tracking):
                    storno_exists += 1
                else:
                    storno_candidates.append((tracking, file_return_date, int(cod_amount_minor or 0)))
                    if len(storno_sample_dates) < 5:
                        storno_sample_dates.append(f"{tracking}:{file_return_date}")

        if (inserted + updated + missing_tracking) % 500 == 0:
            log(f"[XLSX] Obrađen red {idx}...")

    db.commit()

    # Create stornos after shipment import (batch), then rebuild affected months once.
    storno_months: set[tuple[int, int]] = set()
    for tracking, storno_date_iso, amount_minor in storno_candidates:
        try:
            created = db.create_storno_cod(
                carrier=CARRIER_BHP,
                tracking=tracking,
                storno_date_iso=storno_date_iso,
                amount_minor=amount_minor,
                note="Auto storno – povrat pošiljke (BHP)",
                created_at=started,
                rebuild_month=False,
            )
            if created:
                storno_created += 1
                d = date.fromisoformat(storno_date_iso)
                storno_months.add((int(d.year), int(d.month)))
            else:
                storno_exists += 1
        except Exception:
            storno_invalid_date += 1

    for y, m in sorted(storno_months):
        try:
            db.rebuild_cashbook_entries(int(y), int(m), allow_closed=True)
        except Exception:
            pass

    log(
        f"[XLSX] Završeno. novi={inserted} ažurirani={updated} pickup_date_conflicts={pickup_conflicts} preskočeno_nema_tracking={missing_tracking}"
    )
    if pickup_conflict_examples:
        log("[XLSX] Primjeri pickup_date_conflicts (prvih 10):")
        for tracking_no, old_v, new_v in pickup_conflict_examples:
            log(f"  - {tracking_no}: baza={old_v} fajl={new_v}")
    if storno_created or storno_exists or storno_invalid_date:
        log(
            f"[STORNO] Auto storno: created={storno_created} skipped_exists={storno_exists} skipped_invalid_date={storno_invalid_date}"
        )
        if storno_sample_dates:
            log("[STORNO] Primjeri (do 5):")
            for s in storno_sample_dates:
                log(f"  - {s}")
    return inserted, updated, missing_tracking


def import_bhp_uplate_csv(
    db: Database,
    csv_path: Path,
    log: callable,
    *,
    override_paid_date_iso: str | None = None,
    bank_tx_id_for_file: int | None = None,
) -> tuple[int, int, int, set[tuple[int, int]]]:
    inserted = 0
    updated = 0
    skipped = 0
    affected_months: set[tuple[int, int]] = set()

    missing_tracking = 0
    invalid_amount = 0
    invalid_date = 0
    parse_error = 0
    duplicates = 0
    duplicates_by_tracking = 0
    duplicates_by_hash = 0
    examples: dict[str, list[str]] = {
        "missing_tracking": [],
        "invalid_amount": [],
        "invalid_date": [],
        "parse_error": [],
        "duplicate": [],
        "duplicate_tracking": [],
    }

    if override_paid_date_iso is not None and parse_iso_date(override_paid_date_iso) is None:
        raise RuntimeError(f"Neispravan override datum uplate: {override_paid_date_iso!r}")

    guessed = detect_csv_encoding(csv_path, log)
    encodings = [guessed] + [e for e in ("utf-8-sig", "cp1250", "iso-8859-2", "cp1252") if e != guessed]

    created_at = now_utc_iso()
    log(f"[CSV] Početak uvoza: {csv_path}")
    if override_paid_date_iso:
        log(f"[CSV] Datum uplate (override za cijeli fajl): {override_paid_date_iso}")

    def run_with_file(file_obj) -> None:
        nonlocal inserted, updated, skipped
        nonlocal missing_tracking, invalid_amount, invalid_date, parse_error, duplicates
        nonlocal duplicates_by_tracking, duplicates_by_hash
        nonlocal affected_months
        reader = csv.reader(file_obj, delimiter=";")
        for line_no, cols in enumerate(reader, start=1):
            if not cols:
                continue

            tracking, payer_name, paid_amount_minor, paid_date, err = parse_bhp_uplate_row(
                cols, override_paid_date_iso=override_paid_date_iso
            )
            if err:
                skipped += 1
                if err == "parse_error":
                    parse_error += 1
                    if len(examples["parse_error"]) < 10:
                        examples["parse_error"].append(f"linija={line_no} cols={len(cols)} raw={';'.join(cols)}")
                elif err == "missing_tracking":
                    missing_tracking += 1
                    if len(examples["missing_tracking"]) < 10:
                        examples["missing_tracking"].append(f"linija={line_no} raw={';'.join(cols)}")
                elif err == "invalid_amount":
                    invalid_amount += 1
                    if len(examples["invalid_amount"]) < 10:
                        examples["invalid_amount"].append(f"linija={line_no} tracking={tracking} raw={';'.join(cols)}")
                elif err == "invalid_date":
                    invalid_date += 1
                    if len(examples["invalid_date"]) < 10:
                        examples["invalid_date"].append(f"linija={line_no} tracking={tracking} raw={';'.join(cols)}")
                continue

            bank_tx_id: int | None = int(bank_tx_id_for_file) if bank_tx_id_for_file is not None else None
            final_paid_date = paid_date

            # Hard rule: one payment per tracking.
            # Allow inserting again ONLY if:
            #  - more than 12 months since last payment (paid_date)
            #  - AND payer_name differs
            with db._lock:
                existing = db.conn.execute(
                    """
                    SELECT payer_name, paid_date, created_at
                    FROM payments
                    WHERE carrier=? AND tracking=?
                    ORDER BY
                        CASE WHEN paid_date IS NULL OR TRIM(paid_date)='' THEN 1 ELSE 0 END,
                        paid_date DESC,
                        created_at DESC
                    LIMIT 1;
                    """,
                    (CARRIER_BHP, tracking),
                ).fetchone()

            if existing is not None:
                existing_paid_date = str(existing["paid_date"] or "").strip() or None
                allow_reinsert = _is_older_than_12_months(existing_paid_date, final_paid_date) and (
                    not _same_person_name(existing["payer_name"], payer_name)
                )
                if not allow_reinsert:
                    duplicates += 1
                    duplicates_by_tracking += 1
                    skipped += 1
                    if len(examples["duplicate_tracking"]) < 10:
                        examples["duplicate_tracking"].append(
                            f"linija={line_no} tracking={tracking} amount={paid_amount_minor} date={final_paid_date} payer={payer_name or ''} | existing_date={existing_paid_date or 'N/A'} existing_payer={existing['payer_name'] or ''}"
                        )
                    _append_duplicate_payment_log(
                        f"duplicate_tracking file={csv_path.name} linija={line_no} tracking={tracking} amount_minor={paid_amount_minor} paid_date={final_paid_date} payer={payer_name or ''} | existing_paid_date={existing_paid_date or 'N/A'} existing_payer={existing['payer_name'] or ''}"
                    )
                    continue

            raw_hash = compute_payment_row_hash(carrier=CARRIER_BHP, cols=cols)

            payment = Payment(
                carrier=CARRIER_BHP,
                tracking=tracking,
                payer_name=payer_name,
                paid_amount_minor=paid_amount_minor,
                paid_date=final_paid_date,
                raw_hash=raw_hash,
                created_at=created_at,
                bank_tx_id=bank_tx_id,
            )

            result = db.insert_payment(payment)
            if result == "inserted":
                inserted += 1
                d = parse_iso_date(final_paid_date)
                if d is not None:
                    affected_months.add((int(d.year), int(d.month)))
            else:
                duplicates += 1
                duplicates_by_hash += 1
                skipped += 1
                if len(examples["duplicate"]) < 10:
                    examples["duplicate"].append(f"tracking={tracking} amount={paid_amount_minor} date={final_paid_date}")
                _append_duplicate_payment_log(
                    f"duplicate_hash file={csv_path.name} linija={line_no} tracking={tracking} amount_minor={paid_amount_minor} paid_date={final_paid_date} payer={payer_name or ''}"
                )

            if (inserted + updated + skipped) % 1000 == 0:
                log(f"[CSV] Obrađena linija {line_no}...")

    decode_error: Exception | None = None
    used_encoding: str | None = None
    for enc in encodings:
        try:
            # Reset counters in case we retry due to decoding
            inserted = 0
            updated = 0
            skipped = 0
            missing_tracking = 0
            invalid_amount = 0
            invalid_date = 0
            parse_error = 0
            duplicates = 0
            duplicates_by_tracking = 0
            duplicates_by_hash = 0
            for k in examples:
                examples[k].clear()

            with db._lock:
                db.conn.execute("BEGIN;")

            with open(csv_path, "r", encoding=enc, errors="strict", newline="") as file_obj:
                log(f"[CSV] Koristim encoding: {enc}")
                run_with_file(file_obj)

            db.commit()
            used_encoding = enc
            decode_error = None
            break
        except UnicodeDecodeError as e:
            decode_error = e
            log(f"[CSV] Greška dekodiranja sa encoding={enc}: {e}. Pokušavam sljedeći encoding...")
            with db._lock:
                try:
                    db.conn.rollback()
                except Exception:
                    pass
            continue

    if decode_error is not None:
        raise decode_error

    log(
        "[CSV] Završeno. novo={ins} preskočeno={sk} (missing_tracking={mt} invalid_amount={ia} invalid_date={id} parse_error={pe} duplicate={du} duplicate_tracking={dt} duplicate_hash={dh})".format(
            ins=inserted,
            sk=skipped,
            mt=missing_tracking,
            ia=invalid_amount,
            id=invalid_date,
            pe=parse_error,
            du=duplicates,
            dt=duplicates_by_tracking,
            dh=duplicates_by_hash,
        )
    )
    for key, title in (
        ("missing_tracking", "Primjeri: missing_tracking"),
        ("invalid_amount", "Primjeri: invalid_amount"),
        ("invalid_date", "Primjeri: invalid_date"),
        ("parse_error", "Primjeri: parse_error"),
        ("duplicate", "Primjeri: duplicate"),
        ("duplicate_tracking", "Primjeri: duplicate_tracking"),
    ):
        if examples[key]:
            log(f"[CSV] {title} (prvih {len(examples[key])}):")
            for ex in examples[key]:
                log(f"  - {ex}")
    return inserted, updated, skipped, affected_months


class App:
    def __init__(self, root: Tk, db_path: Path) -> None:
        self.root = root
        self.root.title(APP_NAME)
        self.queue: Queue[str] = Queue()
        self.cashbook_unlocked: set[tuple[int, int]] = set()
        self._ops_rows: list[sqlite3.Row] = []
        self._unpaid_rows: list[sqlite3.Row] = []
        self._removed_rows: list[sqlite3.Row] = []
        self._ctrl_unmatched_rows: list[sqlite3.Row] = []
        self._ctrl_mismatch_rows: list[sqlite3.Row] = []
        self._ctrl_bank_unmatched_rows: list[sqlite3.Row] = []

        for folder in (DATA_DIR, LOGS_DIR):
            folder.mkdir(parents=True, exist_ok=True)

        self.settings = SettingsStore(SETTINGS_PATH)

        self.db = Database(db_path)

        notebook = ttk.Notebook(root)
        notebook.pack(fill=BOTH, expand=True)

        dashboard_tab = Frame(notebook)
        import_tab = Frame(notebook)
        matching_tab = Frame(notebook)
        cashbook_tab = Frame(notebook)
        ops_tab = Frame(notebook)
        unpaid_tab = Frame(notebook)
        removed_tab = Frame(notebook)
        control_tab = Frame(notebook)
        settings_tab = Frame(notebook)
        notebook.add(dashboard_tab, text="Pregled")
        notebook.add(import_tab, text="Uvoz")
        notebook.add(matching_tab, text="Pregled usklađivanja")
        notebook.add(cashbook_tab, text="Blagajna")
        notebook.add(ops_tab, text="Pošiljke / Potraživanja")
        notebook.add(unpaid_tab, text="Neuplaćene")
        notebook.add(removed_tab, text="Nepreuzete")
        notebook.add(control_tab, text="Kontrola")
        notebook.add(settings_tab, text="Podešavanja")

        toolbar = Frame(import_tab)
        toolbar.pack(fill="x")

        Button(toolbar, text="Uvoz BHP Export (XLSX)", command=self.on_import_xlsx).pack(
            side=LEFT, padx=6, pady=6
        )
        Button(toolbar, text="Uvoz BHP Uplate (CSV)", command=self.on_import_csv).pack(side=LEFT, padx=6, pady=6)
        Button(toolbar, text="Uvoz bank izvoda (XLSX)", command=self.on_import_bank_statement).pack(
            side=LEFT, padx=6, pady=6
        )
        info = ttk.Label(toolbar, text="(?)", foreground="#0b3d91")
        info.pack(side=LEFT, padx=(10, 0))
        Tooltip(
            info,
            "Redoslijed uvoza (preporučeno):\n"
            "1) Uvoz bank izvoda (XLSX)\n"
            "2) Uvoz BHP Uplate (CSV) – tada automatski povlači bank datum (batch match po fajlu, +10 dana)\n"
            "3) Uvoz BHP Export (XLSX) – pošiljke (može bilo kada)",
            wrap=520,
        )
        Button(toolbar, text="Globalno osvježi", command=self.global_refresh).pack(side=LEFT, padx=6, pady=6)
        Button(toolbar, text="Zatvori", command=self.request_exit).pack(side=RIGHT, padx=6, pady=6)

        self.log_box = ScrolledText(import_tab, height=18)
        self.log_box.pack(fill=BOTH, expand=True, padx=8, pady=8)

        self._build_matching_tab(matching_tab)
        self._build_cashbook_tab(cashbook_tab)
        self._build_dashboard_tab(dashboard_tab)
        self._build_ops_tab(ops_tab)
        self._build_unpaid_tab(unpaid_tab)
        self._build_removed_tab(removed_tab)
        self._build_control_tab(control_tab)
        self._build_settings_tab(settings_tab)

        self._start_log_pump()
        self.log(f"Baza: {self.db.path}")
        self.log("Spremno.")

        self.refresh_dashboard()

    def global_refresh(self) -> None:
        # Log immediately so the user knows the action started (even if background work takes time).
        self.log("[GLOBAL] Pokrenuto globalno osvježavanje…")

        def work():
            # 1) Auto-remove based on tracking events (live) for active shipments.
            moved_by_tracking = 0
            checked = 0
            skipped = 0
            carrier = CARRIER_BHP
            try:
                # Prefer currently loaded operational rows to stay deterministic and fast.
                candidates = []
                for r in getattr(self, "_ops_rows", []) or []:
                    try:
                        if str(r["lifecycle_state"]) in ("UNDELIVERED_REMOVED", "PAID"):
                            continue
                        candidates.append(str(r["tracking"]))
                    except Exception:
                        continue

                # If ops tab isn't loaded yet, fall back to a DB query (bounded).
                if not candidates:
                    rows = self.db.fetch_shipments_operational(
                        carrier=carrier,
                        search=None,
                        pickup_from=None,
                        pickup_to=None,
                        states=["OPEN", "PARTIAL", "MISMATCH"],
                        limit=200,
                    )
                    candidates = [str(r["tracking"]) for r in rows]

                candidates = [normalize_tracking(t) for t in candidates]
                candidates = [t for t in candidates if t]
                candidates = sorted(set(candidates))[:200]

                self.log(f"[TRACK] Provjera statusa po praćenju: {len(candidates)} pošiljki (limit=200)")

                for t in candidates:
                    checked += 1
                    try:
                        events = fetch_bhp_tracking_events(t, timeout_s=10, debug_log=None)
                        if should_auto_remove_from_claims_by_tracking_events(events):
                            try:
                                self.db.remove_shipment_from_claims(
                                    carrier, t, "Vraćeno pošiljaocu", "auto-tracking"
                                )
                                moved_by_tracking += 1
                            except Exception:
                                skipped += 1
                        if checked % 10 == 0:
                            self.log(
                                f"[TRACK] Napredak: {checked}/{len(candidates)} | prebačeno={moved_by_tracking} | preskočeno={skipped}"
                            )
                        time.sleep(0.15)
                    except TrackingCaptchaError:
                        self.log("[TRACK][WARN] reCAPTCHA – automatsko praćenje nije dostupno trenutno.")
                        break
                    except Exception:
                        skipped += 1

            except Exception as exc:
                self.log(f"[ERROR] tracking auto-remove: {exc}")
                self.log(traceback.format_exc().rstrip())

            # 2) Auto-remove based on XLSX status_text (when available).
            moved_by_status = 0
            try:
                moved_by_status = self.db.auto_remove_shipments_by_status(carrier)
            except Exception as exc:
                self.log(f"[ERROR] auto-remove (status): {exc}")
                self.log(traceback.format_exc().rstrip())

            try:
                stats = self.db.run_matching(carrier)
                self.log(
                    f"[MATCH] Pregled osvježen. Pošiljke={stats.get('shipments', 0)} | Neuparene uplate={stats.get('unmatched_payments', 0)}"
                )
            except Exception as exc:
                self.log(f"[ERROR] matching: {exc}")
                self.log(traceback.format_exc().rstrip())

            if checked:
                self.log(f"[TRACK] Provjera završena: checked={checked} moved={moved_by_tracking} skipped={skipped}")
            if moved_by_status:
                self.log(f"[OPS] Auto prebačeno u Nepreuzete (po statusu): {moved_by_status}")

            self.root.after(0, self._refresh_all_views)

        self._run_in_thread(work)

    def _confirm_exit(self) -> bool:
        win = Toplevel(self.root)
        apply_window_icon(win)
        win.title("Potvrda")
        win.transient(self.root)
        win.grab_set()
        win.resizable(False, False)

        result: dict[str, object] = {"ok": False}

        frm = Frame(win)
        frm.pack(fill=BOTH, expand=True, padx=14, pady=14)
        ttk.Label(frm, text="Da li želite završiti rad u aplikaciji?").pack(anchor="w", pady=(0, 10))

        btns = Frame(frm)
        btns.pack(anchor="e")

        def yes():
            result["ok"] = True
            win.destroy()

        def no():
            win.destroy()

        Button(btns, text="Ne", command=no).pack(side=RIGHT, padx=(6, 0))
        Button(btns, text="Da", command=yes).pack(side=RIGHT)

        win.bind("<Escape>", lambda _e: no())
        win.bind("<Return>", lambda _e: yes())
        center_window(win, self.root)
        win.focus_set()
        self.root.wait_window(win)
        return bool(result.get("ok"))

    def request_exit(self) -> None:
        if not self._confirm_exit():
            return
        try:
            self.db.close()
        except Exception:
            pass
        self.root.destroy()

    def _get_reset_password(self) -> str:
        value = self.settings.get("reset_password", DEFAULT_RESET_PASSWORD)
        return str(value) if value else DEFAULT_RESET_PASSWORD

    def _build_settings_tab(self, parent: Frame) -> None:
        frm = Frame(parent)
        frm.pack(fill=BOTH, expand=True, padx=12, pady=12)

        self.settings_db_path_var = StringVar(value=str(self.db.path))
        self.settings_password_var = StringVar(value="")

        ttk.Label(frm, text="Baza (SQLite fajl):").grid(row=0, column=0, sticky="w", pady=(0, 6))
        db_entry = ttk.Entry(frm, width=60, textvariable=self.settings_db_path_var)
        db_entry.grid(row=0, column=1, sticky="we", pady=(0, 6))

        def choose_db():
            filename = filedialog.asksaveasfilename(
                title="Odaberi bazu (SQLite fajl)",
                defaultextension=".sqlite3",
                initialfile=Path(self.settings_db_path_var.get() or "bhpx.sqlite3").name,
                filetypes=[("SQLite baza", "*.sqlite3"), ("SQLite baza", "*.db"), ("Svi fajlovi", "*.*")],
            )
            if filename:
                self.settings_db_path_var.set(filename)

        Button(frm, text="Odaberi...", command=choose_db).grid(row=0, column=2, padx=(8, 0), pady=(0, 6))

        def apply_db():
            raw = (self.settings_db_path_var.get() or "").strip()
            if not raw:
                messagebox.showerror(APP_NAME, "Odaberi fajl baze.")
                return
            new_path = Path(raw)

            def work():
                try:
                    self.db.close()
                except Exception:
                    pass
                self.db = Database(new_path)
                try:
                    self.settings.set("db_path", str(new_path))
                    self.settings.save()
                except Exception:
                    pass
                self.log(f"[SET] Baza postavljena: {new_path}")
                self.root.after(0, self._refresh_all_views)

            self._run_in_thread(work)

        Button(frm, text="Primijeni bazu", command=apply_db).grid(row=1, column=1, sticky="w", pady=(0, 12))

        ttk.Separator(frm).grid(row=2, column=0, columnspan=3, sticky="we", pady=(0, 12))

        ttk.Label(frm, text="Reset baze:").grid(row=3, column=0, sticky="w", pady=(0, 6))

        def reset_db():
            pwd = simpledialog.askstring(APP_NAME, "Unesi lozinku za reset baze:", show="*")
            if pwd is None:
                return
            if pwd != self._get_reset_password():
                messagebox.showerror(APP_NAME, "Pogrešna lozinka.")
                return
            if not messagebox.askyesno(APP_NAME, "Resetovati bazu? Ovo briše sve podatke."):
                return

            db_path = Path((self.settings_db_path_var.get() or "").strip() or str(self.db.path))

            def work():
                try:
                    try:
                        self.db.close()
                    except Exception:
                        pass

                    try:
                        if db_path.exists():
                            db_path.unlink()
                    except Exception:
                        # Fallback: drop tables in-place if the file can't be deleted.
                        conn = sqlite3.connect(db_path)
                        try:
                            conn.executescript(
                                """
                                DROP TABLE IF EXISTS cashbook_entries;
                                DROP TABLE IF EXISTS cashbook_months;
                                DROP TABLE IF EXISTS settlements;
                                DROP TABLE IF EXISTS payments;
                                DROP TABLE IF EXISTS shipments;
                                DROP TABLE IF EXISTS audit_log;
                                PRAGMA user_version = 0;
                                """
                            )
                            conn.commit()
                        finally:
                            conn.close()

                    self.db = Database(db_path)
                    try:
                        self.settings.set("db_path", str(db_path))
                        self.settings.save()
                    except Exception:
                        pass
                    self.log(f"[SET] Baza resetovana: {db_path}")
                    self.root.after(0, self._refresh_all_views)
                except Exception as exc:
                    self.log(f"[ERROR] reset baze: {exc}")
                    try:
                        msg = str(exc)
                        self.root.after(0, lambda m=msg: messagebox.showerror(APP_NAME, m))
                    except Exception:
                        pass

            self._run_in_thread(work)

        Button(frm, text="Reset baze", command=reset_db).grid(row=3, column=1, sticky="w", pady=(0, 6))

        ttk.Label(frm, text="Lozinka (za reset):").grid(row=4, column=0, sticky="w", pady=(6, 6))
        pwd_entry = ttk.Entry(frm, width=24, textvariable=self.settings_password_var, show="*")
        pwd_entry.grid(row=4, column=1, sticky="w", pady=(6, 6))

        def save_pwd():
            pwd = (self.settings_password_var.get() or "").strip()
            if not pwd:
                messagebox.showerror(APP_NAME, "Unesi lozinku.")
                return
            self.settings.set("reset_password", pwd)
            self.settings.save()
            self.settings_password_var.set("")
            self.log("[SET] Lozinka za reset je sačuvana.")
            messagebox.showinfo(APP_NAME, "Lozinka je sačuvana.")

        Button(frm, text="Snimi šifru", command=save_pwd).grid(row=4, column=2, padx=(8, 0), pady=(6, 6))

        frm.columnconfigure(1, weight=1)

    def _refresh_all_views(self) -> None:
        self.settings_db_path_var.set(str(self.db.path))
        self.log(f"Baza: {self.db.path}")
        try:
            self.refresh_matching_view()
        except Exception:
            pass
        try:
            self.refresh_cashbook_view()
        except Exception:
            pass
        try:
            self.refresh_ops_view()
        except Exception:
            pass
        try:
            self.refresh_unpaid_view()
        except Exception:
            pass
        try:
            self.refresh_removed_view()
        except Exception:
            pass
        try:
            self.refresh_control_unmatched()
        except Exception:
            pass
        try:
            self.refresh_control_mismatch()
        except Exception:
            pass
        try:
            self.refresh_control_bank_unmatched()
        except Exception:
            pass
        try:
            self.refresh_dashboard()
        except Exception:
            pass

    def _start_log_pump(self) -> None:
        def pump():
            try:
                while True:
                    msg = self.queue.get_nowait()
                    self.log_box.insert(END, msg + "\n")
                    self.log_box.see(END)
            except Empty:
                pass
            self.root.after(100, pump)

        self.root.after(100, pump)

    def log(self, msg: str) -> None:
        self.queue.put(msg)

    def _ask_yesno_blocking(self, title: str, message: str) -> bool:
        """
        Show a yes/no prompt on the Tk main thread and wait for the result.
        Safe to call from a worker thread.
        """
        result: dict[str, object] = {"value": False}
        done = threading.Event()

        def ask():
            try:
                result["value"] = bool(messagebox.askyesno(title, message, parent=self.root))
            finally:
                done.set()

        try:
            self.root.after(0, ask)
        except Exception:
            # If UI is not available, default to "No".
            return False
        done.wait()
        return bool(result.get("value"))

    def _run_in_thread(self, fn: callable) -> None:
        def runner():
            try:
                fn()
            except Exception as exc:
                self.log(f"[ERROR] {exc}")
                self.log(traceback.format_exc().rstrip())
                try:
                    msg = str(exc)
                    self.root.after(0, lambda m=msg: messagebox.showerror(APP_NAME, m))
                except Exception:
                    pass

        threading.Thread(target=runner, daemon=True).start()

    def on_import_xlsx(self) -> None:
        filename = filedialog.askopenfilename(
            title="Odaberi BHP Export (XLSX)",
            filetypes=[("Excel fajlovi", "*.xlsx"), ("Svi fajlovi", "*.*")],
        )
        if not filename:
            return
        path = Path(filename)

        def work():
            inserted, updated, missing_tracking = import_bhp_export_xlsx(self.db, path, self.log)
            self.log(f"[XLSX] Rezime: novi={inserted} | ažurirani={updated} | preskočeno_nema_tracking={missing_tracking}")
            try:
                self.root.after(
                    0,
                    lambda p=path: (
                        self.settings.set("last_shipments_import_file", p.name),
                        self.settings.set("last_shipments_import_ts", now_utc_iso()),
                        self.settings.save(),
                        self.refresh_dashboard(),
                    ),
                )
            except Exception:
                pass
            stats = self.db.run_matching(CARRIER_BHP)
            self.log(
                "[MATCH] Pošiljke={shipments} | Otvoreno={OPEN} | Djelimično={PARTIAL} | Uplaćeno={PAID} | Neusklađeno={MISMATCH} | Neuparene uplate={unmatched_payments}".format(
                    **stats
                )
            )
            self.root.after(0, self.refresh_matching_view)
            self.root.after(0, self.refresh_ops_view)
            self.root.after(0, self.refresh_unpaid_view)

        self._run_in_thread(work)

    def on_import_bank_statement(self) -> None:
        default_dir = ROOT_DIR / "Izvodi"
        initialdir = str(default_dir) if default_dir.exists() else None
        filename = filedialog.askopenfilename(
            title="Odaberi bank izvod (XLSX)",
            initialdir=initialdir,
            filetypes=[("Excel fajlovi", "*.xlsx"), ("Svi fajlovi", "*.*")],
        )
        if not filename:
            return
        path = Path(filename)

        def work():
            inserted, dupes, total_minor, max_date = import_bank_statement_xlsx(self.db, path, self.log)
            self.log(
                f"[BANK] Rezime: novi={inserted} | duplikati={dupes} | ukupno={format_minor_km(int(total_minor))} KM | max_datum={max_date or 'N/A'}"
            )
            try:
                self.root.after(
                    0,
                    lambda p=path, md=max_date, tm=int(total_minor): (
                        self.settings.set("last_bank_statement_file", p.name),
                        self.settings.set("last_bank_statement_max_date", md or ""),
                        self.settings.set("last_bank_statement_total_minor", tm),
                        self.settings.set("last_bank_statement_ts", now_utc_iso()),
                        self.settings.save(),
                        self.refresh_dashboard(),
                        self.refresh_control_bank_unmatched(),
                    ),
                )
            except Exception:
                pass

        self._run_in_thread(work)

    def on_import_csv(self) -> None:
        filenames = filedialog.askopenfilenames(
            title="Odaberi BHP uplate (CSV) – može više fajlova",
            filetypes=[("CSV fajlovi", "*.csv"), ("Svi fajlovi", "*.*")],
        )
        if not filenames:
            return
        paths = [Path(p) for p in filenames]
        # Sort by date in filename (DDMMYY); unknown dates go last.
        def sort_key(p: Path):
            d = parse_ddmmyy_from_filename(p.name)
            return (1, date.max) if d is None else (0, d)

        paths.sort(key=sort_key)

        bank_mode = bool(self.db.has_bank_transactions())
        reserved_bank_ids: set[int] = set()
        selected: list[tuple[Path, str | None, str, int, int | None]] = []
        for idx, path in enumerate(paths, start=1):
            preview = preview_bhp_uplate_csv(path)
            found_iso = preview.get("found_date_iso")  # type: ignore[assignment]
            name_date = parse_ddmmyy_from_filename(path.name)

            file_total_minor = int(preview.get("total_minor", 0))
            bank_match = None
            bank_match_source = None
            if bank_mode and file_total_minor > 0:
                base = parse_iso_date(found_iso if isinstance(found_iso, str) else None)
                if base is None and name_date is not None:
                    base = name_date
                if base is None:
                    base = date.today()
                start_iso = base.isoformat()
                end_iso = (base + timedelta(days=10)).isoformat()
                try:
                    candidates = self.db.find_unlinked_bank_tx_in_window(
                        amount_minor=file_total_minor, start_date_iso=start_iso, end_date_iso=end_iso, limit=5
                    )
                    candidates = [c for c in candidates if int(c["id"]) not in reserved_bank_ids]
                    if len(candidates) == 1:
                        bank_match = candidates[0]
                        bank_match_source = "window"
                    elif len(candidates) == 0:
                        c2 = self.db.find_unlinked_bank_tx_by_amount(amount_minor=file_total_minor, limit=5)
                        c2 = [c for c in c2 if int(c["id"]) not in reserved_bank_ids]
                        if len(c2) == 1:
                            bank_match = c2[0]
                            bank_match_source = "amount"
                except Exception:
                    bank_match = None
                    bank_match_source = None

            decision = self._prompt_paid_date_for_csv(
                path,
                found_iso if isinstance(found_iso, str) else None,
                name_date,
                file_total_minor,
                int(preview.get("valid_count", 0)),
                int(preview.get("invalid_count", 0)),
                bank_mode=bank_mode,
                bank_match=bank_match,
                bank_match_source=bank_match_source,
            )
            if decision is None:
                self.log("[CSV] Uvoz prekinut.")
                return
            action, chosen_iso = decision
            if action == "skip":
                self.log(f"[CSV] Preskočen fajl: {path.name}")
                continue
            if action == "abort":
                self.log("[CSV] Uvoz prekinut.")
                return

            chosen_bank_id: int | None = None
            if bank_mode:
                if bank_match is not None:
                    chosen_iso = str(bank_match["bank_date"])
                    chosen_bank_id = int(bank_match["id"])
                    reserved_bank_ids.add(chosen_bank_id)

            self.log(
                f"[CSV] Potvrđen uvoz fajla {path.name}: pronađeno={iso_to_ddmmyyyy(found_iso if isinstance(found_iso,str) else None)} | datum_uplate={iso_to_ddmmyyyy(chosen_iso)} | broj_uplata={int(preview.get('valid_count',0))} | ukupno={format_minor_km(file_total_minor)} KM"
            )
            selected.append((path, found_iso, chosen_iso, file_total_minor, chosen_bank_id))

        if not selected:
            self.log("[CSV] Nema fajlova za uvoz (svi preskočeni).")
            return

        def work():
            total_inserted = 0
            total_skipped = 0
            total_files = 0
            affected_months: set[tuple[int, int]] = set()

            for path, found_iso, chosen_iso, file_total_minor, chosen_bank_id in selected:
                total_files += 1
                self.log(
                    f"[CSV] Fajl {total_files}/{len(selected)}: {path} | pronađeno={found_iso or 'N/A'} | override={chosen_iso}"
                )

                bank_tx_id = int(chosen_bank_id) if chosen_bank_id is not None else None
                final_iso = chosen_iso
                if bank_tx_id is not None:
                    try:
                        self.db.claim_bank_tx(bank_tx_id=bank_tx_id, source_file=path.name, note="csv-batch")
                    except Exception:
                        pass

                inserted, _updated, skipped, months = import_bhp_uplate_csv(
                    self.db,
                    path,
                    self.log,
                    override_paid_date_iso=final_iso,
                    bank_tx_id_for_file=bank_tx_id,
                )
                total_inserted += int(inserted)
                total_skipped += int(skipped)
                affected_months |= set(months or set())
                self.log(f"[CSV] Rezime fajla: novi={inserted} | preskočeno={skipped}")

            self.log(f"[CSV] Ukupno (fajlova={len(selected)}): novi={total_inserted} | preskočeno={total_skipped}")

            # Remember last imported payments file for the dashboard (store in settings in UI thread).
            try:
                last_path, _found, last_chosen = selected[-1]
                self.root.after(
                    0,
                    lambda lp=last_path, ld=last_chosen: (
                        self.settings.set("last_payments_import_file", lp.name),
                        self.settings.set("last_payments_import_paid_date", ld),
                        self.settings.set("last_payments_import_ts", now_utc_iso()),
                        self.settings.save(),
                        self.refresh_dashboard(),
                    ),
                )
            except Exception:
                pass

            # Auto-refresh cashbook months touched by this batch (OPEN months only, no MANUAL rows).
            for y, m in sorted(affected_months):
                try:
                    rebuilt = self.db.rebuild_cashbook_entries_safe(y, m)
                    if rebuilt is not None:
                        self.log(f"[CASH] Auto preračunato: {y:04d}-{m:02d}")
                except Exception as exc:
                    self.log(f"[ERROR] cashbook auto preračun {y:04d}-{m:02d}: {exc}")
                    self.log(traceback.format_exc().rstrip())

            stats = self.db.run_matching(CARRIER_BHP)
            self.log(
                "[MATCH] Pošiljke={shipments} | Otvoreno={OPEN} | Djelimično={PARTIAL} | Uplaćeno={PAID} | Neusklađeno={MISMATCH} | Neuparene uplate={unmatched_payments}".format(
                    **stats
                )
            )
            self.root.after(0, self.refresh_matching_view)
            self.root.after(0, self.refresh_cashbook_view)

        self._run_in_thread(work)

    def _prompt_paid_date_for_csv(
        self,
        path: Path,
        found_iso: str | None,
        name_date: date | None,
        total_minor: int,
        valid_count: int,
        invalid_count: int,
        *,
        bank_mode: bool,
        bank_match: sqlite3.Row | None,
        bank_match_source: str | None,
    ) -> tuple[str, str] | None:
        win = Toplevel(self.root)
        apply_window_icon(win)
        win.title(f"Odaberi datum uplate za fajl: {path.name}")
        win.transient(self.root)
        win.grab_set()
        win.resizable(False, False)

        result: dict[str, object] = {"action": None, "date": None}
        default_date = parse_iso_date(found_iso) or date.today()
        if parse_iso_date(found_iso) is None and name_date is not None:
            default_date = name_date

        frm = Frame(win)
        frm.pack(fill=BOTH, expand=True, padx=14, pady=14)

        ttk.Label(frm, text=f"Datum pronađen u fajlu: {iso_to_ddmmyyyy(found_iso)}").grid(
            row=0, column=0, columnspan=4, sticky="w", pady=(0, 6)
        )
        name_hint = name_date.strftime("%d.%m.%Y") if name_date is not None else "N/A"
        ttk.Label(frm, text=f"Datum iz naziva fajla: {name_hint}").grid(
            row=1, column=0, columnspan=4, sticky="w", pady=(0, 6)
        )
        ttk.Label(frm, text=f"Ukupna vrijednost uplata u fajlu: {format_minor_km(int(total_minor))} KM").grid(
            row=2, column=0, columnspan=4, sticky="w", pady=(0, 6)
        )
        ttk.Label(frm, text=f"Broj uplata u fajlu: {int(valid_count)}").grid(
            row=3, column=0, columnspan=4, sticky="w", pady=(0, 10)
        )

        warn_text = ""
        if valid_count <= 0:
            warn_text = "Upozorenje: fajl nema nijednu validnu uplatu (tracking + iznos)."
        elif invalid_count > 0:
            warn_text = f"Napomena: preskočeno invalid redova u pregledu: {int(invalid_count)}"
        warn = ttk.Label(frm, text=warn_text, foreground="#b00020")
        if warn_text:
            warn.grid(row=4, column=0, columnspan=4, sticky="w", pady=(0, 10))

        date_row = 5 if warn_text else 4
        picker = None
        if bank_mode:
            # In bank mode, bank date is the source of truth. Do not allow manual date input.
            if bank_match is not None:
                ttk.Label(frm, text="Datum uplate (iz banke):", font=("Segoe UI", 10, "bold")).grid(
                    row=date_row, column=0, columnspan=4, sticky="w", pady=(0, 6)
                )
                ttk.Label(
                    frm,
                    text=f"Pronađen bank match ({bank_match_source or 'match'}): {iso_to_ddmmyyyy(str(bank_match['bank_date']))} | {format_minor_km(int(total_minor))} KM",
                    foreground="#0b3d91",
                ).grid(row=date_row + 1, column=0, columnspan=4, sticky="w", pady=(0, 6))
                ttk.Label(frm, text=f"Opis: {bank_match['description'] or ''}", foreground="#666666").grid(
                    row=date_row + 2, column=0, columnspan=4, sticky="w", pady=(0, 10)
                )
            else:
                ttk.Label(frm, text="Datum uplate (iz banke):", font=("Segoe UI", 10, "bold")).grid(
                    row=date_row, column=0, columnspan=4, sticky="w", pady=(0, 6)
                )
                ttk.Label(
                    frm,
                    text="Nije pronađen odgovarajući bank izvod za ovaj fajl.\nUvezi bank izvod ili preskoči fajl.",
                    foreground="#b00020",
                ).grid(row=date_row + 1, column=0, columnspan=4, sticky="w", pady=(0, 10))
        else:
            ttk.Label(frm, text="Odaberi datum uplate:").grid(row=date_row, column=0, columnspan=4, sticky="w")
            picker = CalendarPicker(frm, initial=default_date)
            picker.grid(row=date_row + 1, column=0, columnspan=4, sticky="w", pady=(6, 12))

        btns = Frame(frm)
        btns_row_default = date_row + (4 if bank_mode else 3)
        btns.grid(row=btns_row_default, column=0, columnspan=4, sticky="e")

        manual_override_enabled = {"value": False}

        def enable_manual_override():
            # Switch to manual date selection for legacy files without bank statements.
            if manual_override_enabled["value"]:
                return
            manual_override_enabled["value"] = True
            try:
                win.title(f"Odaberi datum uplate (ručno): {path.name}")
            except Exception:
                pass
            # Render a picker under the bank message.
            nonlocal picker
            ttk.Label(frm, text="Odaberi datum uplate (ručno):", font=("Segoe UI", 10, "bold")).grid(
                row=date_row + 3, column=0, columnspan=4, sticky="w", pady=(6, 6)
            )
            picker = CalendarPicker(frm, initial=default_date)
            picker.grid(row=date_row + 4, column=0, columnspan=4, sticky="w", pady=(0, 12))
            # Move buttons below the calendar to avoid overlap/clipping.
            try:
                btns.grid_forget()
            except Exception:
                pass
            btns.grid(row=date_row + 6, column=0, columnspan=4, sticky="e", pady=(6, 0))
            try:
                confirm_btn.configure(state="normal" if valid_count > 0 else "disabled")
            except Exception:
                pass
            try:
                win.resizable(True, True)
                w = int(640 * UI_SCALE)
                h = int(620 * UI_SCALE)
                win.geometry(f"{w}x{h}")
                win.minsize(int(560 * UI_SCALE), int(560 * UI_SCALE))
                win.update_idletasks()
                center_window(win, self.root)
            except Exception:
                pass

        def confirm():
            if bank_mode:
                if bank_match is not None and not manual_override_enabled["value"]:
                    iso = str(bank_match["bank_date"])
                else:
                    iso = picker.selection_get().isoformat() if picker is not None else default_date.isoformat()
            else:
                iso = picker.selection_get().isoformat() if picker is not None else default_date.isoformat()
            if bank_mode and manual_override_enabled["value"]:
                result["action"] = "confirm_manual"
            else:
                result["action"] = "confirm"
            result["date"] = iso
            win.destroy()

        def skip():
            result["action"] = "skip"
            result["date"] = default_date.isoformat()
            win.destroy()

        def abort():
            result["action"] = "abort"
            win.destroy()

        Button(btns, text="Prekini uvoz", command=abort).pack(side=LEFT)
        Button(btns, text="Preskoči fajl", command=skip).pack(side=LEFT, padx=(6, 0))
        if bank_mode and bank_match is None:
            Button(btns, text="Ručno (override datum)", command=enable_manual_override).pack(side=LEFT, padx=(6, 0))
        confirm_btn = Button(btns, text="Potvrdi i uvezi", command=confirm)
        confirm_btn.pack(side=RIGHT)
        if valid_count <= 0 or (bank_mode and bank_match is None):
            confirm_btn.configure(state="disabled")

        win.bind("<Escape>", lambda _e: abort())
        center_window(win, self.root)
        win.focus_set()
        self.root.wait_window(win)

        action = result.get("action")
        if action is None:
            return None
        if action == "confirm" or action == "confirm_manual":
            return "confirm", str(result.get("date"))
        if action == "skip":
            return "skip", str(result.get("date"))
        return "abort", date.today().isoformat()

    def _build_matching_tab(self, parent: Frame) -> None:
        top = Frame(parent)
        top.pack(fill=BOTH, expand=True, padx=8, pady=8)

        left = Frame(top)
        right = Frame(top)
        left.pack(side=LEFT, fill=BOTH, expand=True)
        right.pack(side=RIGHT, fill=BOTH, expand=True)

        ttk.Label(left, text="Pošiljke (kurir + tracking broj)").pack(anchor="w")
        columns = ("tracking", "cod", "paid_total", "difference", "lifecycle_state", "payments_count")
        self.match_tree = ttk.Treeview(left, columns=columns, show="headings", height=14)
        headings = {
            "tracking": "Tracking broj",
            "cod": "Otkupnina (KM)",
            "paid_total": "Ukupno uplaćeno (KM)",
            "difference": "Razlika (KM)",
            "lifecycle_state": "Status",
            "payments_count": "Broj uplata",
        }
        for key in columns:
            self.match_tree.heading(key, text=headings[key])
            self.match_tree.column(key, stretch=True, width=120)
        self.match_tree.column("tracking", width=160)
        self.match_tree.column("lifecycle_state", width=110)
        self.match_tree.column("payments_count", width=90)

        match_scroll = ttk.Scrollbar(left, orient="vertical", command=self.match_tree.yview)
        self.match_tree.configure(yscrollcommand=match_scroll.set)
        self.match_tree.pack(side=LEFT, fill=BOTH, expand=True)
        match_scroll.pack(side=RIGHT, fill="y")
        self.match_tree.bind("<Double-1>", self._on_matching_double_click)

        ttk.Label(right, text="Neuparene uplate (nema pošiljke za tracking broj)").pack(anchor="w")
        pcols = ("tracking", "payer_name", "amount", "paid_date", "created_at")
        self.unmatched_tree = ttk.Treeview(right, columns=pcols, show="headings", height=14)
        pheads = {
            "tracking": "Tracking broj",
            "payer_name": "Ime i prezime",
            "amount": "Iznos (KM)",
            "paid_date": "Datum uplate",
            "created_at": "Datum uvoza",
        }
        for key in pcols:
            self.unmatched_tree.heading(key, text=pheads[key])
            self.unmatched_tree.column(key, stretch=True, width=140)
        self.unmatched_tree.column("tracking", width=160)
        self.unmatched_tree.column("amount", width=90)

        un_scroll = ttk.Scrollbar(right, orient="vertical", command=self.unmatched_tree.yview)
        self.unmatched_tree.configure(yscrollcommand=un_scroll.set)
        self.unmatched_tree.pack(side=LEFT, fill=BOTH, expand=True)
        un_scroll.pack(side=RIGHT, fill="y")
        self.unmatched_tree.bind("<Double-1>", self._on_matching_unmatched_double_click)

        Button(parent, text="Osvježi", command=self.refresh_matching_view).pack(anchor="e", padx=8, pady=(0, 8))

    def refresh_matching_view(self) -> None:
        try:
            rows = self.db.fetch_matching_overview(CARRIER_BHP)
            prows = self.db.fetch_unmatched_payments(CARRIER_BHP)
        except Exception as exc:
            self.log(f"[ERROR] refresh_matching_view: {exc}")
            return

        def clear_and_insert(tree: ttk.Treeview, items: list[tuple]):
            for item_id in tree.get_children():
                tree.delete(item_id)
            for item in items:
                tree.insert("", END, values=item)

        match_items: list[tuple] = []
        for r in rows:
            match_items.append(
                (
                    r["tracking"],
                    format_minor_km(r["cod_amount_minor"]),
                    format_minor_km(r["paid_total_minor"]),
                    format_minor_km(r["difference_minor"]) if r["difference_minor"] is not None else "",
                    ui_status_label(str(r["lifecycle_state"])),
                    int(r["payments_count"]),
                )
            )

        unmatched_items: list[tuple] = []
        for r in prows:
            unmatched_items.append(
                (
                    r["tracking"],
                    r["payer_name"] or "",
                    format_minor_km(int(r["paid_amount_minor"])),
                    r["paid_date"] or "",
                    r["created_at"] or "",
                )
            )

        clear_and_insert(self.match_tree, match_items)
        clear_and_insert(self.unmatched_tree, unmatched_items)
        self.log(f"[MATCH] Pregled osvježen. Pošiljke={len(match_items)} | Neuparene uplate={len(unmatched_items)}")

    def _on_matching_double_click(self, event) -> None:
        item_id = self.match_tree.identify_row(event.y)
        if not item_id:
            return
        values = self.match_tree.item(item_id, "values")
        if not values:
            return
        tracking = str(values[0])
        self.open_tracking_link(tracking, CARRIER_BHP)

    def _on_matching_unmatched_double_click(self, event) -> None:
        item_id = self.unmatched_tree.identify_row(event.y)
        if not item_id:
            return
        values = self.unmatched_tree.item(item_id, "values")
        if not values:
            return
        tracking = str(values[0])
        self.open_tracking_link(tracking, CARRIER_BHP)

    def _build_dashboard_tab(self, parent: Frame) -> None:
        top = Frame(parent)
        top.pack(fill="x", padx=8, pady=8)

        self.dash_period_var = StringVar(value="")
        self.dash_metrics_var = StringVar(value="")
        self.dash_last_imports_var = StringVar(value="")
        self.dash_from_var = StringVar(value="")
        self.dash_to_var = StringVar(value="")

        ttk.Label(top, text="Period (datum preuzimanja):").pack(side=LEFT)
        DateField(top, var=self.dash_from_var, width=12, on_change=self._debounced_refresh_dashboard).pack(
            side=LEFT, padx=(6, 6)
        )
        ttk.Label(top, text="do").pack(side=LEFT)
        DateField(top, var=self.dash_to_var, width=12, on_change=self._debounced_refresh_dashboard).pack(
            side=LEFT, padx=(6, 12)
        )
        Button(top, text="Osvježi", command=self.refresh_dashboard).pack(side=LEFT)
        Button(top, text="Globalno osvježi", command=self.global_refresh).pack(side=LEFT, padx=(8, 0))

        ttk.Label(top, textvariable=self.dash_period_var, foreground="#666666").pack(side=RIGHT)

        box = Frame(parent)
        box.pack(fill=BOTH, expand=True, padx=8, pady=(0, 8))
        ttk.Label(box, textvariable=self.dash_metrics_var, justify="left").pack(anchor="w")
        ttk.Label(box, textvariable=self.dash_last_imports_var, justify="left", foreground="#666666").pack(
            anchor="w", pady=(10, 0)
        )
        self._load_dashboard_filters()

    def refresh_dashboard(self) -> None:
        pickup_from = (self.dash_from_var.get() or "").strip() or None
        pickup_to = (self.dash_to_var.get() or "").strip() or None
        carrier = getattr(self, "ops_carrier_cb", None).get() if getattr(self, "ops_carrier_cb", None) else CARRIER_BHP

        try:
            row = self.db.fetch_dashboard_metrics(carrier=carrier, pickup_from=pickup_from, pickup_to=pickup_to)
        except Exception as exc:
            self.log(f"[ERROR] dashboard: {exc}")
            return

        total_shipments = int(row["total_shipments"] or 0)
        total_active = int(row["total_active"] or 0)
        paid = int(row["paid_count"] or 0)
        partial = int(row["partial_count"] or 0)
        open_ = int(row["open_count"] or 0)
        removed = int(row["removed_count"] or 0)
        mismatch = int(row["mismatch_count"] or 0)

        paid_pct = (paid / total_active * 100.0) if total_active > 0 else 0.0
        removed_pct = (removed / total_shipments * 100.0) if total_shipments > 0 else 0.0

        period = "Period: sve"
        if pickup_from or pickup_to:
            period = f"Aktivni period: {pickup_from or '...'} → {pickup_to or '...'}"
        self.dash_period_var.set(period)

        self.dash_metrics_var.set(
            "\n".join(
                [
                    f"Ukupno pošiljki u periodu: {total_shipments}",
                    f"Uplaćeno: {paid} ({paid_pct:.1f}% od aktivnih)",
                    f"Djelimično uplaćeno: {partial}",
                    f"Otvoreno: {open_}",
                    f"Neusklađeno: {mismatch}",
                    f"Nepreuzeto: {removed} ({removed_pct:.1f}%)",
                    f"Ukupna otkupnina: {format_minor_km(int(row['sum_cod_minor'] or 0))} KM",
                    f"Ukupno uplaćeno: {format_minor_km(int(row['sum_paid_minor'] or 0))} KM",
                    f"Ukupno neuplaćeno (otvoreno/djelimično): {format_minor_km(int(row['sum_unpaid_minor'] or 0))} KM",
                ]
            )
        )

        try:
            last_pay_file = str(self.settings.get("last_payments_import_file", "") or "").strip() or "N/A"
            last_cash_out = self.db.fetch_last_cashbook_payments_out_date()
            last_cash_out_disp = iso_to_ddmmyyyy(last_cash_out) if last_cash_out else "N/A"

            last_bank_tx = self.db.fetch_latest_bank_transaction()
            if last_bank_tx is None:
                last_bank_tx_date = "N/A"
                last_bank_tx_amount = "N/A"
            else:
                last_bank_tx_date = iso_to_ddmmyyyy(str(last_bank_tx["bank_date"] or "") or None)
                last_bank_tx_amount = format_minor_km(int(last_bank_tx["amount_minor"] or 0))

            last_ship = self.db.fetch_latest_shipment_basic(carrier)
            if last_ship is None:
                last_ship_text = "N/A"
                last_ship_pickup = "N/A"
            else:
                last_ship_text = str(last_ship["tracking"] or "") or "N/A"
                last_ship_pickup = iso_to_ddmmyyyy(str(last_ship["pickup_date"] or "") or None)

            self.dash_last_imports_var.set(
                "\n".join(
                    [
                        f"Posljednje uvezene uplate: {last_pay_file} | Zadnji datum upisa u blagajnu: {last_cash_out_disp}",
                        f"Posljednje uvezene narudžbe: {last_ship_text} | Datum preuzimanja: {last_ship_pickup}",
                        f"Zadnji bank izvod: {last_bank_tx_date} | Iznos: {last_bank_tx_amount} KM",
                    ]
                )
            )
        except Exception:
            # Avoid dashboard failure on optional info.
            self.dash_last_imports_var.set("")

        self._save_dashboard_filters()

    def _debounced_refresh_dashboard(self) -> None:
        if hasattr(self, "_dash_refresh_after_id") and self._dash_refresh_after_id:
            try:
                self.root.after_cancel(self._dash_refresh_after_id)
            except Exception:
                pass
        self._dash_refresh_after_id = self.root.after(200, self.refresh_dashboard)

    def global_refresh(self) -> None:
        def work():
            try:
                stats = self.db.run_matching(CARRIER_BHP)
                self.log(
                    "[MATCH] Pošiljke={shipments} | Otvoreno={OPEN} | Djelimično={PARTIAL} | Uplaćeno={PAID} | Neusklađeno={MISMATCH} | Neuparene uplate={unmatched_payments}".format(
                        **stats
                    )
                )
            except Exception as exc:
                self.log(f"[ERROR] global_refresh matching: {exc}")
            self.root.after(0, self._refresh_all_views)

        self._run_in_thread(work)

    def _load_dashboard_filters(self) -> None:
        try:
            f = self.settings.get("dashboard_filters", {}) or {}
            if isinstance(f, dict):
                self.dash_from_var.set(str(f.get("pickup_from", "") or ""))
                self.dash_to_var.set(str(f.get("pickup_to", "") or ""))
        except Exception:
            pass

    def _save_dashboard_filters(self) -> None:
        try:
            self.settings.set(
                "dashboard_filters",
                {"pickup_from": self.dash_from_var.get(), "pickup_to": self.dash_to_var.get()},
            )
            self.settings.save()
        except Exception:
            pass

    def open_tracking_link(self, tracking: str, carrier: str = CARRIER_BHP) -> None:
        if carrier == CARRIER_BHP:
            webbrowser.open_new_tab(bhp_tracking_url(tracking))
            return
        messagebox.showinfo(APP_NAME, f"Nema link template za carrier={carrier}")

    def _build_ops_tab(self, parent: Frame) -> None:
        controls_card = Frame(parent, bd=1, relief="solid", padx=10, pady=10)
        controls_card.pack(fill="x", padx=10, pady=(10, 6))

        self.ops_search_var = StringVar(value="")
        self.ops_from_var = StringVar(value="")
        self.ops_to_var = StringVar(value="")
        self.ops_month_var = StringVar(value="")
        self.ops_month_summary_var = StringVar(value="")
        self.ops_claims_summary_var = StringVar(value="")
        self.ops_state_vars: dict[str, IntVar] = {
            "OPEN": IntVar(value=1),
            "PARTIAL": IntVar(value=1),
            "PAID": IntVar(value=1),
            "MISMATCH": IntVar(value=1),
            "UNDELIVERED_REMOVED": IntVar(value=1),
        }

        row1 = Frame(controls_card)
        row2 = Frame(controls_card)
        row1.pack(fill="x")
        row2.pack(fill="x", pady=(8, 0))

        left_group = Frame(row1)
        mid_group = Frame(row1)
        left_group.pack(side=LEFT, fill="x", expand=True)
        mid_group.pack(side=LEFT, fill="x", expand=True)

        status_group = Frame(row2)
        buttons_group = Frame(row2)
        status_group.pack(side=LEFT, fill="x", expand=True)
        buttons_group.pack(side=RIGHT)

        ttk.Label(left_group, text="Kurir:").pack(side=LEFT)
        self.ops_carrier_cb = ttk.Combobox(left_group, width=10, values=[CARRIER_BHP], state="readonly")
        self.ops_carrier_cb.set(CARRIER_BHP)
        self.ops_carrier_cb.pack(side=LEFT, padx=(6, 14))

        ttk.Label(left_group, text="Mjesec potraživanja:").pack(side=LEFT)
        self.ops_month_cb = ttk.Combobox(left_group, width=10, textvariable=self.ops_month_var, state="normal")
        today = date.today()
        self.ops_month_cb["values"] = [
            f"{today.year:04d}-{m:02d}" for m in range(1, 13)
        ] + [f"{today.year - 1:04d}-{m:02d}" for m in range(1, 13)]
        self.ops_month_cb.pack(side=LEFT, padx=(6, 14))

        ttk.Label(mid_group, text="Pretraga:").pack(side=LEFT)
        ttk.Entry(mid_group, width=24, textvariable=self.ops_search_var).pack(side=LEFT, padx=(6, 14))

        ttk.Label(mid_group, text="Preuzeto od:").pack(side=LEFT)
        DateField(mid_group, var=self.ops_from_var, width=10, on_change=self._debounced_refresh_ops).pack(
            side=LEFT, padx=(6, 6)
        )
        ttk.Label(mid_group, text="do:").pack(side=LEFT)
        DateField(mid_group, var=self.ops_to_var, width=10, on_change=self._debounced_refresh_ops).pack(
            side=LEFT, padx=(6, 14)
        )

        for state in ("OPEN", "PARTIAL", "PAID", "MISMATCH", "UNDELIVERED_REMOVED"):
            ttk.Checkbutton(status_group, text=ui_status_label(state), variable=self.ops_state_vars[state]).pack(
                side=LEFT, padx=4
            )

        ttk.Label(status_group, textvariable=self.ops_claims_summary_var, foreground="#0b3d91").pack(
            side=RIGHT, padx=(10, 0)
        )

        export_filters_btn = Button(
            buttons_group,
            text="Izvoz pregleda (filteri)",
            command=self.export_ops_filtered_xlsx,
            bg="#1f6feb",
            fg="white",
            activebackground="#1a5fd1",
            activeforeground="white",
            bd=0,
            padx=10,
            pady=5,
        )
        export_filters_btn.pack(side=RIGHT, padx=(8, 0))
        Button(buttons_group, text="Izvoz mjeseca", command=self.export_ops_month).pack(side=RIGHT, padx=(8, 0))
        Button(buttons_group, text="Preuzmi tracking - izvod", command=self.export_ops_filtered_with_tracking).pack(
            side=RIGHT, padx=(8, 0)
        )
        Button(buttons_group, text="Osvježi", command=self.refresh_ops_view).pack(side=RIGHT)

        info = ttk.Label(
            parent,
            text="Prikazane su pošiljke u odabranom periodu. Statusi i uplate se automatski usklađuju.",
            foreground="#666666",
        )
        info.pack(fill="x", padx=12, pady=(0, 6))
        ttk.Separator(parent, orient="horizontal").pack(fill="x", padx=12, pady=(0, 6))

        columns = (
            "tracking",
            "recipient",
            "phone",
            "pickup_date",
            "days",
            "cod",
            "paid_total",
            "difference",
            "status",
            "actions",
        )
        self.ops_tree = ttk.Treeview(parent, columns=columns, show="headings", height=16)
        headings = {
            "tracking": "Tracking broj",
            "recipient": "Primalac",
            "phone": "Telefon",
            "pickup_date": "Datum preuzimanja",
            "days": "Dani u dostavi",
            "cod": "Otkupnina (KM)",
            "paid_total": "Ukupno uplaćeno (KM)",
            "difference": "Razlika (KM)",
            "status": "Status",
            "actions": "Akcije",
        }
        for key in columns:
            self.ops_tree.heading(key, text=headings[key])
            self.ops_tree.column(key, stretch=True, width=140)
        self.ops_tree.column("tracking", width=160, stretch=False)
        self.ops_tree.column("pickup_date", width=120, stretch=False)
        self.ops_tree.column("days", width=110, stretch=False, anchor="e")
        self.ops_tree.column("cod", width=120, stretch=False, anchor="e")
        self.ops_tree.column("paid_total", width=140, stretch=False, anchor="e")
        self.ops_tree.column("difference", width=120, stretch=False, anchor="e")
        self.ops_tree.column("status", width=140, stretch=False)
        self.ops_tree.column("actions", width=140, stretch=False)

        scroll = ttk.Scrollbar(parent, orient="vertical", command=self.ops_tree.yview)
        self.ops_tree.configure(yscrollcommand=scroll.set)
        self.ops_tree.pack(side=LEFT, fill=BOTH, expand=True, padx=(8, 0), pady=(0, 8))
        scroll.pack(side=RIGHT, fill="y", padx=(0, 8), pady=(0, 8))

        # Very subtle status background hints (premium look).
        self.ops_tree.tag_configure("state_OPEN", background="#fffdf5")
        self.ops_tree.tag_configure("state_PARTIAL", background="#fff9e8")
        self.ops_tree.tag_configure("state_PAID", background="#f3fbf5")
        self.ops_tree.tag_configure("state_MISMATCH", background="#fff3f3")
        self.ops_tree.tag_configure("state_UNDELIVERED_REMOVED", background="#f7f7f7")
        self.ops_tree.tag_configure("diff_zero", foreground="#666666")
        self.ops_tree.tag_configure("diff_neg", foreground="#b00020")

        self.ops_ctx_menu = Menu(self.root, tearoff=0)
        try:
            menu_font = tkfont.nametofont("TkMenuFont").copy()
            size = int(menu_font.cget("size"))
            # Tk uses negative sizes for pixel units; larger magnitude => larger.
            if size < 0:
                menu_font.configure(size=int(round(size * 1.8)))
            else:
                menu_font.configure(size=int(round(size * 1.8)))
            self.ops_ctx_menu.configure(font=menu_font)
        except Exception:
            pass

        self.ops_ctx_menu.add_command(label="Skini iz potraživanja", command=self.remove_selected_from_claims)
        self.ops_ctx_menu.add_command(label="Kopiraj tracking broj", command=self._copy_selected_tracking)
        self.ops_ctx_menu.add_command(label="Uredi uplatu (iznos)", command=self._edit_payment_amount_for_selected)

        bottom = Frame(parent)
        bottom.pack(fill="x", padx=8, pady=(0, 8))
        Button(bottom, text="Skini iz potraživanja (odabrano)", command=self.remove_selected_from_claims).pack(
            side=LEFT
        )
        ttk.Label(bottom, textvariable=self.ops_month_summary_var).pack(side=RIGHT)

        self.ops_tree.bind("<Double-1>", self._on_ops_double_click)
        self.ops_tree.bind("<Button-3>", self._on_ops_right_click)
        self.ops_month_cb.bind("<<ComboboxSelected>>", lambda _e: self.apply_ops_month_filter())
        self.ops_month_cb.bind("<Return>", lambda _e: self.apply_ops_month_filter())

        self._load_ops_filters()
        self.refresh_ops_view()

    def _on_ops_right_click(self, event) -> None:
        try:
            item = self.ops_tree.identify_row(event.y)
            if item:
                self.ops_tree.selection_set(item)
                self.ops_tree.focus(item)

            tracking = self._ops_selected_tracking()
            state = None
            if tracking:
                row = next((r for r in self._ops_rows if str(r["tracking"]) == tracking), None)
                if row is not None:
                    state = str(row["lifecycle_state"])

            can_remove = bool(tracking) and state not in (None, "PAID")
            self.ops_ctx_menu.entryconfigure(0, state=("normal" if can_remove else "disabled"))
            self.ops_ctx_menu.entryconfigure(1, state=("normal" if tracking else "disabled"))
            self.ops_ctx_menu.entryconfigure(2, state=("normal" if tracking else "disabled"))
            self.ops_ctx_menu.tk_popup(event.x_root, event.y_root)
        finally:
            try:
                self.ops_ctx_menu.grab_release()
            except Exception:
                pass

    def _copy_selected_tracking(self) -> None:
        tracking = self._ops_selected_tracking()
        if not tracking:
            return
        try:
            self.root.clipboard_clear()
            self.root.clipboard_append(tracking)
            self.log(f"[OPS] Kopirano u clipboard: {tracking}")
        except Exception as exc:
            self.log(f"[ERROR] clipboard: {exc}")

    def _edit_payment_amount_for_selected(self) -> None:
        tracking = self._ops_selected_tracking()
        if not tracking:
            return
        self._open_edit_payment_dialog(tracking, self.ops_carrier_cb.get() or CARRIER_BHP)

    def _open_edit_payment_dialog(self, tracking: str, carrier: str) -> None:
        tracking = normalize_tracking(tracking) or ""
        if not tracking:
            return

        win = Toplevel(self.root)
        apply_window_icon(win)
        win.title(f"Ispravka uplate: {tracking}")
        win.transient(self.root)
        win.grab_set()
        win.resizable(True, True)
        # Give enough space for the editor panel (prevents clipped controls on smaller widths).
        win.geometry(f"{int(1040 * UI_SCALE)}x{int(560 * UI_SCALE)}")
        win.minsize(int(920 * UI_SCALE), int(520 * UI_SCALE))

        top = Frame(win)
        top.pack(fill="x", padx=12, pady=(12, 8))
        ttk.Label(top, text=f"Tracking broj: {tracking}", font=("Segoe UI", 11, "bold")).pack(side=LEFT)

        info_var = StringVar(value="Učitavam uplate…")
        ttk.Label(win, textvariable=info_var, foreground="#666666").pack(anchor="w", padx=12, pady=(0, 8))

        content = Frame(win)
        content.pack(fill=BOTH, expand=True, padx=12, pady=(0, 12))
        content.grid_columnconfigure(0, weight=1)
        content.grid_columnconfigure(1, weight=0)
        content.grid_rowconfigure(0, weight=1)

        columns = ("id", "paid_date", "payer_name", "amount_km", "created_at")
        tree = ttk.Treeview(content, columns=columns, show="headings", height=14)
        tree.heading("id", text="ID")
        tree.heading("paid_date", text="Datum uplate")
        tree.heading("payer_name", text="Uplatilac")
        tree.heading("amount_km", text="Iznos (KM)")
        tree.heading("created_at", text="Uvezeno")
        tree.column("id", width=60, stretch=False, anchor="e")
        tree.column("paid_date", width=110, stretch=False)
        tree.column("payer_name", width=220, stretch=True)
        tree.column("amount_km", width=110, stretch=False, anchor="e")
        tree.column("created_at", width=160, stretch=False)

        table = Frame(content)
        table.grid(row=0, column=0, sticky="nsew")
        table.grid_rowconfigure(0, weight=1)
        table.grid_columnconfigure(0, weight=1)

        scroll = ttk.Scrollbar(table, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scroll.set)
        tree.grid(row=0, column=0, sticky="nsew")
        scroll.grid(row=0, column=1, sticky="ns")

        editor = Frame(content)
        editor.grid(row=0, column=1, sticky="n", padx=(12, 0))
        editor.grid_columnconfigure(1, weight=1)

        state: dict[str, object] = {"rows": []}
        new_amount_var = StringVar(value="")
        reason_var = StringVar(value="")

        def reload_rows() -> None:
            for item in tree.get_children():
                tree.delete(item)
            rows = self.db.fetch_payments_for_tracking(carrier, tracking)
            state["rows"] = rows
            if not rows:
                info_var.set("Nema uplata za ovaj tracking.")
                return
            info_var.set(f"Uplata pronađeno: {len(rows)} (odaberi red i klikni 'Uredi iznos')")
            for r in rows:
                tree.insert(
                    "",
                    END,
                    values=(
                        int(r["id"]),
                        r["paid_date"] or "",
                        r["payer_name"] or "",
                        _payment_amount_display(int(r["paid_amount_minor"])),
                        r["created_at"] or "",
                    ),
                )
            # Select first row by default.
            try:
                first = tree.get_children()[0]
                tree.selection_set(first)
                tree.focus(first)
                tree.see(first)
                on_select(None)
            except Exception:
                pass

        def selected_payment() -> sqlite3.Row | None:
            sel = tree.selection()
            if not sel:
                return None
            values = tree.item(sel[0], "values")
            if not values:
                return None
            pid = int(values[0])
            for r in state.get("rows", []) or []:
                if int(r["id"]) == pid:
                    return r
            return None

        def on_select(_e) -> None:
            r = selected_payment()
            if r is None:
                new_amount_var.set("")
                return
            new_amount_var.set(_payment_amount_display(int(r["paid_amount_minor"])))

        def save_amount():
            r = selected_payment()
            if r is None:
                messagebox.showinfo(APP_NAME, "Odaberi uplatu.", parent=win)
                return
            pid = int(r["id"])
            old_minor = int(r["paid_amount_minor"])
            # UI field is "KM" (major units). A plain "154" means 154,00 KM (not 1,54).
            new_minor = parse_km_major_to_minor(new_amount_var.get())
            if new_minor is None:
                messagebox.showerror(APP_NAME, "Neispravan iznos.", parent=win)
                return

            try:
                self.db.update_payment_amount_minor(pid, int(new_minor))
                self.db.audit(
                    "EDIT_PAYMENT_AMOUNT",
                    f"payment {pid}",
                    json.dumps(
                        {
                            "carrier": carrier,
                            "tracking": tracking,
                            "old_minor": old_minor,
                            "new_minor": int(new_minor),
                            "payer_name": r["payer_name"] or "",
                            "paid_date": r["paid_date"] or "",
                            "reason": (reason_var.get() or "").strip(),
                        },
                        ensure_ascii=False,
                    ),
                )
                self.log(
                    f"[PAY] Ispravljeno: tracking={tracking} payment_id={pid} {format_minor_km(old_minor)} -> {format_minor_km(int(new_minor))}"
                )
            except Exception as exc:
                self.log(f"[ERROR] edit payment: {exc}")
                self.log(traceback.format_exc().rstrip())
                messagebox.showerror(APP_NAME, f"Greška: {exc}", parent=win)
                return

            try:
                self.db.run_matching(carrier)
            except Exception as exc:
                self.log(f"[ERROR] matching after edit: {exc}")

            try:
                st = self.db.fetch_shipment_settlement_by_tracking(carrier, tracking)
                if st is not None:
                    info_var.set(
                        "Sačuvano. Status: {status} | COD: {cod} KM | Uplaćeno: {paid} KM | Razlika: {diff} KM".format(
                            status=ui_status_label(str(st["lifecycle_state"])),
                            cod=format_minor_km(int(st["cod_amount_minor"] or 0)),
                            paid=format_minor_km(int(st["paid_total_minor"] or 0)),
                            diff=format_minor_km(int(st["difference_minor"] or 0)),
                        )
                    )
                    self.log(
                        "[PAY] Nakon izmjene: tracking={t} status={s} cod={cod} paid={paid} diff={diff}".format(
                            t=tracking,
                            s=str(st["lifecycle_state"]),
                            cod=int(st["cod_amount_minor"] or 0),
                            paid=int(st["paid_total_minor"] or 0),
                            diff=int(st["difference_minor"] or 0),
                        )
                    )
            except Exception:
                pass

            self._refresh_all_views()
            reload_rows()
            try:
                messagebox.showinfo(APP_NAME, "Sačuvano.", parent=win)
            except Exception:
                pass

        tree.bind("<<TreeviewSelect>>", on_select)

        ttk.Label(editor, text="Novi iznos (KM):").grid(row=0, column=0, sticky="w", pady=(0, 6))
        amount_entry = ttk.Entry(editor, width=14, textvariable=new_amount_var)
        amount_entry.grid(row=0, column=1, sticky="w", pady=(0, 6), padx=(6, 6))
        Button(editor, text="Sačuvaj", command=save_amount).grid(row=0, column=2, sticky="w", pady=(0, 6))

        ttk.Label(editor, text="Razlog:").grid(row=1, column=0, sticky="w", pady=(0, 6))
        ttk.Entry(editor, width=28, textvariable=reason_var).grid(
            row=1, column=1, columnspan=2, sticky="ew", pady=(0, 6), padx=(6, 0)
        )

        ttk.Separator(editor, orient="horizontal").grid(row=2, column=0, columnspan=3, sticky="ew", pady=(8, 8))
        Button(editor, text="Zatvori", command=win.destroy).grid(row=3, column=0, columnspan=3, sticky="e")

        reload_rows()
        try:
            amount_entry.focus_set()
        except Exception:
            pass
        try:
            win.update_idletasks()
        except Exception:
            pass
        center_window(win, self.root)
        win.focus_set()
        self.root.wait_window(win)

    def _debounced_refresh_ops(self) -> None:
        if hasattr(self, "_ops_refresh_after_id") and self._ops_refresh_after_id:
            try:
                self.root.after_cancel(self._ops_refresh_after_id)
            except Exception:
                pass
        self._ops_refresh_after_id = self.root.after(200, self.refresh_ops_view)

    def _ops_selected_tracking(self) -> str | None:
        sel = self.ops_tree.selection()
        if not sel:
            return None
        values = self.ops_tree.item(sel[0], "values")
        return str(values[0]) if values else None

    def _ops_neighbor_trackings(self, tracking: str) -> tuple[str | None, str | None]:
        tracking = normalize_tracking(tracking) or tracking
        items = list(self.ops_tree.get_children())
        for i, iid in enumerate(items):
            vals = self.ops_tree.item(iid, "values")
            if not vals:
                continue
            if normalize_tracking(str(vals[0])) == tracking:
                next_t = None
                prev_t = None
                if i + 1 < len(items):
                    nvals = self.ops_tree.item(items[i + 1], "values")
                    if nvals:
                        next_t = str(nvals[0])
                if i - 1 >= 0:
                    pvals = self.ops_tree.item(items[i - 1], "values")
                    if pvals:
                        prev_t = str(pvals[0])
                return next_t, prev_t
        return None, None

    def _ops_select_tracking(self, tracking: str) -> bool:
        tracking = normalize_tracking(tracking) or tracking
        for iid in self.ops_tree.get_children():
            vals = self.ops_tree.item(iid, "values")
            if not vals:
                continue
            if normalize_tracking(str(vals[0])) == tracking:
                self.ops_tree.selection_set(iid)
                self.ops_tree.focus(iid)
                try:
                    self.ops_tree.see(iid)
                except Exception:
                    pass
                try:
                    self.ops_tree.focus_set()
                except Exception:
                    pass
                return True
        return False

    def _ops_select_fallback_row(self) -> None:
        items = self.ops_tree.get_children()
        if not items:
            return
        iid = items[0]
        self.ops_tree.selection_set(iid)
        self.ops_tree.focus(iid)
        try:
            self.ops_tree.see(iid)
            self.ops_tree.focus_set()
        except Exception:
            pass

    def _on_ops_double_click(self, event) -> None:
        try:
            item_id = self.ops_tree.identify_row(event.y)
            if not item_id:
                return
            col = self.ops_tree.identify_column(event.x)
            self.ops_tree.selection_set(item_id)
            self.ops_tree.focus(item_id)
            values = self.ops_tree.item(item_id, "values")
            if not values:
                return
            tracking = str(values[0])
            self.log(f"[TRACK] Dvoklik: {tracking}")

            # Keep action accessible: double-click on "Akcije" triggers it; otherwise open review modal.
            if col == "#10":
                action_text = str(values[9])
                if action_text.startswith("Skini"):
                    self.remove_selected_from_claims()
                    return

            # Ctrl + dvoklik -> otvori u pregledniku (brzi shortcut)
            if int(getattr(event, "state", 0)) & 0x0004:
                self.open_tracking_link(tracking, self.ops_carrier_cb.get() or CARRIER_BHP)
                return

            # Always show the in-app tracking review; browser is available from the modal.
            self._show_tracking_review(tracking, self.ops_carrier_cb.get() or CARRIER_BHP)
        except Exception as exc:
            self.log(f"[ERROR] tracking modal: {exc}")
            self.log(traceback.format_exc().rstrip())
            try:
                messagebox.showerror(APP_NAME, f"Greška pri otvaranju praćenja: {exc}", parent=self.root)
            except Exception:
                pass

    def _show_tracking_review(self, tracking: str, carrier: str) -> None:
        # Remember current position in the ops list so we can auto-advance after the review.
        next_tracking, prev_tracking = self._ops_neighbor_trackings(tracking)

        win = Toplevel(self.root)
        apply_window_icon(win)
        win.title(f"Praćenje pošiljke: {tracking}")
        win.transient(self.root)
        win.grab_set()
        # Size based on UI scale so bottom buttons are never clipped.
        w = int(980 * UI_SCALE)
        h = int(560 * UI_SCALE)
        win.geometry(f"{w}x{h}")
        try:
            win.minsize(int(860 * UI_SCALE), int(520 * UI_SCALE))
        except Exception:
            pass
        win.resizable(True, True)

        header = Frame(win)
        header.pack(fill="x", padx=12, pady=(12, 6))

        ttk.Label(header, text=f"Tracking broj: {tracking}", font=("Segoe UI", 11, "bold")).pack(side=LEFT)
        status_txt = ""
        row = next((r for r in getattr(self, "_ops_rows", []) if str(r["tracking"]) == tracking), None)
        if row is not None:
            status_txt = ui_status_label(str(row["lifecycle_state"]))
        if status_txt:
            ttk.Label(header, text=f"Status: {status_txt}", foreground="#666666").pack(side=LEFT, padx=(12, 0))

        info_var = StringVar(value="Učitavam podatke…")
        ttk.Label(win, textvariable=info_var, foreground="#666666").pack(anchor="w", padx=12, pady=(0, 6))

        table_area = Frame(win)
        table_area.pack(fill=BOTH, expand=True, padx=12, pady=(0, 8))

        columns = ("dt", "country", "location", "event", "note")
        # Keep a bit of space for the bottom action buttons.
        tree = ttk.Treeview(table_area, columns=columns, show="headings", height=14)
        tree.heading("dt", text="Datum i vrijeme")
        tree.heading("country", text="Zemlja")
        tree.heading("location", text="Lokacija")
        tree.heading("event", text="Događaj")
        tree.heading("note", text="Napomena")
        tree.column("dt", width=150, stretch=False)
        tree.column("country", width=180, stretch=False)
        tree.column("location", width=240, stretch=True)
        tree.column("event", width=260, stretch=True)
        tree.column("note", width=220, stretch=True)

        scroll = ttk.Scrollbar(table_area, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scroll.set)
        tree.pack(side=LEFT, fill=BOTH, expand=True, padx=(0, 0), pady=(0, 0))
        scroll.pack(side=RIGHT, fill="y", padx=(6, 0), pady=(0, 0))

        btns = Frame(win)
        btns.pack(fill="x", padx=12, pady=(0, 12))

        def _advance_ops_selection():
            # Prefer selecting the next row; fallback to previous; then first row.
            if next_tracking and self._ops_select_tracking(next_tracking):
                return
            if prev_tracking and self._ops_select_tracking(prev_tracking):
                return
            self._ops_select_fallback_row()

        def close():
            win.destroy()

        can_remove = True
        if row is not None and str(row["lifecycle_state"]) == "PAID":
            can_remove = False

        def do_remove():
            if not can_remove:
                return
            ok = self.remove_tracking_from_claims(tracking)
            if ok:
                # After removal+refresh, select the next row so the user doesn't lose their place.
                try:
                    self.root.after(0, _advance_ops_selection)
                except Exception:
                    pass
                close()

        remove_btn = Button(btns, text="Skini sa potraživanja", command=do_remove)
        if not can_remove:
            remove_btn.configure(state="disabled")
        else:
            remove_btn.configure(
                bg="#d1fae5",  # subtle green
                activebackground="#bbf7d0",
            )
        remove_btn.pack(side=LEFT)
        def do_keep():
            try:
                self.root.after(0, _advance_ops_selection)
            except Exception:
                pass
            close()

        keep_btn = Button(btns, text="Ostavi na potraživanjima", command=do_keep)
        keep_btn.configure(
            bg="#fee2e2",  # subtle red
            activebackground="#fecaca",
        )
        keep_btn.pack(side=LEFT, padx=(8, 0))

        center_window(win, self.root)
        try:
            win.lift()
            win.focus_force()
            win.after(50, win.lift)
        except Exception:
            pass

        def populate(events: list[dict[str, str]] | None, err: str | None) -> None:
            for item in tree.get_children():
                tree.delete(item)
            if err:
                info_var.set(f"[Greška] {err}")
                return
            if not events:
                info_var.set("Nema događaja za prikaz.")
                return
            info_var.set(f"Pronađeno događaja: {len(events)}")
            for ev in events:
                tree.insert(
                    "",
                    "end",
                    values=(
                        ev.get("datetime", ""),
                        ev.get("country", ""),
                        ev.get("location", ""),
                        ev.get("event", ""),
                        ev.get("note", ""),
                    ),
                )

        def worker():
            try:
                evs = fetch_bhp_tracking_events(tracking, timeout_s=10, debug_log=None)
                self.root.after(0, lambda: populate(evs, None))
                return
            except TrackingNoDataError as exc:
                msg = "Automatsko praćenje nije dostupno (BHP WebForms / reCAPTCHA)."
                self.root.after(0, lambda: info_var.set(msg))
                return
            except TrackingCaptchaError:
                msg = "Automatsko praćenje nije dostupno (BHP WebForms / reCAPTCHA)."
                self.root.after(0, lambda: info_var.set(msg))
                return
            except Exception as exc:
                self.root.after(0, lambda: populate([], "Nije moguće dohvatiti podatke sa BHP sistema."))

        threading.Thread(target=worker, daemon=True).start()
        self.root.wait_window(win)

    def refresh_ops_view(self) -> None:
        carrier = self.ops_carrier_cb.get() or CARRIER_BHP
        search = (self.ops_search_var.get() or "").strip()
        pickup_from = (self.ops_from_var.get() or "").strip() or None
        pickup_to = (self.ops_to_var.get() or "").strip() or None
        states = [k for k, v in self.ops_state_vars.items() if int(v.get()) == 1]

        try:
            rows = self.db.fetch_shipments_operational(
                carrier=carrier,
                search=search or None,
                pickup_from=pickup_from,
                pickup_to=pickup_to,
                states=states or None,
                limit=2000,
            )
        except Exception as exc:
            self.log(f"[ERROR] ops refresh: {exc}")
            return

        self._ops_rows = rows
        today = date.today()

        for item_id in self.ops_tree.get_children():
            self.ops_tree.delete(item_id)

        for r in rows:
            state = str(r["lifecycle_state"])
            days = ""
            if state in ("OPEN", "PARTIAL"):
                d = days_between(r["pickup_date"], today)
                if d is None:
                    days = ""
                elif d > 14:
                    days = f"{d} !!"
                elif d > 7:
                    days = f"{d} !"
                else:
                    days = str(d)

            actions = "Skini iz potraživanja" if state != "PAID" else "-"
            tag = f"state_{state}"
            diff_tag = None
            try:
                diff_minor = r["difference_minor"]
                if diff_minor is None:
                    diff_tag = None
                else:
                    d = int(diff_minor)
                    if d == 0:
                        diff_tag = "diff_zero"
                    elif d < 0:
                        diff_tag = "diff_neg"
            except Exception:
                diff_tag = None
            self.ops_tree.insert(
                "",
                END,
                values=(
                    r["tracking"],
                    r["recipient_name"] or "",
                    r["recipient_phone"] or "",
                    r["pickup_date"] or "",
                    days,
                    format_minor_km(r["cod_amount_minor"]),
                    format_minor_km(int(r["paid_total_minor"])),
                    format_minor_km(r["difference_minor"]) if r["difference_minor"] is not None else "",
                    ui_status_label(state),
                    actions,
                ),
                tags=(tag,) if diff_tag is None else (tag, diff_tag),
            )

        # Claims summary (what we still need to collect): OPEN/PARTIAL only, sum positive differences.
        claim_rows = [r for r in rows if str(r["lifecycle_state"]) in ("OPEN", "PARTIAL")]
        claim_total_minor = 0
        for r in claim_rows:
            try:
                diff = int(r["difference_minor"] or 0)
            except Exception:
                diff = 0
            if diff > 0:
                claim_total_minor += diff
        self.ops_claims_summary_var.set(
            f"Potraživanja: {len(claim_rows)} | Ukupno: {format_minor_km(claim_total_minor)}"
        )

        self.log(f"[OPS] Učitano redova={len(rows)}")
        self._save_ops_filters()
        self.refresh_ops_month_summary()
        self.refresh_dashboard()

    def _load_ops_filters(self) -> None:
        try:
            ops = self.settings.get("ops_filters", {}) or {}
            if isinstance(ops, dict):
                self.ops_search_var.set(str(ops.get("search", "") or ""))
                self.ops_from_var.set(str(ops.get("pickup_from", "") or ""))
                self.ops_to_var.set(str(ops.get("pickup_to", "") or ""))
                self.ops_month_var.set(str(ops.get("month", "") or ""))
                carrier = str(ops.get("carrier", CARRIER_BHP) or CARRIER_BHP)
                if carrier in (CARRIER_BHP,):
                    self.ops_carrier_cb.set(carrier)
                states = ops.get("states", None)
                if isinstance(states, dict):
                    for k, v in states.items():
                        if k in self.ops_state_vars:
                            self.ops_state_vars[k].set(1 if v else 0)
        except Exception:
            pass

    def _save_ops_filters(self) -> None:
        try:
            states = {k: bool(v.get()) for k, v in self.ops_state_vars.items()}
            self.settings.set(
                "ops_filters",
                {
                    "carrier": self.ops_carrier_cb.get() or CARRIER_BHP,
                    "search": self.ops_search_var.get(),
                    "pickup_from": self.ops_from_var.get(),
                    "pickup_to": self.ops_to_var.get(),
                    "month": self.ops_month_var.get(),
                    "states": states,
                },
            )
            self.settings.save()
        except Exception:
            pass

    def apply_ops_month_filter(self) -> None:
        text = (self.ops_month_var.get() or "").strip()
        if not text:
            return
        try:
            year_s, month_s = text.split("-", 1)
            year = int(year_s)
            month = int(month_s)
            start, next_start = month_bounds(year, month)
            last_day = (date.fromisoformat(next_start) - timedelta(days=1)).isoformat()
            self.ops_from_var.set(start)
            self.ops_to_var.set(last_day)
            self.refresh_ops_view()
        except Exception:
            messagebox.showerror(APP_NAME, "Format mjeseca: YYYY-MM (npr. 2026-01)")

    def refresh_ops_month_summary(self) -> None:
        text = (self.ops_month_var.get() or "").strip()
        if not text:
            self.ops_month_summary_var.set("")
            return
        try:
            year_s, month_s = text.split("-", 1)
            year = int(year_s)
            month = int(month_s)
        except Exception:
            self.ops_month_summary_var.set("")
            return
        try:
            row = self.db.fetch_claims_month_summary(carrier=self.ops_carrier_cb.get() or CARRIER_BHP, year=year, month=month)
            open_c = int(row["open_count"] or 0)
            partial_c = int(row["partial_count"] or 0)
            unpaid = int(row["unpaid_minor"] or 0)
            status = "Mjesec operativno zatvoren ✔" if (open_c == 0 and partial_c == 0) else "Mjesec otvoren"
            self.ops_month_summary_var.set(
                f"Otvoreno u mjesecu: {open_c} | Djelimično: {partial_c} | Neuplaćeno: {format_minor_km(unpaid)} KM | {status}"
            )
        except Exception as exc:
            self.ops_month_summary_var.set(f"Greška: {exc}")

    def export_ops_month(self) -> None:
        text = (self.ops_month_var.get() or "").strip()
        if not text:
            messagebox.showinfo(APP_NAME, "Odaberi mjesec potraživanja (YYYY-MM).")
            return
        try:
            year_s, month_s = text.split("-", 1)
            year = int(year_s)
            month = int(month_s)
        except Exception:
            messagebox.showerror(APP_NAME, "Format mjeseca: YYYY-MM (npr. 2026-01)")
            return

        start, next_start = month_bounds(year, month)
        last_day = (date.fromisoformat(next_start) - timedelta(days=1)).isoformat()

        today = date.today().isoformat()
        default_name = f"ExportMjeseca_{year:04d}-{month:02d}_{today}.xlsx"
        filename = filedialog.asksaveasfilename(
            title=f"Izvoz mjeseca {year:04d}-{month:02d} (XLSX)",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel fajlovi", "*.xlsx")],
        )
        if not filename:
            return

        carrier = self.ops_carrier_cb.get() or CARRIER_BHP
        try:
            unpaid = self.db.fetch_unpaid_shipments(carrier=carrier, pickup_from=start, pickup_to=last_day, limit=5000)
            removed = self.db.fetch_removed_shipments(carrier=carrier, pickup_from=start, pickup_to=last_day, limit=5000)
            mismatch = self.db.fetch_mismatch_shipments(carrier=carrier, pickup_from=start, pickup_to=last_day, limit=5000)
            unmatched_payments = self.db.fetch_unmatched_payments_by_paid_date(
                carrier=carrier, paid_from=start, paid_to=last_day, limit=5000
            )
        except Exception as exc:
            self.log(f"[ERROR] month export query: {exc}")
            messagebox.showerror(APP_NAME, str(exc))
            return

        sheets: list[tuple[str, list[str], list[list[object]]]] = []

        sheets.append(
            (
                "Neuplaćene",
                [
                    "Tracking broj",
                    "Primalac",
                    "Datum preuzimanja",
                    "Otkupnina (KM)",
                    "Ukupno uplaćeno (KM)",
                    "Razlika (KM)",
                    "Status",
                ],
                [
                    [
                        r["tracking"],
                        r["recipient_name"] or "",
                        r["pickup_date"] or "",
                        format_minor_km(r["cod_amount_minor"]),
                        format_minor_km(int(r["paid_total_minor"])),
                        format_minor_km(r["difference_minor"]) if r["difference_minor"] is not None else "",
                        ui_status_label(str(r["lifecycle_state"])),
                    ]
                    for r in unpaid
                ],
            )
        )
        sheets.append(
            (
                "Nepreuzete",
                ["Tracking broj", "Primalac", "Datum preuzimanja", "Otkupnina (KM)", "Razlog", "Datum uklanjanja"],
                [
                    [
                        r["tracking"],
                        r["recipient_name"] or "",
                        r["pickup_date"] or "",
                        format_minor_km(r["cod_amount_minor"]),
                        r["removed_reason"] or "",
                        r["removed_at"] or "",
                    ]
                    for r in removed
                ],
            )
        )
        sheets.append(
            (
                "Neusklađene",
                [
                    "Tracking broj",
                    "Primalac",
                    "Datum preuzimanja",
                    "Otkupnina (KM)",
                    "Ukupno uplaćeno (KM)",
                    "Razlika (KM)",
                    "Broj uplata",
                    "Status",
                ],
                [
                    [
                        r["tracking"],
                        r["recipient_name"] or "",
                        r["pickup_date"] or "",
                        format_minor_km(r["cod_amount_minor"]),
                        format_minor_km(int(r["paid_total_minor"])),
                        format_minor_km(r["difference_minor"]) if r["difference_minor"] is not None else "",
                        int(r["payments_count"]),
                        ui_status_label(str(r["lifecycle_state"])),
                    ]
                    for r in mismatch
                ],
            )
        )
        sheets.append(
            (
                "Neuparene uplate",
                ["Tracking broj", "Ime i prezime", "Iznos (KM)", "Datum uplate", "Datum uvoza"],
                [
                    [
                        r["tracking"],
                        r["payer_name"] or "",
                        format_minor_km(int(r["paid_amount_minor"])),
                        r["paid_date"] or "",
                        r["created_at"] or "",
                    ]
                    for r in unmatched_payments
                ],
            )
        )

        export_xlsx_workbook(Path(filename), sheets)
        self.log(
            f"[IZVOZ] Mjesec {year:04d}-{month:02d}: neuplaćene={len(unpaid)} nepreuzete={len(removed)} neusklađene={len(mismatch)} neuparene_uplate={len(unmatched_payments)} -> {filename}"
        )

    def export_ops_filtered_xlsx(self) -> None:
        """
        Export exactly what the user currently sees in 'Pošiljke / Potraživanja' based on active filters
        (search + pickup date range + lifecycle status checkboxes + carrier).
        This is independent of the 'Mjesec potraživanja' dropdown.
        """
        carrier = self.ops_carrier_cb.get() or CARRIER_BHP
        search = (self.ops_search_var.get() or "").strip() or None
        pickup_from = (self.ops_from_var.get() or "").strip() or None
        pickup_to = (self.ops_to_var.get() or "").strip() or None
        states = [k for k, v in self.ops_state_vars.items() if int(v.get()) == 1]

        try:
            rows = self.db.fetch_shipments_operational(
                carrier=carrier,
                search=search,
                pickup_from=pickup_from,
                pickup_to=pickup_to,
                states=states or None,
                limit=100000,
            )
        except Exception as exc:
            self.log(f"[ERROR] ops filtered export query: {exc}")
            messagebox.showerror(APP_NAME, str(exc), parent=self.root)
            return

        period = "sve"
        if pickup_from or pickup_to:
            period = f"{pickup_from or '...'}_do_{pickup_to or '...'}"
        today = date.today().isoformat()
        default_name = f"Posiljke_Potraživanja_{period}_{today}.xlsx".replace(":", "-")
        filename = filedialog.asksaveasfilename(
            title="Izvoz pregleda (filteri) – XLSX",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel fajlovi", "*.xlsx")],
        )
        if not filename:
            return

        headers = [
            "Tracking broj",
            "Primalac",
            "Telefon",
            "Datum preuzimanja",
            "Dani u dostavi",
            "Otkupnina (KM)",
            "Ukupno uplaćeno (KM)",
            "Razlika (KM)",
            "Status",
            "Broj uplata",
            "Razlog (nepreuzeto)",
            "Datum uklanjanja",
        ]
        today_d = date.today()
        out_rows: list[list[object]] = []
        for r in rows:
            state = str(r["lifecycle_state"])
            days_val = ""
            if state in ("OPEN", "PARTIAL"):
                d = days_between(r["pickup_date"], today_d)
                days_val = "" if d is None else int(d)
            out_rows.append(
                [
                    r["tracking"],
                    r["recipient_name"] or "",
                    r["recipient_phone"] or "",
                    r["pickup_date"] or "",
                    days_val,
                    format_minor_km(r["cod_amount_minor"]),
                    format_minor_km(int(r["paid_total_minor"] or 0)),
                    format_minor_km(r["difference_minor"]) if r["difference_minor"] is not None else "",
                    ui_status_label(state),
                    int(r["payments_count"] or 0),
                    r["removed_reason"] or "",
                    r["removed_at"] or "",
                ]
            )

        export_xlsx(Path(filename), "Pošiljke", headers, out_rows)
        self.log(
            f"[IZVOZ] Pregled (filteri): carrier={carrier} from={pickup_from or ''} to={pickup_to or ''} rows={len(out_rows)} -> {filename}"
        )

    def export_ops_filtered_with_tracking(self) -> None:
        """
        Fetch latest tracking events from BHP WebForms for all rows in the current ops view (filters),
        then export an XLSX similar to 'Izvoz pregleda (filteri)' with an additional column:
          - Zadnji tracking upis (dd.mm.yyyy HH:MM – događaj)

        Throttling:
          - random 4–7s pause between requests
          - after each 30 processed, pause 30s
        """
        carrier = self.ops_carrier_cb.get() or CARRIER_BHP
        search = (self.ops_search_var.get() or "").strip() or None
        pickup_from = (self.ops_from_var.get() or "").strip() or None
        pickup_to = (self.ops_to_var.get() or "").strip() or None
        states = [k for k, v in self.ops_state_vars.items() if int(v.get()) == 1]

        try:
            rows = self.db.fetch_shipments_operational(
                carrier=carrier,
                search=search,
                pickup_from=pickup_from,
                pickup_to=pickup_to,
                states=states or None,
                limit=100000,
            )
        except Exception as exc:
            self.log(f"[ERROR] tracking export query: {exc}")
            messagebox.showerror(APP_NAME, str(exc), parent=self.root)
            return

        if not rows:
            messagebox.showinfo(APP_NAME, "Nema redova za izvoz.", parent=self.root)
            return

        period = "sve"
        if pickup_from or pickup_to:
            period = f"{pickup_from or '...'}_do_{pickup_to or '...'}"
        today = date.today().isoformat()
        default_name = f"Tracking_Izvod_{period}_{today}.xlsx".replace(":", "-")
        filename = filedialog.asksaveasfilename(
            title="Preuzmi tracking - izvod (XLSX)",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel fajlovi", "*.xlsx")],
        )
        if not filename:
            return

        if not messagebox.askyesno(
            APP_NAME,
            f"Preuzeti tracking podatke za {len(rows)} pošiljki?\n\n"
            "Napomena: ovo može trajati duže (pauze 4–7s između upita, +30s nakon svakih 30).",
            parent=self.root,
        ):
            return

        stop_event = threading.Event()
        self._track_export_stop_event = stop_event  # type: ignore[attr-defined]

        progress = Toplevel(self.root)
        apply_window_icon(progress)
        progress.title("Preuzmi tracking - izvod")
        progress.transient(self.root)
        progress.grab_set()
        progress.resizable(False, False)
        w = int(420 * UI_SCALE)
        h = int(170 * UI_SCALE)
        progress.geometry(f"{w}x{h}")

        wrap = max(int(360 * UI_SCALE), 260)
        lbl = ttk.Label(progress, text="Učitavanje…", justify="center", anchor="center", wraplength=wrap)
        lbl.pack(fill="both", expand=True, padx=16, pady=(16, 10))
        btns = Frame(progress)
        btns.pack(fill="x", padx=16, pady=(0, 16))

        def cancel():
            stop_event.set()
            try:
                lbl.configure(text="Prekidam…")
            except Exception:
                pass

        Button(btns, text="Prekini", command=cancel).pack(side=RIGHT)
        center_window(progress, self.root)

        def fmt_last_event(events: list[dict[str, str]] | None) -> tuple[str, str]:
            if not events:
                return "N/A", ""
            last = events[-1]
            dt = (last.get("datetime") or "").strip()
            ev = (last.get("event") or "").strip()
            if not dt and not ev:
                return "N/A", ""
            # dt is typically 'DD.MM.YYYY HH:MM:SS' -> date is first 10 chars.
            dt_date = ""
            try:
                if len(dt) >= 10:
                    dt_date = dt[:10]
            except Exception:
                dt_date = ""
            status = ev or "N/A"
            return status, dt_date

        def work():
            tracking_last: dict[str, tuple[str, str]] = {}
            total = len(rows)
            self.log(f"[TRACK-EXPORT] Start: rows={total} carrier={carrier} from={pickup_from or ''} to={pickup_to or ''}")

            for idx, r in enumerate(rows, start=1):
                if stop_event.is_set():
                    self.log(f"[TRACK-EXPORT] Prekinuto na {idx-1}/{total}.")
                    break
                tracking = str(r["tracking"])

                self.root.after(0, lambda i=idx, t=tracking: lbl.configure(text=f"{i}/{total} – {t}"))
                try:
                    evs = fetch_bhp_tracking_events(tracking, timeout_s=10, debug_log=None)
                    tracking_last[tracking] = fmt_last_event(evs)
                except TrackingNoDataError:
                    tracking_last[tracking] = ("N/A", "")
                except TrackingCaptchaError:
                    tracking_last[tracking] = ("N/A", "")
                except Exception as exc:
                    tracking_last[tracking] = ("N/A", "")
                    self.log(f"[TRACK-EXPORT] {tracking}: ERROR {exc}")

                if idx % 30 == 0 and idx < total and not stop_event.is_set():
                    self.log("[TRACK-EXPORT] Pauza 30s (batch 30).")
                    for _ in range(30):
                        if stop_event.is_set():
                            break
                        time.sleep(1)
                else:
                    if idx < total and not stop_event.is_set():
                        time.sleep(random.uniform(4.0, 7.0))

            # Build export rows (same as filtered export + extra column).
            headers = [
                "Tracking broj",
                "Primalac",
                "Telefon",
                "Datum preuzimanja",
                "Dani u dostavi",
                "Otkupnina (KM)",
                "Ukupno uplaćeno (KM)",
                "Razlika (KM)",
                "Status",
                "Broj uplata",
                "Razlog (nepreuzeto)",
                "Datum uklanjanja",
                "Zadnji tracking status",
                "Datum zadnjeg trackinga",
            ]
            today_d = date.today()
            out_rows: list[list[object]] = []
            for r in rows:
                state = str(r["lifecycle_state"])
                days_val = ""
                if state in ("OPEN", "PARTIAL"):
                    d = days_between(r["pickup_date"], today_d)
                    days_val = "" if d is None else int(d)
                t = str(r["tracking"])
                last_status, last_dt = tracking_last.get(t, ("", ""))
                out_rows.append(
                    [
                        t,
                        r["recipient_name"] or "",
                        r["recipient_phone"] or "",
                        r["pickup_date"] or "",
                        days_val,
                        format_minor_km(r["cod_amount_minor"]),
                        format_minor_km(int(r["paid_total_minor"] or 0)),
                        format_minor_km(r["difference_minor"]) if r["difference_minor"] is not None else "",
                        ui_status_label(state),
                        int(r["payments_count"] or 0),
                        r["removed_reason"] or "",
                        r["removed_at"] or "",
                        last_status,
                        last_dt,
                    ]
                )

            try:
                export_xlsx(Path(filename), "Pošiljke", headers, out_rows)
                self.log(f"[TRACK-EXPORT] Završeno: rows={len(out_rows)} -> {filename}")
            except Exception as exc:
                self.log(f"[ERROR] tracking export xlsx: {exc}")
                self.log(traceback.format_exc().rstrip())
                self.root.after(0, lambda: messagebox.showerror(APP_NAME, f"Greška pri izvozu: {exc}", parent=self.root))
            finally:
                self.root.after(0, lambda: (progress.destroy(), self.root.focus_force()))

        threading.Thread(target=work, daemon=True).start()
        self.root.wait_window(progress)

    def _remove_modal(self, *, tracking: str, cod_amount_minor: int) -> tuple[str, str | None, str | None, bool] | None:
        win = Toplevel(self.root)
        apply_window_icon(win)
        win.title("Skini iz potraživanja")
        win.transient(self.root)
        win.grab_set()

        reason_var = StringVar(value="Nepreuzeto")
        note_var = StringVar(value="")
        storno_var = BooleanVar(value=(int(cod_amount_minor or 0) > 0))
        storno_date_var = StringVar(value=date.today().isoformat())
        result: dict[str, object] = {"ok": False}

        frm = Frame(win)
        frm.pack(fill=BOTH, expand=True, padx=12, pady=12)

        ttk.Label(frm, text=f"Tracking: {tracking}").grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 6))
        ttk.Label(frm, text=f"Otkupnina: {format_minor_km(int(cod_amount_minor or 0))} KM").grid(
            row=1, column=0, columnspan=2, sticky="w", pady=(0, 10)
        )

        ttk.Label(frm, text="Razlog:").grid(row=2, column=0, sticky="w", pady=(0, 6))
        reason_cb = ttk.Combobox(
            frm,
            width=18,
            values=["Nepreuzeto", "Vraćeno", "Storno", "Ostalo"],
            state="readonly",
            textvariable=reason_var,
        )
        reason_cb.grid(row=2, column=1, sticky="w", pady=(0, 6))

        ttk.Label(frm, text="Napomena (opcionalno):").grid(row=3, column=0, sticky="w", pady=(0, 6))
        note_entry = ttk.Entry(frm, width=40, textvariable=note_var)
        note_entry.grid(row=3, column=1, sticky="w", pady=(0, 6))

        ttk.Separator(frm, orient="horizontal").grid(row=4, column=0, columnspan=2, sticky="ew", pady=(6, 8))
        storno_chk = ttk.Checkbutton(frm, text="Storno blagajne (umanji ULAZ)", variable=storno_var)
        storno_chk.grid(row=5, column=0, columnspan=2, sticky="w", pady=(0, 6))

        ttk.Label(frm, text="Datum storna:").grid(row=6, column=0, sticky="w", pady=(0, 6))
        storno_date_field = DateField(frm, var=storno_date_var, width=12)
        storno_date_field.grid(row=6, column=1, sticky="w", pady=(0, 6))

        if int(cod_amount_minor or 0) <= 0:
            storno_var.set(False)
            storno_chk.configure(state="disabled")
            storno_date_field.entry.configure(state="disabled")
            for child in storno_date_field.winfo_children():
                try:
                    child.configure(state="disabled")
                except Exception:
                    pass

        btns = Frame(frm)
        btns.grid(row=7, column=0, columnspan=2, sticky="e")

        def ok():
            if storno_var.get():
                iso = parse_date_to_iso(storno_date_var.get())
                if not iso or parse_iso_date(iso) is None:
                    messagebox.showerror(APP_NAME, "Neispravan datum storna.", parent=win)
                    return
                storno_date_var.set(iso)
            result["ok"] = True
            win.destroy()

        def cancel():
            win.destroy()

        Button(btns, text="Odustani", command=cancel).pack(side=RIGHT, padx=(6, 0))
        Button(btns, text="Potvrdi", command=ok).pack(side=RIGHT)

        note_entry.focus_set()
        center_window(win, self.root)
        self.root.wait_window(win)
        if not result.get("ok"):
            return None
        reason = reason_var.get().strip() or "Ostalo"
        note = (note_var.get() or "").strip() or None
        storno_date_iso = storno_date_var.get().strip() if storno_var.get() else None
        return reason, note, storno_date_iso, bool(storno_var.get())

    def remove_selected_from_claims(self) -> None:
        tracking = self._ops_selected_tracking()
        if not tracking:
            messagebox.showinfo(APP_NAME, "Nema odabira.", parent=self.root)
            return

        self.remove_tracking_from_claims(tracking)

    def remove_tracking_from_claims(self, tracking: str) -> bool:
        tracking = normalize_tracking(tracking) or ""
        if not tracking:
            messagebox.showinfo(APP_NAME, "Nema odabira.", parent=self.root)
            return False

        row = next((r for r in self._ops_rows if str(r["tracking"]) == tracking), None)
        if row is None:
            messagebox.showerror(APP_NAME, "Nije pronađen zapis.", parent=self.root)
            return False
        if str(row["lifecycle_state"]) == "PAID":
            messagebox.showwarning(APP_NAME, "Uplaćena pošiljka se ne skida iz potraživanja.", parent=self.root)
            return False

        cod_minor = int(row["cod_amount_minor"] or 0)
        modal = self._remove_modal(tracking=tracking, cod_amount_minor=cod_minor)
        if modal is None:
            return False
        reason, note, storno_date_iso, do_storno = modal

        if not messagebox.askyesno(
            APP_NAME, f"Potvrdi skidanje pošiljke {tracking} ({reason})?", parent=self.root
        ):
            return False

        try:
            self.db.remove_shipment_from_claims(self.ops_carrier_cb.get() or CARRIER_BHP, tracking, reason, note)
            self.log(f"[OPS] Skinuto iz potraživanja: {tracking} | razlog={reason}")

            # Optional: create storno to balance cashbook ULAZ (for COD>0).
            if do_storno and storno_date_iso and cod_minor > 0:
                try:
                    carrier = self.ops_carrier_cb.get() or CARRIER_BHP
                    created = self.db.create_storno_cod(
                        carrier=carrier,
                        tracking=tracking,
                        storno_date_iso=storno_date_iso,
                        amount_minor=cod_minor,
                        note=f"Ručno storno – {reason}",
                    )
                    if created:
                        self.log(
                            f"[STORNO] Kreiran storno: {tracking} | datum={storno_date_iso} | iznos={format_minor_km(cod_minor)} KM"
                        )
                    else:
                        self.log(f"[STORNO] Preskočeno (već postoji): {tracking}")
                except Exception as exc:
                    self.log(f"[ERROR] storno: {exc}")
                    self.log(traceback.format_exc().rstrip())

            self.refresh_ops_view()
            self.refresh_removed_view()
            # Update current month cashbook snapshot (safe).
            try:
                today = date.today()
                rebuilt = self.db.rebuild_cashbook_entries_safe(int(today.year), int(today.month))
                if rebuilt is not None:
                    self.log(f"[CASH] Auto preračunato: {today.year:04d}-{today.month:02d}")
            except Exception:
                pass
            self.refresh_cashbook_view()
            return True
        except Exception as exc:
            self.log(f"[ERROR] remove_from_claims: {exc}")
            self.log(traceback.format_exc().rstrip())
            messagebox.showerror(APP_NAME, f"Greška: {exc}", parent=self.root)
            return False

    def _build_unpaid_tab(self, parent: Frame) -> None:
        controls = Frame(parent)
        controls.pack(fill="x", padx=8, pady=8)

        self.unpaid_from_var = StringVar(value="")
        self.unpaid_to_var = StringVar(value="")

        ttk.Label(controls, text="Preuzeto od:").pack(side=LEFT)
        DateField(controls, var=self.unpaid_from_var, width=10, on_change=self._debounced_refresh_unpaid).pack(
            side=LEFT, padx=(4, 6)
        )
        ttk.Label(controls, text="do:").pack(side=LEFT)
        DateField(controls, var=self.unpaid_to_var, width=10, on_change=self._debounced_refresh_unpaid).pack(
            side=LEFT, padx=(4, 12)
        )

        Button(controls, text="Osvježi", command=self.refresh_unpaid_view).pack(side=LEFT, padx=6)
        Button(controls, text="Izvoz u XLSX", command=self.export_unpaid_xlsx).pack(side=RIGHT, padx=6)

        columns = ("tracking", "recipient", "pickup_date", "days", "cod", "paid_total", "difference", "status")
        self.unpaid_tree = ttk.Treeview(parent, columns=columns, show="headings", height=16)
        heads = {
            "tracking": "Tracking broj",
            "recipient": "Primalac",
            "pickup_date": "Datum preuzimanja",
            "days": "Dani u dostavi",
            "cod": "Otkupnina (KM)",
            "paid_total": "Uplaćeno (KM)",
            "difference": "Razlika (KM)",
            "status": "Status",
        }
        for key in columns:
            self.unpaid_tree.heading(key, text=heads[key])
            self.unpaid_tree.column(key, stretch=True, width=140)
        self.unpaid_tree.column("tracking", width=160, stretch=False)
        self.unpaid_tree.column("pickup_date", width=120, stretch=False)
        self.unpaid_tree.column("days", width=110, stretch=False, anchor="e")
        self.unpaid_tree.column("cod", width=120, stretch=False, anchor="e")
        self.unpaid_tree.column("paid_total", width=120, stretch=False, anchor="e")
        self.unpaid_tree.column("difference", width=120, stretch=False, anchor="e")
        self.unpaid_tree.column("status", width=110, stretch=False)

        scroll = ttk.Scrollbar(parent, orient="vertical", command=self.unpaid_tree.yview)
        self.unpaid_tree.configure(yscrollcommand=scroll.set)
        self.unpaid_tree.pack(side=LEFT, fill=BOTH, expand=True, padx=(8, 0), pady=(0, 8))
        scroll.pack(side=RIGHT, fill="y", padx=(0, 8), pady=(0, 8))

        self.unpaid_tree.tag_configure("state_OPEN", background="#fffdf5")
        self.unpaid_tree.tag_configure("state_PARTIAL", background="#fff9e8")
        self.unpaid_tree.tag_configure("diff_zero", foreground="#666666")
        self.unpaid_tree.tag_configure("diff_neg", foreground="#b00020")

        self.unpaid_tree.bind("<Double-1>", self._on_unpaid_double_click)
        self._load_unpaid_filters()
        self.refresh_unpaid_view()

    def _debounced_refresh_unpaid(self) -> None:
        if hasattr(self, "_unpaid_refresh_after_id") and self._unpaid_refresh_after_id:
            try:
                self.root.after_cancel(self._unpaid_refresh_after_id)
            except Exception:
                pass
        self._unpaid_refresh_after_id = self.root.after(200, self.refresh_unpaid_view)

    def _on_unpaid_double_click(self, event) -> None:
        item_id = self.unpaid_tree.identify_row(event.y)
        if not item_id:
            return
        values = self.unpaid_tree.item(item_id, "values")
        if not values:
            return
        tracking = str(values[0])
        self.open_tracking_link(tracking, CARRIER_BHP)

    def refresh_unpaid_view(self) -> None:
        pickup_from = (self.unpaid_from_var.get() or "").strip() or None
        pickup_to = (self.unpaid_to_var.get() or "").strip() or None
        try:
            rows = self.db.fetch_unpaid_shipments(
                carrier=CARRIER_BHP, pickup_from=pickup_from, pickup_to=pickup_to, limit=2000
            )
        except Exception as exc:
            self.log(f"[ERROR] unpaid refresh: {exc}")
            return

        self._unpaid_rows = rows
        today = date.today()

        for item_id in self.unpaid_tree.get_children():
            self.unpaid_tree.delete(item_id)

        for r in rows:
            state = str(r["lifecycle_state"])
            d = days_between(r["pickup_date"], today)
            if d is None:
                days = ""
            elif d > 14:
                days = f"{d} !!"
            elif d > 7:
                days = f"{d} !"
            else:
                days = str(d)
            diff_tag = None
            try:
                diff_minor = r["difference_minor"]
                if diff_minor is None:
                    diff_tag = None
                else:
                    d = int(diff_minor)
                    if d == 0:
                        diff_tag = "diff_zero"
                    elif d < 0:
                        diff_tag = "diff_neg"
            except Exception:
                diff_tag = None
            self.unpaid_tree.insert(
                "",
                END,
                values=(
                    r["tracking"],
                    r["recipient_name"] or "",
                    r["pickup_date"] or "",
                    days,
                    format_minor_km(r["cod_amount_minor"]),
                    format_minor_km(int(r["paid_total_minor"])),
                    format_minor_km(r["difference_minor"]) if r["difference_minor"] is not None else "",
                    ui_status_label(state),
                ),
                tags=(f"state_{state}",) if diff_tag is None else (f"state_{state}", diff_tag),
            )

        self.log(f"[UNPAID] Neuplaćene – učitano redova={len(rows)}")
        self._save_unpaid_filters()
        self.refresh_dashboard()

    def _load_unpaid_filters(self) -> None:
        try:
            f = self.settings.get("unpaid_filters", {}) or {}
            if isinstance(f, dict):
                self.unpaid_from_var.set(str(f.get("pickup_from", "") or ""))
                self.unpaid_to_var.set(str(f.get("pickup_to", "") or ""))
        except Exception:
            pass

    def _save_unpaid_filters(self) -> None:
        try:
            self.settings.set(
                "unpaid_filters",
                {"pickup_from": self.unpaid_from_var.get(), "pickup_to": self.unpaid_to_var.get()},
            )
            self.settings.save()
        except Exception:
            pass

    def export_unpaid_xlsx(self) -> None:
        today = date.today().isoformat()
        default_name = f"Neuplacene_{today}.xlsx"
        filename = filedialog.asksaveasfilename(
            title="Izvoz Neuplaćene (XLSX)",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel fajlovi", "*.xlsx")],
        )
        if not filename:
            return
        rows: list[list[object]] = []
        for r in self._unpaid_rows:
            rows.append(
                [
                    r["tracking"],
                    r["recipient_name"] or "",
                    r["pickup_date"] or "",
                    format_minor_km(r["cod_amount_minor"]),
                    format_minor_km(int(r["paid_total_minor"])),
                    format_minor_km(r["difference_minor"]) if r["difference_minor"] is not None else "",
                    ui_status_label(str(r["lifecycle_state"])),
                ]
            )
        export_xlsx(
            Path(filename),
            "Neuplacene",
            [
                "Tracking broj",
                "Primalac",
                "Datum preuzimanja",
                "Otkupnina (KM)",
                "Ukupno uplaćeno (KM)",
                "Razlika (KM)",
                "Status",
            ],
            rows,
        )
        self.log(f"[IZVOZ] Neuplaćene: {filename}")

    def _build_removed_tab(self, parent: Frame) -> None:
        controls = Frame(parent)
        controls.pack(fill="x", padx=8, pady=8)

        Button(controls, text="Osvježi", command=self.refresh_removed_view).pack(side=LEFT, padx=6)
        Button(controls, text="Izvoz u XLSX", command=self.export_removed_xlsx).pack(side=RIGHT, padx=6)

        columns = ("tracking", "recipient", "pickup_date", "cod", "reason", "removed_at")
        self.removed_tree = ttk.Treeview(parent, columns=columns, show="headings", height=16)
        heads = {
            "tracking": "Tracking broj",
            "recipient": "Primalac",
            "pickup_date": "Datum preuzimanja",
            "cod": "Otkupnina (KM)",
            "reason": "Razlog",
            "removed_at": "Datum uklanjanja",
        }
        for key in columns:
            self.removed_tree.heading(key, text=heads[key])
            self.removed_tree.column(key, stretch=True, width=160)
        self.removed_tree.column("tracking", width=160, stretch=False)
        self.removed_tree.column("pickup_date", width=120, stretch=False)
        self.removed_tree.column("cod", width=120, stretch=False, anchor="e")
        self.removed_tree.column("removed_at", width=170, stretch=False)

        scroll = ttk.Scrollbar(parent, orient="vertical", command=self.removed_tree.yview)
        self.removed_tree.configure(yscrollcommand=scroll.set)
        self.removed_tree.pack(side=LEFT, fill=BOTH, expand=True, padx=(8, 0), pady=(0, 8))
        scroll.pack(side=RIGHT, fill="y", padx=(0, 8), pady=(0, 8))

        self.removed_tree.tag_configure("state_UNDELIVERED_REMOVED", background="#f7f7f7")

        self.removed_tree.bind("<Double-1>", self._on_removed_double_click)
        self.refresh_removed_view()

    def _on_removed_double_click(self, event) -> None:
        item_id = self.removed_tree.identify_row(event.y)
        if not item_id:
            return
        values = self.removed_tree.item(item_id, "values")
        if not values:
            return
        tracking = str(values[0])
        self.open_tracking_link(tracking, CARRIER_BHP)

    def refresh_removed_view(self) -> None:
        try:
            rows = self.db.fetch_removed_shipments(carrier=CARRIER_BHP, limit=2000)
        except Exception as exc:
            self.log(f"[ERROR] removed refresh: {exc}")
            return

        self._removed_rows = rows
        for item_id in self.removed_tree.get_children():
            self.removed_tree.delete(item_id)

        for r in rows:
            self.removed_tree.insert(
                "",
                END,
                values=(
                    r["tracking"],
                    r["recipient_name"] or "",
                    r["pickup_date"] or "",
                    format_minor_km(r["cod_amount_minor"]),
                    r["removed_reason"] or "",
                    r["removed_at"] or "",
                ),
                tags=("state_UNDELIVERED_REMOVED",),
            )

        self.log(f"[REMOVED] Nepreuzete – učitano redova={len(rows)}")

    def export_removed_xlsx(self) -> None:
        today = date.today().isoformat()
        default_name = f"Nepreuzete_{today}.xlsx"
        filename = filedialog.asksaveasfilename(
            title="Izvoz Nepreuzete (XLSX)",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel fajlovi", "*.xlsx")],
        )
        if not filename:
            return
        rows: list[list[object]] = []
        for r in self._removed_rows:
            rows.append(
                [
                    r["tracking"],
                    r["recipient_name"] or "",
                    r["pickup_date"] or "",
                    format_minor_km(r["cod_amount_minor"]),
                    r["removed_reason"] or "",
                    r["removed_at"] or "",
                ]
            )
        export_xlsx(
            Path(filename),
            "Nepreuzete",
            ["Tracking broj", "Primalac", "Datum preuzimanja", "Otkupnina (KM)", "Razlog", "Datum uklanjanja"],
            rows,
        )
        self.log(f"[IZVOZ] Nepreuzete: {filename}")

    def _build_control_tab(self, parent: Frame) -> None:
        notebook = ttk.Notebook(parent)
        notebook.pack(fill=BOTH, expand=True)
        unmatched_tab = Frame(notebook)
        mismatch_tab = Frame(notebook)
        bank_all_tab = Frame(notebook)
        bank_unmatched_tab = Frame(notebook)
        notebook.add(unmatched_tab, text="Neuparene uplate")
        notebook.add(mismatch_tab, text="Neusklađene")
        notebook.add(bank_all_tab, text="Transakcije banke")
        notebook.add(bank_unmatched_tab, text="Neuparene uplate (banka)")

        self._build_control_unmatched(unmatched_tab)
        self._build_control_mismatch(mismatch_tab)
        self._build_control_bank_all(bank_all_tab)
        self._build_control_bank_unmatched(bank_unmatched_tab)

    def _build_control_unmatched(self, parent: Frame) -> None:
        controls = Frame(parent)
        controls.pack(fill="x", padx=8, pady=8)
        Button(controls, text="Osvježi", command=self.refresh_control_unmatched).pack(side=LEFT, padx=6)
        Button(controls, text="Izvoz u XLSX", command=self.export_control_unmatched_xlsx).pack(side=RIGHT, padx=6)

        cols = ("tracking", "payer", "amount", "paid_date", "created_at")
        self.ctrl_unmatched_tree = ttk.Treeview(parent, columns=cols, show="headings", height=16)
        heads = {
            "tracking": "Tracking broj",
            "payer": "Ime i prezime",
            "amount": "Iznos (KM)",
            "paid_date": "Datum uplate",
            "created_at": "Datum uvoza",
        }
        for k in cols:
            self.ctrl_unmatched_tree.heading(k, text=heads[k])
            self.ctrl_unmatched_tree.column(k, stretch=True, width=160)
        self.ctrl_unmatched_tree.column("tracking", width=160, stretch=False)
        self.ctrl_unmatched_tree.column("amount", width=110, stretch=False, anchor="e")
        self.ctrl_unmatched_tree.column("paid_date", width=120, stretch=False)
        self.ctrl_unmatched_tree.column("created_at", width=170, stretch=False)

        scroll = ttk.Scrollbar(parent, orient="vertical", command=self.ctrl_unmatched_tree.yview)
        self.ctrl_unmatched_tree.configure(yscrollcommand=scroll.set)
        self.ctrl_unmatched_tree.pack(side=LEFT, fill=BOTH, expand=True, padx=(8, 0), pady=(0, 8))
        scroll.pack(side=RIGHT, fill="y", padx=(0, 8), pady=(0, 8))

        self.ctrl_unmatched_tree.bind("<Double-1>", self._on_ctrl_unmatched_double_click)
        self.refresh_control_unmatched()

    def _on_ctrl_unmatched_double_click(self, event) -> None:
        item_id = self.ctrl_unmatched_tree.identify_row(event.y)
        if not item_id:
            return
        values = self.ctrl_unmatched_tree.item(item_id, "values")
        if not values:
            return
        tracking = str(values[0])
        self.open_tracking_link(tracking, CARRIER_BHP)

    def refresh_control_unmatched(self) -> None:
        try:
            rows = self.db.fetch_unmatched_payments(CARRIER_BHP, limit=2000)
        except Exception as exc:
            self.log(f"[ERROR] control unmatched refresh: {exc}")
            return
        self._ctrl_unmatched_rows = rows
        for item_id in self.ctrl_unmatched_tree.get_children():
            self.ctrl_unmatched_tree.delete(item_id)
        for r in rows:
            self.ctrl_unmatched_tree.insert(
                "",
                END,
                values=(
                    r["tracking"],
                    r["payer_name"] or "",
                    format_minor_km(int(r["paid_amount_minor"])),
                    r["paid_date"] or "",
                    r["created_at"] or "",
                ),
            )
        self.log(f"[CTRL] Neuparene uplate – učitano redova={len(rows)}")

    def export_control_unmatched_xlsx(self) -> None:
        today = date.today().isoformat()
        default_name = f"NeupareneUplate_{today}.xlsx"
        filename = filedialog.asksaveasfilename(
            title="Izvoz Neuparene uplate (XLSX)",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel fajlovi", "*.xlsx")],
        )
        if not filename:
            return
        rows: list[list[object]] = []
        for r in self._ctrl_unmatched_rows:
            rows.append(
                [
                    r["tracking"],
                    r["payer_name"] or "",
                    format_minor_km(int(r["paid_amount_minor"])),
                    r["paid_date"] or "",
                    r["created_at"] or "",
                ]
            )
        export_xlsx(
            Path(filename),
            "NeupareneUplate",
            ["Tracking broj", "Ime i prezime", "Iznos (KM)", "Datum uplate", "Datum uvoza"],
            rows,
        )
        self.log(f"[IZVOZ] Neuparene uplate: {filename}")

    def _build_control_mismatch(self, parent: Frame) -> None:
        controls = Frame(parent)
        controls.pack(fill="x", padx=8, pady=8)
        Button(controls, text="Osvježi", command=self.refresh_control_mismatch).pack(side=LEFT, padx=6)
        Button(controls, text="Izvoz u XLSX", command=self.export_control_mismatch_xlsx).pack(side=RIGHT, padx=6)

        cols = ("tracking", "recipient", "pickup_date", "cod", "paid_total", "difference", "payments_count", "status")
        self.ctrl_mismatch_tree = ttk.Treeview(parent, columns=cols, show="headings", height=16)
        heads = {
            "tracking": "Tracking broj",
            "recipient": "Primalac",
            "pickup_date": "Datum preuzimanja",
            "cod": "Otkupnina (KM)",
            "paid_total": "Uplaćeno (KM)",
            "difference": "Razlika (KM)",
            "payments_count": "Broj uplata",
            "status": "Status",
        }
        for k in cols:
            self.ctrl_mismatch_tree.heading(k, text=heads[k])
            self.ctrl_mismatch_tree.column(k, stretch=True, width=150)
        self.ctrl_mismatch_tree.column("tracking", width=160, stretch=False)
        self.ctrl_mismatch_tree.column("pickup_date", width=120, stretch=False)
        self.ctrl_mismatch_tree.column("cod", width=120, stretch=False, anchor="e")
        self.ctrl_mismatch_tree.column("paid_total", width=120, stretch=False, anchor="e")
        self.ctrl_mismatch_tree.column("difference", width=120, stretch=False, anchor="e")
        self.ctrl_mismatch_tree.column("payments_count", width=100, stretch=False, anchor="e")
        self.ctrl_mismatch_tree.column("status", width=110, stretch=False)

        scroll = ttk.Scrollbar(parent, orient="vertical", command=self.ctrl_mismatch_tree.yview)
        self.ctrl_mismatch_tree.configure(yscrollcommand=scroll.set)
        self.ctrl_mismatch_tree.pack(side=LEFT, fill=BOTH, expand=True, padx=(8, 0), pady=(0, 8))
        scroll.pack(side=RIGHT, fill="y", padx=(0, 8), pady=(0, 8))

        self.ctrl_mismatch_tree.tag_configure("state_MISMATCH", background="#fdecea")

        self.ctrl_mismatch_tree.bind("<Double-1>", self._on_ctrl_mismatch_double_click)
        self.refresh_control_mismatch()

    def _on_ctrl_mismatch_double_click(self, event) -> None:
        item_id = self.ctrl_mismatch_tree.identify_row(event.y)
        if not item_id:
            return
        values = self.ctrl_mismatch_tree.item(item_id, "values")
        if not values:
            return
        tracking = str(values[0])
        self.open_tracking_link(tracking, CARRIER_BHP)

    def refresh_control_mismatch(self) -> None:
        try:
            rows = self.db.fetch_mismatch_shipments(carrier=CARRIER_BHP, limit=2000)
        except Exception as exc:
            self.log(f"[ERROR] control mismatch refresh: {exc}")
            return
        self._ctrl_mismatch_rows = rows
        for item_id in self.ctrl_mismatch_tree.get_children():
            self.ctrl_mismatch_tree.delete(item_id)
        for r in rows:
            self.ctrl_mismatch_tree.insert(
                "",
                END,
                values=(
                    r["tracking"],
                    r["recipient_name"] or "",
                    r["pickup_date"] or "",
                    format_minor_km(r["cod_amount_minor"]),
                    format_minor_km(int(r["paid_total_minor"])),
                    format_minor_km(r["difference_minor"]) if r["difference_minor"] is not None else "",
                    int(r["payments_count"]),
                    ui_status_label(str(r["lifecycle_state"])),
                ),
                tags=("state_MISMATCH",),
            )
        self.log(f"[CTRL] Neusklađene – učitano redova={len(rows)}")

    def export_control_mismatch_xlsx(self) -> None:
        today = date.today().isoformat()
        default_name = f"Neusklađene_{today}.xlsx"
        filename = filedialog.asksaveasfilename(
            title="Izvoz Neusklađene (XLSX)",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel fajlovi", "*.xlsx")],
        )
        if not filename:
            return
        rows: list[list[object]] = []
        for r in self._ctrl_mismatch_rows:
            rows.append(
                [
                    r["tracking"],
                    r["recipient_name"] or "",
                    r["pickup_date"] or "",
                    format_minor_km(r["cod_amount_minor"]),
                    format_minor_km(int(r["paid_total_minor"])),
                    format_minor_km(r["difference_minor"]) if r["difference_minor"] is not None else "",
                    int(r["payments_count"]),
                    ui_status_label(str(r["lifecycle_state"])),
                ]
            )
        export_xlsx(
            Path(filename),
            "Neusklađene",
            [
                "Tracking broj",
                "Primalac",
                "Datum preuzimanja",
                "Otkupnina (KM)",
                "Ukupno uplaćeno (KM)",
                "Razlika (KM)",
                "Broj uplata",
                "Status",
            ],
            rows,
        )
        self.log(f"[IZVOZ] Neusklađene: {filename}")

    def _build_control_bank_all(self, parent: Frame) -> None:
        controls = Frame(parent)
        controls.pack(fill="x", padx=8, pady=8)
        Button(controls, text="Osvježi", command=self.refresh_control_bank_all).pack(side=LEFT, padx=6)

        cols = ("bank_date", "amount", "status", "source_file", "claimed_at", "description")
        self.ctrl_bank_all_tree = ttk.Treeview(parent, columns=cols, show="headings", height=16)
        heads = {
            "bank_date": "Datum",
            "amount": "Iznos (KM)",
            "status": "Status",
            "source_file": "CSV fajl",
            "claimed_at": "Upareno",
            "description": "Opis",
        }
        for k in cols:
            self.ctrl_bank_all_tree.heading(k, text=heads[k])
            self.ctrl_bank_all_tree.column(k, stretch=True, width=220)
        self.ctrl_bank_all_tree.column("bank_date", width=120, stretch=False)
        self.ctrl_bank_all_tree.column("amount", width=110, stretch=False, anchor="e")
        self.ctrl_bank_all_tree.column("status", width=110, stretch=False)
        self.ctrl_bank_all_tree.column("source_file", width=340, stretch=True)
        self.ctrl_bank_all_tree.column("claimed_at", width=130, stretch=False)
        self.ctrl_bank_all_tree.column("description", width=520, stretch=True)

        scroll = ttk.Scrollbar(parent, orient="vertical", command=self.ctrl_bank_all_tree.yview)
        self.ctrl_bank_all_tree.configure(yscrollcommand=scroll.set)
        self.ctrl_bank_all_tree.pack(side=LEFT, fill=BOTH, expand=True, padx=(8, 0), pady=(0, 8))
        scroll.pack(side=RIGHT, fill="y", padx=(0, 8), pady=(0, 8))

        self.refresh_control_bank_all()

    def _build_control_bank_unmatched(self, parent: Frame) -> None:
        controls = Frame(parent)
        controls.pack(fill="x", padx=8, pady=8)
        Button(controls, text="Osvježi", command=self.refresh_control_bank_unmatched).pack(side=LEFT, padx=6)

        cols = ("bank_date", "amount", "description")
        self.ctrl_bank_unmatched_tree = ttk.Treeview(parent, columns=cols, show="headings", height=16)
        heads = {
            "bank_date": "Datum",
            "amount": "Iznos (KM)",
            "description": "Opis",
        }
        for k in cols:
            self.ctrl_bank_unmatched_tree.heading(k, text=heads[k])
            self.ctrl_bank_unmatched_tree.column(k, stretch=True, width=220)
        self.ctrl_bank_unmatched_tree.column("bank_date", width=120, stretch=False)
        self.ctrl_bank_unmatched_tree.column("amount", width=110, stretch=False, anchor="e")
        self.ctrl_bank_unmatched_tree.column("description", width=520, stretch=True)

        scroll = ttk.Scrollbar(parent, orient="vertical", command=self.ctrl_bank_unmatched_tree.yview)
        self.ctrl_bank_unmatched_tree.configure(yscrollcommand=scroll.set)
        self.ctrl_bank_unmatched_tree.pack(side=LEFT, fill=BOTH, expand=True, padx=(8, 0), pady=(0, 8))
        scroll.pack(side=RIGHT, fill="y", padx=(0, 8), pady=(0, 8))

        self.refresh_control_bank_unmatched()

    def refresh_control_bank_all(self) -> None:
        try:
            rows = self.db.fetch_all_bank_transactions_with_claims(limit=5000)
        except Exception as exc:
            self.log(f"[ERROR] control bank all refresh: {exc}")
            return
        self._ctrl_bank_all_rows = rows
        for item_id in self.ctrl_bank_all_tree.get_children():
            self.ctrl_bank_all_tree.delete(item_id)

        paired = 0
        for r in rows:
            claimed_at = str(r["claimed_at"] or "") or None
            source_file = str(r["source_file"] or "") or ""
            is_paired = bool(claimed_at)
            if is_paired:
                paired += 1
            self.ctrl_bank_all_tree.insert(
                "",
                END,
                values=(
                    iso_to_ddmmyyyy(str(r["bank_date"] or "") or None),
                    format_minor_km(int(r["amount_minor"] or 0)),
                    "Upareno" if is_paired else "Neupareno",
                    source_file if source_file else ("Neupareno" if not is_paired else ""),
                    iso_to_ddmmyyyy(claimed_at),
                    r["description"] or "",
                ),
            )
        self.log(f"[CTRL] Transakcije banke – učitano={len(rows)} | upareno={paired} | neupareno={len(rows) - paired}")

    def refresh_control_bank_unmatched(self) -> None:
        try:
            rows = self.db.fetch_unmatched_bank_transactions(limit=2000)
        except Exception as exc:
            self.log(f"[ERROR] control bank unmatched refresh: {exc}")
            return
        self._ctrl_bank_unmatched_rows = rows
        for item_id in self.ctrl_bank_unmatched_tree.get_children():
            self.ctrl_bank_unmatched_tree.delete(item_id)
        for r in rows:
            self.ctrl_bank_unmatched_tree.insert(
                "",
                END,
                values=(
                    iso_to_ddmmyyyy(str(r["bank_date"] or "") or None),
                    format_minor_km(int(r["amount_minor"] or 0)),
                    r["description"] or "",
                ),
            )
        self.log(f"[CTRL] Neuparene uplate (banka) – učitano redova={len(rows)}")

    def _build_cashbook_tab(self, parent: Frame) -> None:
        controls = Frame(parent)
        controls.pack(fill="x", padx=8, pady=8)

        today = date.today()
        self.cash_year_var = IntVar(value=today.year)
        self.cash_month_var = IntVar(value=today.month)
        self.cash_status_var = StringVar(value="")
        self.cash_opening_var = StringVar(value="")
        self.cash_totals_var = StringVar(value="")

        ttk.Label(controls, text="Godina:").pack(side=LEFT)
        self.cash_year_cb = ttk.Combobox(
            controls,
            width=8,
            values=[str(today.year - 1), str(today.year), str(today.year + 1)],
            state="normal",
        )
        self.cash_year_cb.set(str(today.year))
        self.cash_year_cb.pack(side=LEFT, padx=(4, 12))

        ttk.Label(controls, text="Mjesec:").pack(side=LEFT)
        self.cash_month_cb = ttk.Combobox(
            controls, width=6, values=[str(i) for i in range(1, 13)], state="readonly"
        )
        self.cash_month_cb.set(str(today.month))
        self.cash_month_cb.pack(side=LEFT, padx=(4, 12))

        ttk.Label(controls, textvariable=self.cash_status_var).pack(side=LEFT, padx=(0, 12))

        ttk.Label(controls, text="Doneseni saldo (KM):").pack(side=LEFT)
        self.cash_opening_entry = ttk.Entry(controls, width=14, textvariable=self.cash_opening_var)
        self.cash_opening_entry.pack(side=LEFT, padx=(4, 6))
        self.cash_opening_save_btn = Button(controls, text="Sačuvaj", command=self.save_cashbook_opening)
        self.cash_opening_save_btn.pack(side=LEFT, padx=(0, 12))

        self.cash_recalc_btn = Button(controls, text="Preračunaj", command=self.recalculate_cashbook)
        self.cash_recalc_btn.pack(side=LEFT, padx=6)
        self.cash_close_btn = Button(controls, text="Zatvori mjesec", command=self.close_cashbook)
        self.cash_close_btn.pack(side=LEFT, padx=6)
        self.cash_unlock_btn = Button(controls, text="Otključaj", command=self.unlock_cashbook)
        self.cash_unlock_btn.pack(side=LEFT, padx=6)
        Button(controls, text="Izvoz dnevnika (XLSX)", command=self.export_cashbook_xlsx).pack(side=LEFT, padx=6)
        Button(controls, text="Osvježi", command=self.refresh_cashbook_view).pack(side=RIGHT, padx=6)

        columns = ("rb", "datum", "opis", "ulaz", "izlaz")
        self.cash_tree = ttk.Treeview(parent, columns=columns, show="headings", height=16)
        self.cash_tree.heading("rb", text="R.br.")
        self.cash_tree.heading("datum", text="Datum")
        self.cash_tree.heading("opis", text="Opis")
        self.cash_tree.heading("ulaz", text="Ulaz (KM)")
        self.cash_tree.heading("izlaz", text="Izlaz (KM)")
        self.cash_tree.column("rb", width=50, stretch=False, anchor="e")
        self.cash_tree.column("datum", width=100, stretch=False)
        self.cash_tree.column("opis", width=360, stretch=True)
        self.cash_tree.column("ulaz", width=120, stretch=False, anchor="e")
        self.cash_tree.column("izlaz", width=120, stretch=False, anchor="e")

        scroll = ttk.Scrollbar(parent, orient="vertical", command=self.cash_tree.yview)
        self.cash_tree.configure(yscrollcommand=scroll.set)
        self.cash_tree.pack(side=LEFT, fill=BOTH, expand=True, padx=(8, 0), pady=(0, 8))
        scroll.pack(side=RIGHT, fill="y", padx=(0, 8), pady=(0, 8))

        summary = Frame(parent)
        summary.pack(fill="x", padx=8, pady=(0, 8))
        ttk.Label(summary, textvariable=self.cash_totals_var, justify="left").pack(anchor="w")

        self.cash_year_cb.bind("<<ComboboxSelected>>", lambda _e: self.refresh_cashbook_view())
        self.cash_month_cb.bind("<<ComboboxSelected>>", lambda _e: self.refresh_cashbook_view())
        self.refresh_cashbook_view()

    def _cashbook_selected(self) -> tuple[int, int]:
        try:
            y = int(self.cash_year_cb.get())
            m = int(self.cash_month_cb.get())
        except Exception:
            y = int(date.today().year)
            m = int(date.today().month)
        return y, m

    def _cashbook_is_unlocked(self, year: int, month: int) -> bool:
        return (int(year), int(month)) in self.cashbook_unlocked

    def refresh_cashbook_view(self) -> None:
        year, month = self._cashbook_selected()
        try:
            month_row = self.db.get_or_create_cashbook_month(year, month)
            month_id = int(month_row["id"])
            entries = self.db.fetch_cashbook_entries(month_id)
            if not entries and month_row["status"] == "OPEN":
                month_row = self.db.rebuild_cashbook_entries(year, month)
                entries = self.db.fetch_cashbook_entries(int(month_row["id"]))
        except Exception as exc:
            self.log(f"[ERROR] cashbook refresh: {exc}")
            return

        is_unlocked = self._cashbook_is_unlocked(year, month)
        status = str(month_row["status"])
        closed_at = month_row["closed_at"] or ""
        status_label = "Otvoren" if status == "OPEN" else ("Zatvoren" if status == "CLOSED" else status)
        status_text = f"Status: {status_label}"
        if status == "CLOSED":
            status_text += f" (zatvoreno: {closed_at})"
        if is_unlocked:
            status_text += " [OTKLJUČANO]"
        self.cash_status_var.set(status_text)

        self.cash_opening_var.set(format_minor_km(int(month_row["opening_balance_minor"])))

        for item_id in self.cash_tree.get_children():
            self.cash_tree.delete(item_id)
        for r in entries:
            self.cash_tree.insert(
                "",
                END,
                values=(
                    int(r["rb"]),
                    r["entry_date"],
                    r["description"],
                    format_minor_km(int(r["in_minor"])),
                    format_minor_km(int(r["out_minor"])),
                ),
            )

        opening = int(month_row["opening_balance_minor"])
        total_in = int(month_row["total_in_minor"])
        total_out = int(month_row["total_out_minor"])
        closing = int(month_row["closing_balance_minor"])
        ukupni_promet = opening + total_in

        self.cash_totals_var.set(
            "\n".join(
                [
                    f"Promet blagajne: ULaz {format_minor_km(total_in)} / IZlaz {format_minor_km(total_out)}",
                    f"Saldo od (doneseni saldo): {format_minor_km(opening)}",
                    f"Ukupni promet (doneseni saldo + ulaz): {format_minor_km(ukupni_promet)}",
                    f"Odbija se izdatak: {format_minor_km(total_out)}",
                    f"Saldo (zaključno stanje): {format_minor_km(closing)}",
                ]
            )
        )

        can_edit = status == "OPEN" or is_unlocked
        self.cash_opening_entry.configure(state=("normal" if can_edit else "disabled"))
        self.cash_opening_save_btn.configure(state=("normal" if can_edit else "disabled"))
        self.cash_recalc_btn.configure(state=("normal" if can_edit else "disabled"))
        self.cash_close_btn.configure(state=("normal" if status == "OPEN" else "disabled"))
        self.cash_unlock_btn.configure(state=("normal" if status == "CLOSED" and not is_unlocked else "disabled"))

    def save_cashbook_opening(self) -> None:
        year, month = self._cashbook_selected()
        try:
            month_row = self.db.get_or_create_cashbook_month(year, month)
        except Exception as exc:
            self.log(f"[ERROR] cashbook opening load: {exc}")
            return

        status = str(month_row["status"])
        is_unlocked = self._cashbook_is_unlocked(year, month)
        if status == "CLOSED" and not is_unlocked:
            messagebox.showwarning(APP_NAME, "Mjesec je zatvoren. Potrebno je otključavanje.")
            return

        reason = None
        if status == "CLOSED" and is_unlocked:
            reason = simpledialog.askstring(APP_NAME, "Razlog izmjene doneseni saldo (zatvoren mjesec):")
            if not reason:
                return

        try:
            opening_minor = parse_km_text_to_minor(self.cash_opening_var.get())
        except Exception:
            messagebox.showerror(APP_NAME, "Neispravan iznos. Format: KM (npr 10,50)")
            return

        try:
            updated = self.db.set_cashbook_opening(int(month_row["id"]), opening_minor)
            if status == "CLOSED":
                self.db.audit(
                    "EDIT_OPENING",
                    f"cashbook {year:04d}-{month:02d}",
                    f"opening_balance_minor={opening_minor}; reason={reason}",
                )
            self.log(f"[CASH] Doneseni saldo sačuvan {year:04d}-{month:02d}: {opening_minor}")
            self.refresh_cashbook_view()
        except Exception as exc:
            self.log(f"[ERROR] cashbook opening save: {exc}")

    def recalculate_cashbook(self) -> None:
        year, month = self._cashbook_selected()
        try:
            month_row = self.db.get_or_create_cashbook_month(year, month)
        except Exception as exc:
            self.log(f"[ERROR] cashbook recalc load: {exc}")
            return

        status = str(month_row["status"])
        is_unlocked = self._cashbook_is_unlocked(year, month)
        if status == "CLOSED" and not is_unlocked:
            messagebox.showwarning(APP_NAME, "Mjesec je zatvoren. Potrebno je otključavanje.")
            return

        reason = None
        allow_closed = False
        if status == "CLOSED" and is_unlocked:
            allow_closed = True
            reason = simpledialog.askstring(APP_NAME, "Razlog preračuna (zatvoren mjesec):")
            if not reason:
                return

        try:
            self.db.rebuild_cashbook_entries(year, month, allow_closed=allow_closed)
            if status == "CLOSED":
                self.db.audit(
                    "RECALC",
                    f"cashbook {year:04d}-{month:02d}",
                    f"reason={reason}",
                )
            self.log(f"[CASH] Preračunato {year:04d}-{month:02d}")
            self.refresh_cashbook_view()
        except Exception as exc:
            self.log(f"[ERROR] cashbook recalc: {exc}")

    def close_cashbook(self) -> None:
        year, month = self._cashbook_selected()
        try:
            month_row = self.db.get_or_create_cashbook_month(year, month)
        except Exception as exc:
            self.log(f"[ERROR] cashbook close load: {exc}")
            return

        if str(month_row["status"]) != "OPEN":
            messagebox.showwarning(APP_NAME, "Mjesec nije otvoren.")
            return

        if not messagebox.askyesno(APP_NAME, f"Zatvoriti mjesec {year:04d}-{month:02d}?"):
            return

        try:
            self.db.rebuild_cashbook_entries(year, month)
            closed_month, next_month = self.db.close_cashbook_month(year, month)
            expected_next_opening = int(closed_month["closing_balance_minor"])
            actual_next_opening = int(next_month["opening_balance_minor"] or 0)
            self.db.audit(
                "CLOSE",
                f"cashbook {year:04d}-{month:02d}",
                f"closing_balance_minor={expected_next_opening}; next_opening_minor={actual_next_opening}",
            )
            self.log(
                f"[CASH] Zatvoren mjesec {year:04d}-{month:02d}. Zaključno={int(closed_month['closing_balance_minor'])} | Sljedeći={int(next_month['year']):04d}-{int(next_month['month']):02d}"
            )
            if actual_next_opening != expected_next_opening:
                self.log(
                    f"[CASH][WARN] Sljedeći mjesec već postoji i nije automatski usklađen (status={str(next_month['status'])}). Očekivani opening={expected_next_opening}, a u bazi je {actual_next_opening}."
                )
            self.refresh_cashbook_view()
        except Exception as exc:
            self.log(f"[ERROR] cashbook close: {exc}")

    def unlock_cashbook(self) -> None:
        year, month = self._cashbook_selected()
        try:
            month_row = self.db.get_or_create_cashbook_month(year, month)
        except Exception as exc:
            self.log(f"[ERROR] cashbook unlock load: {exc}")
            return

        if str(month_row["status"]) != "CLOSED":
            messagebox.showinfo(APP_NAME, "Otključavanje je potrebno samo za zatvoren mjesec.")
            return

        pwd = simpledialog.askstring(APP_NAME, "Lozinka:", show="*")
        if pwd is None:
            return
        # Same global password as DB reset (stored in settings).
        if pwd != self._get_reset_password():
            messagebox.showerror(APP_NAME, "Pogrešna lozinka.")
            self.db.audit("UNLOCK_FAILED", f"cashbook {year:04d}-{month:02d}", "bad_password")
            return

        self.cashbook_unlocked.add((int(year), int(month)))
        self.db.audit("UNLOCK", f"cashbook {year:04d}-{month:02d}", "ok")
        self.log(f"[CASH] Otključano {year:04d}-{month:02d}")
        self.refresh_cashbook_view()

    def export_cashbook_xlsx(self) -> None:
        year, month = self._cashbook_selected()
        default_name = f"Dnevnik_blagajne_{year:04d}-{month:02d}.xlsx"
        filename = filedialog.asksaveasfilename(
            title="Izvoz dnevnika blagajne (XLSX)",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel fajlovi", "*.xlsx"), ("Svi fajlovi", "*.*")],
        )
        if not filename:
            return

        out_path = Path(filename)

        def work():
            export_cashbook_dnevnik_xlsx(db=self.db, year=year, month=month, out_path=out_path, log=self.log)

        self._run_in_thread(work)


def main() -> None:
    global _APP_LOCK_FH

    for folder in (DATA_DIR, LOGS_DIR):
        folder.mkdir(parents=True, exist_ok=True)

    _APP_LOCK_FH = _try_acquire_single_instance_lock()
    if _APP_LOCK_FH is None:
        _show_already_running_message()
        return

    settings = SettingsStore(SETTINGS_PATH)
    configured_db = settings.get("db_path", None)
    db_path = DB_PATH
    if configured_db:
        try:
            db_path = Path(str(configured_db))
        except Exception:
            db_path = DB_PATH

    root = Tk()
    apply_window_icon(root)
    apply_ui_scale(root)
    root.update_idletasks()
    root.geometry("1200x700")
    try:
        # Center main window on screen
        sw = root.winfo_screenwidth()
        sh = root.winfo_screenheight()
        w, h = 1200, 700
        root.geometry(f"{w}x{h}+{int((sw - w) / 2)}+{int((sh - h) / 2)}")
    except Exception:
        pass

    app = App(root, db_path)

    root.protocol("WM_DELETE_WINDOW", app.request_exit)
    root.mainloop()


def _try_acquire_single_instance_lock():
    try:
        import msvcrt
    except Exception:
        return object()

    try:
        fh = open(LOCK_PATH, "a+b")
        fh.seek(0)
        try:
            msvcrt.locking(fh.fileno(), msvcrt.LK_NBLCK, 1)
        except OSError:
            try:
                fh.close()
            except Exception:
                pass
            return None
        return fh
    except Exception:
        return object()


def _show_already_running_message() -> None:
    try:
        root = Tk()
        apply_window_icon(root)
        root.withdraw()
        messagebox.showerror(APP_NAME, "Aplikacija je već pokrenuta.")
        root.destroy()
    except Exception:
        print("Aplikacija je već pokrenuta.", file=sys.stderr)


if __name__ == "__main__":
    main()
